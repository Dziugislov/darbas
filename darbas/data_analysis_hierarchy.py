import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import matplotlib.lines as mlines
from matplotlib.patches import Circle
from sklearn.cluster import AgglomerativeClustering
from sklearn.preprocessing import StandardScaler
from datetime import datetime
import scipy.cluster.hierarchy as shc
from scipy.spatial.distance import pdist, squareform
import calendar
import os
import glob
import read_ts
import matplotlib.pyplot as plt
import numpy as np
from scipy.spatial.distance import cdist
import matplotlib.pyplot as plt
import scipy.cluster.hierarchy as shc
from scipy.spatial.distance import pdist
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import numpy as np
from matplotlib.lines import Line2D
import matplotlib
import openpyxl

from input import MIN_TRADES, MAX_TRADES, MIN_ELEMENTS_PER_CLUSTER, DEFAULT_NUM_CLUSTERS
from input import TICKER, START_DATE, END_DATE, TRAIN_TEST_SPLIT, ATR_PERIOD, TRADING_CAPITAL
from SMA_Strategy import SMAStrategy

import json

def main():
    def load_parameters():
        try:
            with open("parameters.json", "r") as file:
                parameters = json.load(file)
                big_point_value = parameters["big_point_value"]
                slippage = parameters["slippage"]
                capital = parameters.get("capital", TRADING_CAPITAL)
                atr_period = parameters.get("atr_period", ATR_PERIOD)
                return big_point_value, slippage, capital, atr_period
        except FileNotFoundError:
            print("Parameters file not found. Ensure it was saved correctly in data_gather.py.")
            return None, None, TRADING_CAPITAL, ATR_PERIOD
        except KeyError as e:
            print(f"Missing key in parameters.json: {e}")
            print(f"Available keys: {', '.join(parameters.keys())}")
            return None, None, TRADING_CAPITAL, ATR_PERIOD

    # Load the parameters
    big_point_value, slippage, capital, atr_period = load_parameters()

    print(f"Big Point Value: {big_point_value}")
    print(f"Dynamic Slippage: {slippage}")
    print(f"Capital for Position Sizing: {capital:,}")
    print(f"ATR Period: {atr_period}")

    # Setup paths
    WORKING_DIR = "."  # Current directory
    DATA_DIR = os.path.join(WORKING_DIR, "data")
    SYMBOL = TICKER.replace('=F', '')

    # Define the output folder where the plots will be saved
    output_dir = os.path.join(WORKING_DIR, 'output', SYMBOL)
    os.makedirs(output_dir, exist_ok=True)  # Create the directory if it doesn't exist

    # Function to save plots in the created folder
    def save_plot(plot_name):
        plt.savefig(os.path.join(output_dir, plot_name))  # Save plot to the symbol-specific folder
        plt.close()  # Close the plot to free up memory

    # Function to save results to csv in the same output directory
    def save_results(data_frame, file_name):
        csv_path = os.path.join(output_dir, file_name)
        data_frame.to_csv(csv_path, index=False)
        print(f"Results saved to {csv_path}")

    def find_futures_file(symbol, data_dir):
        """Find a data file for the specified futures symbol"""
        # First try a pattern that specifically looks for @SYMBOL
        pattern = f"*@{symbol}_*.dat"
        files = glob.glob(os.path.join(data_dir, pattern))
        
        if not files:
            # Try a pattern that looks for the symbol with an underscore or at boundary
            pattern = f"*_@{symbol}_*.dat"
            files = glob.glob(os.path.join(data_dir, pattern))
        
        if not files:
            # Try a more specific boundary pattern for the symbol
            pattern = f"*_{symbol}_*.dat"
            files = glob.glob(os.path.join(data_dir, pattern))
        
        if not files:
            # Last resort: less specific but better than nothing
            pattern = f"*@{symbol}*.dat"
            files = glob.glob(os.path.join(data_dir, pattern))
        
        return files[0] if files else None

    # Load the SMA simulation results
    def analyze_sma_results(file_path='sma_all_results.txt'):
        print(f"Loading simulation results from {file_path}...")

        # Load the data from the CSV file
        data = pd.read_csv(file_path)

        # Print basic information about the data
        print(f"Loaded {len(data)} simulation results")
        print(f"Short SMA range: {data['short_SMA'].min()} to {data['short_SMA'].max()}")
        print(f"Long SMA range: {data['long_SMA'].min()} to {data['long_SMA'].max()}")
        print(f"Sharpe ratio range: {data['sharpe_ratio'].min():.4f} to {data['sharpe_ratio'].max():.4f}")

        # Find the best Sharpe ratio
        best_idx = data['sharpe_ratio'].idxmax()
        best_short_sma = data.loc[best_idx, 'short_SMA']
        best_long_sma = data.loc[best_idx, 'long_SMA']
        best_sharpe = data.loc[best_idx, 'sharpe_ratio']
        best_trades = data.loc[best_idx, 'trades']

        print(f"\nBest parameters:")
        print(f"Short SMA: {best_short_sma}")
        print(f"Long SMA: {best_long_sma}")
        print(f"Sharpe Ratio: {best_sharpe:.6f}")
        print(f"Number of Trades: {best_trades}")

        # Create a pivot table for the heatmap
        heatmap_data = data.pivot_table(
            index='long_SMA',
            columns='short_SMA',
            values='sharpe_ratio'
        )

        # Create the heatmap visualization
        plt.figure(figsize=(12, 10))

        # Create a mask for invalid combinations (where short_SMA >= long_SMA)
        mask = np.zeros_like(heatmap_data, dtype=bool)
        for i, long_sma in enumerate(heatmap_data.index):
            for j, short_sma in enumerate(heatmap_data.columns):
                if short_sma >= long_sma:
                    mask[i, j] = True

        # Plot the heatmap with the mask
        ax = sns.heatmap(
            heatmap_data,
            mask=mask,
            cmap='coolwarm',  # Blue to red colormap
            annot=False,  # Don't annotate each cell with its value
            fmt='.4f',
            linewidths=0,
            cbar_kws={'label': 'Sharpe Ratio'}
        )

        # Invert the y-axis so smaller long_SMA values are at the top
        ax.invert_yaxis()

        # Find the position of the best Sharpe ratio in the heatmap
        best_y = heatmap_data.index.get_loc(best_long_sma)
        best_x = heatmap_data.columns.get_loc(best_short_sma)

        # Add a star to mark the best Sharpe ratio
        # We need to add 0.5 to center the marker in the cell
        ax.add_patch(Circle((best_x + 0.5, best_y + 0.5), 0.4, facecolor='none',
                            edgecolor='white', lw=2))
        plt.plot(best_x + 0.5, best_y + 0.5, 'w*', markersize=10)

        # Set labels and title
        plt.title(f'SMA Optimization Heatmap (Best Sharpe: {best_sharpe:.4f} at {best_short_sma}/{best_long_sma})',
                fontsize=14)
        plt.xlabel('Short SMA (days)', fontsize=12)
        plt.ylabel('Long SMA (days)', fontsize=12)

        # Rotate tick labels for better readability
        plt.xticks(rotation=45, ha='right')
        plt.yticks(rotation=0)

        # Add a text annotation for the best parameters
        plt.annotate(
            f'Best: Short={best_short_sma}, Long={best_long_sma}\nSharpe={best_sharpe:.4f}, Trades={best_trades}',
            xy=(best_x + 0.5, best_y + 0.5),
            xytext=(best_x + 5, best_y + 5),
            arrowprops=dict(arrowstyle="->", color='white'),
            color='white',
            backgroundcolor='black',
            bbox=dict(boxstyle="round,pad=0.3", fc="black", alpha=0.7)
        )

        # Display and save the plot
        plt.tight_layout()

        # Return the data and best parameters
        return data, best_short_sma, best_long_sma, best_sharpe, best_trades

    def compute_hierarchical_medoids(X, labels, valid_clusters):
        """
        Compute medoids for each valid cluster from hierarchical clustering
        A medoid is the data point in a cluster that has the minimum average distance to all other points in the cluster

        Parameters:
        X: numpy array of shape (n_samples, n_features) - Original data points (not scaled)
        labels: numpy array of shape (n_samples,) - Cluster labels for each data point
        valid_clusters: set - Set of valid cluster IDs

        Returns:
        list of tuples - Each tuple contains (short_SMA, long_SMA, sharpe_ratio, trades) for each medoid
        """
        # Initialize list to store medoids for each cluster
        medoids = []

        # Process each valid cluster
        for cluster_id in valid_clusters:
            # Extract points belonging to this cluster
            cluster_indices = np.where(labels == cluster_id)[0]
            cluster_points = X[cluster_indices]

            if len(cluster_points) == 0:
                continue

            if len(cluster_points) == 1:
                # If only one point in cluster, it's the medoid
                medoids.append(tuple(cluster_points[0]))
                continue

            # Calculate pairwise distances within cluster
            min_total_distance = float('inf')
            medoid = None

            for i, point1 in enumerate(cluster_points):
                total_distance = 0
                for j, point2 in enumerate(cluster_points):
                    # Calculate Euclidean distance between points
                    distance = np.sqrt(np.sum((point1 - point2) ** 2))
                    total_distance += distance

                if total_distance < min_total_distance:
                    min_total_distance = total_distance
                    medoid = point1

            if medoid is not None:
                medoids.append(tuple(medoid))

            # Print medoid info
            print(f"Cluster {cluster_id} medoid: Short SMA={int(medoid[0])}, Long SMA={int(medoid[1])}, "
                f"Sharpe={medoid[2]:.4f}, Trades={int(medoid[3])}")

        return medoids

    def create_dendrogram(X_scaled, method='ward', figsize=(12, 8), color_threshold=None, truncate_mode=None,
                        p=10, save_path=None):
        """
        Create and display a dendrogram for hierarchical clustering

        Parameters:
        X_scaled: numpy array - Scaled data points
        method: str - Linkage method ('ward', 'complete', 'average', 'single')
        figsize: tuple - Figure size
        color_threshold: float - The threshold to apply for coloring the dendrogram
        truncate_mode: str - 'lastp' for last p branches, 'level' for no more than p levels
        p: int - Used with truncate_mode
        save_path: str - Path to save the dendrogram image

        Returns:
        Z: numpy array - The hierarchical clustering linkage matrix
        """
        # Use a sample if there are too many points
        if len(X_scaled) > 1000:
            np.random.seed(42)  # For reproducibility
            sample_indices = np.random.choice(len(X_scaled), size=1000, replace=False)
            X_sample = X_scaled[sample_indices]
            print(f"Using a random sample of 1000 points for dendrogram creation (out of {len(X_scaled)} total points)")
        else:
            X_sample = X_scaled

        # Compute the distance matrix
        dist_matrix = pdist(X_sample, metric='euclidean')

        # Compute the linkage matrix
        Z = shc.linkage(dist_matrix, method=method)

        # Print some statistics about the linkage
        print(f"\nDendrogram using {method} linkage method:")
        print(f"Number of data points: {len(X_sample)}")
        print(f"Cophenetic correlation: {shc.cophenet(Z, dist_matrix)[0]:.4f}")

        # Create a new figure
        plt.figure(figsize=figsize)

        # Set the background color
        plt.gca().set_facecolor('white')

        # Set a title
        plt.title(f'Hierarchical Clustering Dendrogram ({method} linkage)', fontsize=14)

        # Generate the dendrogram
        dendrogram = shc.dendrogram(
            Z,
            truncate_mode='lastp',  # Only show the last p merges
            p=30,  # Show only 30 merges
            leaf_rotation=90.,
            leaf_font_size=10.,
            show_contracted=True,
            color_threshold=color_threshold
        )

        # Add labels and axes
        plt.xlabel('Sample Index or Cluster Size', fontsize=12)
        plt.ylabel('Distance', fontsize=12)

        # Add a horizontal line to indicate a distance threshold if specified
        if color_threshold is not None:
            plt.axhline(y=color_threshold, color='crimson', linestyle='--',
                        label=f'Threshold: {color_threshold:.2f}')
            plt.legend()

        # Adjust layout
        plt.tight_layout()

        # Save the plot
        save_plot('Hierarchical_Dendrogram.png')

        return Z

    def hierarchical_cluster_analysis(file_path='sma_all_results.txt', min_trades=MIN_TRADES, max_trades=MAX_TRADES,
                        min_elements_per_cluster=MIN_ELEMENTS_PER_CLUSTER):
        """
        Perform hierarchical clustering analysis on SMA optimization results to find robust parameter regions
        
        Parameters:
        file_path: str - Path to the file with SMA results
        min_trades: int - Minimum number of trades to consider
        max_trades: int - Maximum number of trades to consider
        min_elements_per_cluster: int - Minimum number of elements per cluster
        
        Returns:
        tuple - (X_filtered_full, medoids, top_medoids, max_sharpe_point, labels)
        """
        print(f"\n----- HIERARCHICAL CLUSTER ANALYSIS -----")
        print(f"Loading data from {file_path}...")

        # Load the data
        df = pd.read_csv(file_path)

        # Convert data to numpy array for easier processing
        X_full = df[['short_SMA', 'long_SMA', 'sharpe_ratio', 'trades']].values

        # STEP 1: First filter by trades and valid SMA combinations
        X_filtered_full = X_full[(X_full[:, 0] < X_full[:, 1]) &  # short_SMA < long_SMA
                    (X_full[:, 3] >= min_trades) &  # trades >= min_trades
                    (X_full[:, 3] <= max_trades)]  # trades <= max_trades

        if len(X_filtered_full) == 0:
            print(f"No data points meet the criteria after filtering! Adjust min_trades ({min_trades}) and max_trades ({max_trades}).")
            return None, None, None, None, None

        # STEP 2: Use exactly 3 dimensions for clustering - same as K-means
        X_filtered = X_filtered_full[:, 0:3]  # Only short_SMA, long_SMA, and sharpe_ratio
        
        print(f"Filtered data to {len(X_filtered)} points with {min_trades}-{max_trades} trades")

        # Extract the fields for better scaling visibility
        short_sma_values = X_filtered[:, 0]
        long_sma_values = X_filtered[:, 1]
        sharpe_values = X_filtered[:, 2]
        trades_values = X_filtered_full[:, 3]

        print(f"Short SMA range: {short_sma_values.min()} to {short_sma_values.max()}")
        print(f"Long SMA range: {long_sma_values.min()} to {long_sma_values.max()}")
        print(f"Sharpe ratio range: {sharpe_values.min():.4f} to {sharpe_values.max():.4f}")
        print(f"Trades range: {trades_values.min()} to {trades_values.max()}")

        # Scale the data for clustering - using StandardScaler for each dimension
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X_filtered)  # Only scale the 3 dimensions we use for clustering

        # Print scaling info for verification
        print("\nScaled data information:")
        scaled_short = X_scaled[:, 0]
        scaled_long = X_scaled[:, 1]
        scaled_sharpe = X_scaled[:, 2]

        print(f"Scaled Short SMA range: {scaled_short.min():.4f} to {scaled_short.max():.4f}")
        print(f"Scaled Long SMA range: {scaled_long.min():.4f} to {scaled_long.max():.4f}")
        print(f"Scaled Sharpe ratio range: {scaled_sharpe.min():.4f} to {scaled_sharpe.max():.4f}")

        # Create a dendrogram to help visualize (optional, can be commented out)
        print("\nCreating dendrogram to visualize hierarchical structure...")
        linkage_method = 'ward'  # Ward minimizes the variance within clusters
        Z = create_dendrogram(X_scaled, method=linkage_method, figsize=(12, 8))

        # STEP 3: Determine number of clusters - use the same DEFAULT_NUM_CLUSTERS as K-means
        print(f"Using default number of clusters: {DEFAULT_NUM_CLUSTERS}")
        k = DEFAULT_NUM_CLUSTERS

        # STEP 4: Apply hierarchical clustering
        print(f"Performing hierarchical clustering with {k} clusters using {linkage_method} linkage...")
        hierarchical = AgglomerativeClustering(
            n_clusters=k,
            linkage=linkage_method
        )

        # Fit the model and get cluster labels
        labels = hierarchical.fit_predict(X_scaled)

        # STEP 5: Count elements per cluster
        unique_labels, counts = np.unique(labels, return_counts=True)
        cluster_sizes = dict(zip(unique_labels, counts))

        print("\nCluster sizes:")
        for cluster_id, size in cluster_sizes.items():
            print(f"Cluster {cluster_id}: {size} elements")

        # STEP 6: Filter clusters with enough elements - SAME as K-means approach
        valid_clusters = {i for i, count in cluster_sizes.items() if count >= min_elements_per_cluster}
        
        if not valid_clusters:
            print(f"No clusters have at least {min_elements_per_cluster} elements! Using all clusters.")
            valid_clusters = set(unique_labels)
        
        # Don't filter the data points here yet - just keep track of which clusters are valid

        # STEP 7: Compute medoids using the original data (not scaled)
        print("Computing medoids for each cluster...")
        medoids = compute_hierarchical_medoids(X_filtered_full, labels, valid_clusters)

        # Find max Sharpe ratio point overall
        max_sharpe_idx = np.argmax(df['sharpe_ratio'].values)
        max_sharpe_point = df.iloc[max_sharpe_idx][['short_SMA', 'long_SMA', 'sharpe_ratio', 'trades']].values

        # STEP 8: Sort medoids by Sharpe ratio
        medoids_sorted = sorted(medoids, key=lambda x: float(x[2]), reverse=True)
        top_medoids = medoids_sorted[:5]  # Get top 5 medoids by Sharpe ratio

        # Print results
        print("\n----- HIERARCHICAL CLUSTERING RESULTS -----")
        print(f"Max Sharpe point: Short SMA={int(max_sharpe_point[0])}, Long SMA={int(max_sharpe_point[1])}, "
            f"Sharpe={max_sharpe_point[2]:.4f}, Trades={int(max_sharpe_point[3])}")

        print("\nTop 5 Medoids (by Sharpe ratio):")
        for idx, medoid in enumerate(top_medoids, 1):
            print(f"Top {idx}: Short SMA={int(medoid[0])}, Long SMA={int(medoid[1])}, "
                f"Sharpe={float(medoid[2]):.4f}, Trades={int(medoid[3])}")

        # Create visualization - pass all the original filtered data and valid_clusters
        create_hierarchical_cluster_visualization(X_filtered_full, medoids, top_medoids, max_sharpe_point, labels, valid_clusters)
        
        return X_filtered_full, medoids, top_medoids, max_sharpe_point, labels

    def plot_strategy_performance(short_sma, long_sma, top_medoids=None, big_point_value=1, slippage=0, capital=1000000, atr_period=14):
        print(f"\n----- PLOTTING STRATEGY PERFORMANCE -----")
        print(f"Using Short SMA: {short_sma}, Long SMA: {long_sma}")
        print(f"Trading with ATR-based position sizing (Capital: ${capital:,}, ATR Period: {atr_period})")
        if top_medoids:
            print(f"Including top {len(top_medoids)} medoids")

        # Load data from local file
        print(f"Loading {TICKER} data from local files...")
        data_file = find_futures_file(SYMBOL, DATA_DIR)
        if not data_file:
            print(f"Error: No data file found for {TICKER} in {DATA_DIR}")
            exit(1)

        print(f"Found data file: {os.path.basename(data_file)}")
        print(f"File size: {os.path.getsize(data_file)} bytes")

        # Load the data from local file
        all_data = read_ts.read_ts_ohlcv_dat(data_file)
        data_obj = all_data[0]
        ohlc_data = data_obj.data.copy()

        # Convert the OHLCV data to the format expected by the strategy
        data = ohlc_data.rename(columns={
            'datetime': 'Date',
            'open': 'Open',
            'high': 'High',
            'low': 'Low',
            'close': 'Close',
            'volume': 'Volume'
        })
        data.set_index('Date', inplace=True)

        # Add a warm-up period before the start date
        original_start_idx = None
        if START_DATE and END_DATE:
            # Calculate warm-up period for SMA calculation (longest SMA + buffer)
            warm_up_days = max(short_sma, long_sma) * 3  # Use 3x the longest SMA as warm-up
            
            # Convert dates to datetime
            start_date = pd.to_datetime(START_DATE)
            end_date = pd.to_datetime(END_DATE)
            
            # Adjust start date for warm-up
            adjusted_start = start_date - pd.Timedelta(days=warm_up_days)
            
            # Filter with the extended date range
            extended_data = data[(data.index >= adjusted_start) & (data.index <= end_date)]
            
            # Store the index where the actual analysis should start
            if not extended_data.empty:
                # Find the closest index to our original start date
                original_start_idx = extended_data.index.get_indexer([start_date], method='nearest')[0]
            
            print(f"Added {warm_up_days} days warm-up period before {START_DATE}")
            data = extended_data
        
        # Create a dictionary to store results for each strategy
        strategies = {
            'Best': {'short_sma': short_sma, 'long_sma': long_sma}
        }

        # Add medoids in their original order, including original Sharpe and Trades
        if top_medoids:
            print("\nUSING THESE MEDOIDS (IN ORIGINAL ORDER):")
            for i, medoid in enumerate(top_medoids, 1):
                strategies[f'Medoid {i}'] = {
                    'short_sma': int(medoid[0]),
                    'long_sma': int(medoid[1]),
                    'original_sharpe': float(medoid[2]),  # Store the original Sharpe ratio
                    'original_trades': int(medoid[3])     # Store the original number of trades
                }
                print(f"Medoid {i}: SMA({int(medoid[0])}/{int(medoid[1])}) - Original Sharpe: {float(medoid[2]):.4f}, Trades: {int(medoid[3])}")

        # Apply the proper strategy for each parameter set
        for name, params in strategies.items():
            # Calculate centered SMAs directly
            data[f'SMA_Short_{name}'] = data['Close'].rolling(window=params['short_sma'], center=True).mean()
            data[f'SMA_Long_{name}'] = data['Close'].rolling(window=params['long_sma'], center=True).mean()
            
            # Create a strategy instance for each parameter set
            sma_strategy = SMAStrategy(
                short_sma=params['short_sma'],
                long_sma=params['long_sma'],
                big_point_value=big_point_value,
                slippage=slippage,
                capital=capital,
                atr_period=atr_period
            )

            # Apply the strategy
            data = sma_strategy.apply_strategy(
                data.copy(),
                strategy_name=name
            )

        # Trim data to the original date range if we added warm-up period
        if original_start_idx is not None:
            data_for_evaluation = data.iloc[original_start_idx:]
            print(f"Trimmed warm-up period. Original data length: {len(data)}, Evaluation data length: {len(data_for_evaluation)}")
        else:
            data_for_evaluation = data
        
        # Calculate split index for in-sample/out-of-sample using the trimmed data
        split_index = int(len(data_for_evaluation) * TRAIN_TEST_SPLIT)
        split_date = data_for_evaluation.index[split_index]

        # Create color palette for strategies
        colors = ['blue', 'green', 'purple', 'orange', 'brown', 'pink']

        # Create the performance visualization with just two panels
        plt.figure(figsize=(14, 12))

        # Plot price and SMA (first subplot)
        plt.subplot(2, 1, 1)
        plt.plot(data_for_evaluation.index, data_for_evaluation['Close'], label=f'{SYMBOL} Price', color='black', alpha=0.5)
        
        # Use the centered SMAs for plotting
        for name, params in strategies.items():
            if name == 'Best':
                plt.plot(data_for_evaluation.index, data_for_evaluation[f'SMA_Short_{name}'], 
                        label=f'Short SMA ({params["short_sma"]})', color='orange')
                plt.plot(data_for_evaluation.index, data_for_evaluation[f'SMA_Long_{name}'], 
                        label=f'Long SMA ({params["long_sma"]})', color='blue')
        
        # Mark the train/test split
        plt.axvline(x=split_date, color='black', linestyle='--', alpha=0.7)
        
        plt.legend(loc='upper left')
        plt.title(f'{SYMBOL} Price and SMA Indicators')
        plt.ylabel('Price')
        plt.grid(True, alpha=0.3)

        # Plot cumulative P&L (second subplot)
        plt.subplot(2, 1, 2)

        for i, (name, params) in enumerate(strategies.items()):
            color = colors[i % len(colors)]

            # Plot full period P&L
            plt.plot(data_for_evaluation.index, data_for_evaluation[f'Cumulative_PnL_{name}'],
                    label=f'{name} ({params["short_sma"]}/{params["long_sma"]})', color=color)

            # Plot out-of-sample portion with thicker line
            plt.plot(data_for_evaluation.index[split_index:], data_for_evaluation[f'Cumulative_PnL_{name}'].iloc[split_index:],
                    color=color, linewidth=2.5, alpha=0.7)

        plt.axvline(x=split_date, color='black', linestyle='--',
                    label=f'Train/Test Split ({int(TRAIN_TEST_SPLIT * 100)}%/{int((1 - TRAIN_TEST_SPLIT) * 100)}%)')
        plt.axhline(y=0.0, color='gray', linestyle='-', alpha=0.5, label='Break-even')
        plt.legend(loc='upper left')
        plt.title('Strategy Cumulative P&L')
        plt.ylabel('P&L ($)')
        plt.grid(True, alpha=0.3)

        plt.tight_layout()
        save_plot('Hierarchical_Strategy_Performance.png')

        # Create a list to store performance data for saving to file
        performance_data = []

        # Print detailed performance metrics for all strategies with in-sample and out-of-sample breakdown
        print("\n----- PERFORMANCE SUMMARY -----")

        # IN-SAMPLE PERFORMANCE
        print("\nIN-SAMPLE PERFORMANCE:")
        header = f"{'Strategy':<10} | {'Short/Long':<10} | {'P&L':>12} | {'Sharpe':>7} | {'Trades':>6}"
        separator = "-" * len(header)
        print(separator)
        print(header)
        print(separator)

        # Calculate in-sample metrics using trimmed data
        for name, params in strategies.items():
            short = params['short_sma']
            long = params['long_sma']

            # Get in-sample data from trimmed dataset
            in_sample = data_for_evaluation.iloc[:split_index]

            # Calculate in-sample metrics
            in_sample_daily_pnl = in_sample[f'Daily_PnL_{name}']
            in_sample_cumulative_pnl = in_sample[f'Daily_PnL_{name}'].sum()
            
            # Calculate average position size for in-sample
            avg_pos_size = in_sample[f'Position_Size_{name}'].mean()

            # Use original Sharpe and Trades for medoids, calculate for Best
            if name.startswith('Medoid'):
                in_sample_sharpe = params['original_sharpe']
                in_sample_trades = params['original_trades']
            else:
                # Calculate Sharpe ratio (annualized) for Best
                if in_sample_daily_pnl.std() > 0:
                    in_sample_sharpe = in_sample_daily_pnl.mean() / in_sample_daily_pnl.std() * np.sqrt(252)
                else:
                    in_sample_sharpe = 0
                # Count in-sample trades for Best
                in_sample_trades = in_sample[f'Position_Change_{name}'].sum()
            
            # Print in the original order
            row = f"{name:<10} | {short:>4}/{long:<5} | ${in_sample_cumulative_pnl:>10,.2f} | {in_sample_sharpe:>6.3f} | {in_sample_trades:>6}"
            print(row)
            
            # Store data for later use
            performance_data.append({
                'Period': 'In-Sample',
                'Strategy': name,
                'Short_SMA': short,
                'Long_SMA': long,
                'PnL': in_sample_cumulative_pnl,
                'Avg_Position': avg_pos_size,
                'Sharpe': in_sample_sharpe,
                'Trades': in_sample_trades
            })

        print(separator)

        # OUT-OF-SAMPLE PERFORMANCE
        print("\nOUT-OF-SAMPLE PERFORMANCE:")
        print(separator)
        print(header)
        print(separator)

        # Calculate out-of-sample metrics using trimmed data
        for name, params in strategies.items():
            short = params['short_sma']
            long = params['long_sma']

            # Get out-of-sample data from trimmed dataset
            out_sample = data_for_evaluation.iloc[split_index:]

            # Calculate out-of-sample metrics
            out_sample_daily_pnl = out_sample[f'Daily_PnL_{name}']
            out_sample_cumulative_pnl = out_sample[f'Daily_PnL_{name}'].sum()
            
            # Calculate average position size for out-of-sample
            avg_pos_size = out_sample[f'Position_Size_{name}'].mean()

            # Calculate Sharpe ratio (annualized)
            if out_sample_daily_pnl.std() > 0:
                out_sample_sharpe = out_sample_daily_pnl.mean() / out_sample_daily_pnl.std() * np.sqrt(252)
            else:
                out_sample_sharpe = 0

            # Count out-of-sample trades
            out_sample_trades = out_sample[f'Position_Change_{name}'].sum()
            
            # Print in the original order
            row = f"{name:<10} | {short:>4}/{long:<5} | ${out_sample_cumulative_pnl:>10,.2f} | {out_sample_sharpe:>6.3f} | {out_sample_trades:>6}"
            print(row)
            
            # Store data for later use
            performance_data.append({
                'Period': 'Out-of-Sample',
                'Strategy': name,
                'Short_SMA': short,
                'Long_SMA': long,
                'PnL': out_sample_cumulative_pnl,
                'Avg_Position': avg_pos_size,
                'Sharpe': out_sample_sharpe,
                'Trades': out_sample_trades
            })

        print(separator)

        # FULL PERIOD PERFORMANCE
        print("\nFULL PERIOD PERFORMANCE:")
        print(separator)
        print(header)
        print(separator)

        # Calculate full period metrics using trimmed data
        for name, params in strategies.items():
            short = params['short_sma']
            long = params['long_sma']

            # Calculate full period metrics using trimmed data
            full_daily_pnl = data_for_evaluation[f'Daily_PnL_{name}']
            full_cumulative_pnl = full_daily_pnl.sum()
            
            # Calculate average position size for full period
            avg_pos_size = data_for_evaluation[f'Position_Size_{name}'].mean()
            max_pos_size = data_for_evaluation[f'Position_Size_{name}'].max()

            # Calculate Sharpe ratio (annualized)
            if full_daily_pnl.std() > 0:
                full_sharpe = full_daily_pnl.mean() / full_daily_pnl.std() * np.sqrt(252)
            else:
                full_sharpe = 0

            # Count full period trades
            full_trades = data_for_evaluation[f'Position_Change_{name}'].sum()
            
            # Print in the original order
            row = f"{name:<10} | {short:>4}/{long:<5} | ${full_cumulative_pnl:>10,.2f} | {full_sharpe:>6.3f} | {full_trades:>6}"
            print(row)
            
            # Store data for later use
            performance_data.append({
                'Period': 'Full',
                'Strategy': name,
                'Short_SMA': short,
                'Long_SMA': long,
                'PnL': full_cumulative_pnl,
                'Avg_Position': avg_pos_size,
                'Max_Position': max_pos_size,
                'Sharpe': full_sharpe,
                'Trades': full_trades
            })
            
            # Additional metrics for the best strategy
            if name == 'Best':
                print(f"\nAdditional metrics for Best strategy:")
                print(f"Maximum position size: {max_pos_size:.2f} contracts")
                print(f"Average ATR value: {data_for_evaluation['ATR_Best'].mean():.4f}")
                
                # Calculate drawdown using trimmed data
                peak = data_for_evaluation[f'Cumulative_PnL_{name}'].cummax()
                drawdown = data_for_evaluation[f'Cumulative_PnL_{name}'] - peak
                max_drawdown = drawdown.min()
                
                print(f"Maximum drawdown: ${max_drawdown:.2f}")
                
                # Calculate win rate using trimmed data
                daily_win_rate = (data_for_evaluation[f'Daily_PnL_{name}'] > 0).mean() * 100
                print(f"Daily win rate: {daily_win_rate:.2f}%")

        print(separator)

        # Save performance data to CSV in the ticker output directory
        performance_df = pd.DataFrame(performance_data)
        save_results(performance_df, f"{SYMBOL}_hierarchical_performance_summary.csv")

        return data_for_evaluation  # Return the trimmed data instead of full data

    def bimonthly_out_of_sample_comparison(data, best_short_sma, best_long_sma, top_medoids, min_sharpe=0.2, 
                                big_point_value=big_point_value, slippage=slippage,
                                capital=capital, atr_period=atr_period):
        """
        Compare bimonthly (2-month) performance between the best Sharpe strategy and a portfolio of top medoids
        using ATR-based position sizing.
        
        Parameters:
        data: DataFrame with market data (should already be trimmed to exclude warmup period)
        best_short_sma: int - The short SMA period for the best Sharpe strategy
        best_long_sma: int - The long SMA period for the best Sharpe strategy
        top_medoids: list - List of top medoids, each as (short_sma, long_sma, sharpe, trades)
        min_sharpe: float - Minimum Sharpe ratio threshold for medoids to be included
        big_point_value: float - Big point value for the futures contract
        slippage: float - Slippage value in price units
        capital: float - Capital allocation for position sizing
        atr_period: int - Period for ATR calculation
        """
        print(f"\n----- BIMONTHLY OUT-OF-SAMPLE COMPARISON -----")
        print(f"Best Sharpe: ({best_short_sma}/{best_long_sma})")
        print(f"Using ATR-based position sizing (Capital: ${capital:,}, ATR Period: {atr_period})")
        
        # Handle the case where top_medoids is None
        if top_medoids is None:
            print("No medoids provided. Comparison cannot be performed.")
            return None
        
        # ADDED: Ensure we have trimmed data
        if data is None:
            # Load data from local file
            print(f"Loading {TICKER} data from local files...")
            data_file = find_futures_file(SYMBOL, DATA_DIR)
            if not data_file:
                print(f"Error: No data file found for {TICKER} in {DATA_DIR}")
                exit(1)
            
            print(f"Found data file: {os.path.basename(data_file)}")
            print(f"File size: {os.path.getsize(data_file)} bytes")

            # Load the data from local file
            all_data = read_ts.read_ts_ohlcv_dat(data_file)
            data_obj = all_data[0]
            ohlc_data = data_obj.data.copy()

            # Convert the OHLCV data to the format expected by the strategy
            data = ohlc_data.rename(columns={
                'datetime': 'Date',
                'open': 'Open',
                'high': 'High',
                'low': 'Low',
                'close': 'Close',
                'volume': 'Volume'
            })
            data.set_index('Date', inplace=True)
            
            # Add a warm-up period before the start date
            original_start_idx = None
            if START_DATE and END_DATE:
                # Calculate warm-up period for SMA calculation (longest SMA + buffer)
                warm_up_days = max(best_short_sma, best_long_sma) * 3  # Use 3x the longest SMA as warm-up
                
                # Convert dates to datetime
                start_date = pd.to_datetime(START_DATE)
                end_date = pd.to_datetime(END_DATE)
                
                # Adjust start date for warm-up
                adjusted_start = start_date - pd.Timedelta(days=warm_up_days)
                
                # Filter with the extended date range
                extended_data = data[(data.index >= adjusted_start) & (data.index <= end_date)]
                
                # Store the index where the actual analysis should start
                if not extended_data.empty:
                    # Find the closest index to our original start date
                    original_start_idx = extended_data.index.get_indexer([start_date], method='nearest')[0]
                
                print(f"Added {warm_up_days} days warm-up period before {START_DATE}")
                data = extended_data
                
                # Trim data to the original date range
                if original_start_idx is not None:
                    data = data.iloc[original_start_idx:]
                    print(f"Trimmed warm-up period. Data length: {len(data)}")
        
        # The top_medoids are already sorted by Sharpe ratio in cluster_analysis
        # Just print them for verification
        print("\nUSING TOP MEDOIDS (BY SHARPE RATIO):")
        for i, medoid in enumerate(top_medoids, 1):
            print(f"Medoid {i}: ({int(medoid[0])}/{int(medoid[1])}) - Sharpe: {float(medoid[2]):.4f}, Trades: {int(medoid[3])}")
        
        # Take at most 3 medoids and filter by minimum Sharpe
        filtered_medoids = []
        for i, m in enumerate(top_medoids[:3]):
            # Check if we can access the required elements
            try:
                # Extract Sharpe ratio and check if it meets the threshold
                short_sma = m[0]
                long_sma = m[1]
                sharpe = float(m[2])  # Convert to float to handle numpy types
                trades = m[3]
                
                if sharpe >= min_sharpe:
                    filtered_medoids.append(m)
                    print(f"Selected medoid {i+1} with Sharpe {sharpe:.4f}")
            except (IndexError, TypeError) as e:
                print(f"Error processing medoid: {e}")
        
        if not filtered_medoids:
            print(f"No medoids have a Sharpe ratio >= {min_sharpe}. Comparison cannot be performed.")
            return None
        
        print(f"Creating portfolio of {len(filtered_medoids)} medoids with Sharpe ratio >= {min_sharpe}:")
        for i, medoid in enumerate(filtered_medoids, 1):
            print(f"Final Medoid {i}: ({int(medoid[0])}/{int(medoid[1])}) - Sharpe: {float(medoid[2]):.4f}")
        
        # Create strategies
        strategies = {
            'Best': {'short_sma': best_short_sma, 'long_sma': best_long_sma}
        }
        
        # Add filtered medoids
        for i, medoid in enumerate(filtered_medoids, 1):
            strategies[f'Medoid_{i}'] = {'short_sma': int(medoid[0]), 'long_sma': int(medoid[1])}
        
        # Apply each strategy to the data
        for name, params in strategies.items():
            strategy = SMAStrategy(
                short_sma=params['short_sma'],
                long_sma=params['long_sma'],
                big_point_value=big_point_value,
                slippage=slippage,
                capital=capital,
                atr_period=atr_period
            )
            
            # Apply the strategy
            data = strategy.apply_strategy(
                data.copy(),
                strategy_name=name
            )
        
        # Get the out-of-sample split date
        split_index = int(len(data) * TRAIN_TEST_SPLIT)
        split_date = data.index[split_index]
        print(f"Out-of-sample period starts on: {split_date.strftime('%Y-%m-%d')}")
        
        # Get out-of-sample data
        oos_data = data.iloc[split_index:].copy()
        
        # Add a year and bimonthly period columns for grouping (each year has 6 bimonthly periods)
        oos_data['year'] = oos_data.index.year.astype(int)
        oos_data['bimonthly'] = ((oos_data.index.month - 1) // 2 + 1).astype(int)
        
        # Create simplified period labels with just the start month (YYYY-MM)
        oos_data['period_label'] = oos_data.apply(
            lambda row: f"{int(row['year'])}-{int((row['bimonthly'] - 1) * 2 + 1):02d}",
            axis=1
        )
        
        # Create a DataFrame to store bimonthly Sharpe ratios
        bimonthly_sharpe = []
        
        # Group by year and bimonthly period, calculate Sharpe ratio for each period
        for period_label, group in oos_data.groupby('period_label'):
            # Skip periods with too few trading days
            if len(group) < 10:
                continue
                
            # Create a bimonthly result entry
            year, start_month = period_label.split('-')
            year = int(year)
            start_month = int(start_month)
            
            bimonthly_result = {
                'period_label': period_label,
                'date': pd.Timestamp(year=year, month=start_month, day=15),  # Middle of first month in period
                'trading_days': len(group),
            }
            
            # Calculate Sharpe ratio for each strategy in this period
            for name in strategies.keys():
                # Get returns for this strategy in this period
                returns = group[f'Daily_PnL_{name}']
                
                # Calculate Sharpe ratio (annualized)
                if len(returns) > 1 and returns.std() > 0:
                    sharpe = returns.mean() / returns.std() * np.sqrt(252)
                else:
                    sharpe = 0
                    
                bimonthly_result[f'{name}_sharpe'] = sharpe
                bimonthly_result[f'{name}_return'] = returns.sum()  # Total P&L for the period
            
            # Calculate the normalized average of medoid Sharpe ratios
            medoid_sharpes = [bimonthly_result[f'Medoid_{i}_sharpe'] for i in range(1, len(filtered_medoids) + 1)]
            bimonthly_result['Avg_Medoid_sharpe'] = sum(medoid_sharpes) / len(filtered_medoids)
            
            # Calculate the normalized average of medoid returns
            medoid_returns = [bimonthly_result[f'Medoid_{i}_return'] for i in range(1, len(filtered_medoids) + 1)]
            bimonthly_result['Avg_Medoid_return'] = sum(medoid_returns) / len(filtered_medoids)
            
            bimonthly_sharpe.append(bimonthly_result)
        
        # Convert to DataFrame
        bimonthly_sharpe_df = pd.DataFrame(bimonthly_sharpe)
        
        # Sort the DataFrame by date for proper chronological display
        if not bimonthly_sharpe_df.empty:
            bimonthly_sharpe_df = bimonthly_sharpe_df.sort_values('date')
        else:
            print(f"WARNING: No bimonthly periods found for {SYMBOL}. Cannot create chart.")
            return None
        
        # Add rounded values to dataframe for calculations
        bimonthly_sharpe_df['Best_sharpe_rounded'] = np.round(bimonthly_sharpe_df['Best_sharpe'], 2)
        bimonthly_sharpe_df['Avg_Medoid_sharpe_rounded'] = np.round(bimonthly_sharpe_df['Avg_Medoid_sharpe'], 2)
        
        # Print detailed comparison of Sharpe ratios
        print("\nDetailed Sharpe ratio comparison by period:")
        print(f"{'Period':<12} | {'Best Sharpe':>12} | {'Medoid Portfolio':>16} | {'Difference':>12} | {'Portfolio Wins':<14}")
        print("-" * 80)
        
        for idx, row in bimonthly_sharpe_df.iterrows():
            period = row['period_label']
            best_sharpe = row['Best_sharpe']
            avg_medoid_sharpe = row['Avg_Medoid_sharpe']
            best_rounded = row['Best_sharpe_rounded']
            avg_medoid_rounded = row['Avg_Medoid_sharpe_rounded']
            
            diff = avg_medoid_sharpe - best_sharpe
            portfolio_wins = avg_medoid_sharpe > best_sharpe
            
            print(f"{period:<12} | {best_sharpe:12.6f} | {avg_medoid_sharpe:16.6f} | {diff:12.6f} | {portfolio_wins!s:<14}")
        
        # Calculate win rate using raw values
        portfolio_wins = sum(bimonthly_sharpe_df['Avg_Medoid_sharpe'] > bimonthly_sharpe_df['Best_sharpe'])
        total_periods = len(bimonthly_sharpe_df)
        win_percentage = (portfolio_wins / total_periods) * 100 if total_periods > 0 else 0
        
        # Calculate win rate using rounded values (for alternative comparison)
        rounded_wins = sum(bimonthly_sharpe_df['Avg_Medoid_sharpe_rounded'] > bimonthly_sharpe_df['Best_sharpe_rounded'])
        rounded_win_percentage = (rounded_wins / total_periods) * 100 if total_periods > 0 else 0
        
        print(f"\nBimonthly periods analyzed: {total_periods}")
        print(f"Medoid Portfolio Wins: {portfolio_wins} of {total_periods} periods ({win_percentage:.2f}%)")
        print(f"Using rounded values (2 decimal places): {rounded_wins} of {total_periods} periods ({rounded_win_percentage:.2f}%)")
        
        # Create a bar plot to compare bimonthly Sharpe ratios
        plt.figure(figsize=(14, 8))
        
        # Set up x-axis dates
        x = np.arange(len(bimonthly_sharpe_df))
        width = 0.35  # Width of the bars
        
        # Create bars
        plt.bar(x - width/2, bimonthly_sharpe_df['Best_sharpe'], width, 
            label=f'Best Sharpe ({best_short_sma}/{best_long_sma})', color='blue')
        plt.bar(x + width/2, bimonthly_sharpe_df['Avg_Medoid_sharpe'], width, 
            label=f'Medoid Portfolio ({len(filtered_medoids)} strategies)', color='green')
        
        # Add a horizontal line at Sharpe = 0
        plt.axhline(y=0, color='black', linestyle='-', alpha=0.3)
        
        # Create medoid description for the title
        medoid_desc = ", ".join([f"({int(m[0])}/{int(m[1])})" for m in filtered_medoids])
        
        # Customize the plot - using rounded win percentage instead of raw
        plt.title(f'Bimonthly Sharpe Ratio Comparison (Out-of-Sample Period)\n' + 
                f'Medoid Portfolio [{medoid_desc}] outperformed {rounded_win_percentage:.2f}% of the time', 
                fontsize=14)
        plt.xlabel('Bimonthly Period (Start Month)', fontsize=12)
        plt.ylabel('Sharpe Ratio (Annualized)', fontsize=12)
        
        # Simplified x-tick labels with just the period start month
        plt.xticks(x, bimonthly_sharpe_df['period_label'], rotation=45)
        
        plt.grid(axis='y', linestyle='--', alpha=0.7)
        
        # Create legend with both strategies - moved to bottom to avoid overlap with title
        plt.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=2,
                frameon=True, fancybox=True, framealpha=0.9, fontsize=10)
        
        # Add a text box with rounded win percentage instead of raw - moved to right side
        plt.annotate(f'Medoid Portfolio Win Rate: {rounded_win_percentage:.2f}%\n'
                    f'({rounded_wins} out of {total_periods} periods)\n'
                    f'Portfolio: {len(filtered_medoids)} medoids with Sharpe ≥ {min_sharpe}\n'
                    f'ATR-Based Position Sizing (${capital:,}, {atr_period} days)',
                    xy=(0.7, 0.95), xycoords='axes fraction',
                    bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="gray", alpha=0.8),
                    fontsize=12)
        
        plt.tight_layout(rect=[0, 0, 1, 0.95])  # Add extra space at the bottom for the legend
        save_plot('Hierarchical_Bimonthly_Comparison.png')
        
        # Save bimonthly comparison to CSV
        save_results(bimonthly_sharpe_df, f"{SYMBOL}_hierarchical_bimonthly_comparison.csv")
        
        # Save win percentage and medoid parameters to Excel file
        try:
            import openpyxl
            
            # Path to the Excel file
            excel_file = r"C:\Users\Admin\Documents\darbas\Results.xlsx"
            
            # Check if file exists
            if not os.path.exists(excel_file):
                print(f"Excel file not found at: {excel_file}")
                return bimonthly_sharpe_df
                
            print(f"Updating Excel file with Hierarchical results for {SYMBOL}...")
            
            # Load the workbook
            wb = openpyxl.load_workbook(excel_file)
            
            # Get the active sheet
            sheet = wb.active
            
            # Find the row with the ticker symbol or the first empty row
            row = 3  # Start from row 3 (assuming rows 1-2 have headers)
            ticker_row = None
            
            while True:
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value == SYMBOL:
                    # Found the ticker symbol
                    ticker_row = row
                    break
                elif cell_value is None:
                    # Found an empty row
                    ticker_row = row
                    # Write the ticker symbol in column A
                    sheet.cell(row=ticker_row, column=1).value = SYMBOL
                    break
                row += 1
            
            # Round the win percentage to one decimal place
            rounded_win_percentage_1dp = round(rounded_win_percentage, 1)
            
            # Write the win percentage in column C (Hierarchy)
            sheet.cell(row=ticker_row, column=3).value = rounded_win_percentage_1dp
            
            # Write the medoid parameters in the respective cluster columns
            # Hierarchy clusters start at column I (9)
            for i, medoid in enumerate(filtered_medoids):
                if i >= 3:  # Only use up to 3 clusters
                    break
                    
                # Calculate column index: I=9, J=10, K=11 for Cluster1, Cluster2, Cluster3
                column_idx = 9 + i
                
                # Format as "short/long" (e.g., "5/20")
                param_value = f"{int(medoid[0])}/{int(medoid[1])}"
                
                # Write to Excel
                sheet.cell(row=ticker_row, column=column_idx).value = param_value
            
            # Write the best Sharpe parameters in column M (13)
            best_sharpe_params = f"{best_short_sma}/{best_long_sma}"
            sheet.cell(row=ticker_row, column=13).value = best_sharpe_params
            
            # Save the workbook
            wb.save(excel_file)
            
            print(f"Excel file updated successfully. Added {SYMBOL} with Hierarchical win rate {rounded_win_percentage_1dp}% in row {ticker_row}")
            print(f"Added best Sharpe parameters {best_sharpe_params} in column M")
            
        except Exception as e:
            print(f"Error updating Excel file: {e}")
        
        # Return the bimonthly Sharpe ratio data
        return bimonthly_sharpe_df

    def create_hierarchical_cluster_visualization(X_filtered_full, medoids, top_medoids, max_sharpe_point, labels, valid_clusters):
        """
        Create a continuous heatmap visualization with hierarchical cluster centers overlaid.
        
        Parameters:
        X_filtered_full: array - All filter-compliant data points with shape (n_samples, 4)
        medoids: list - List of medoids from valid clusters
        top_medoids: list - List of top medoids by Sharpe ratio
        max_sharpe_point: array - Point with maximum Sharpe ratio
        labels: array - Cluster labels for each point in X_filtered_full
        valid_clusters: set - Set of valid cluster IDs that meet the min_elements_per_cluster requirement
        """
        print("Creating hierarchical cluster visualization...")

        # Load the full dataset for pivot table creation
        data = pd.read_csv('sma_all_results.txt')
        
        # Create filtered dataframe from X_filtered_full but include ALL points
        # We will NOT filter by valid_clusters yet - this matches the K-means approach
        filtered_df = pd.DataFrame(X_filtered_full, columns=['short_SMA', 'long_SMA', 'sharpe_ratio', 'trades'])
        
        # Create a pivot table for the heatmap using ALL filtered data points (not just valid clusters)
        heatmap_data = filtered_df.pivot_table(
            index='long_SMA',
            columns='short_SMA',
            values='sharpe_ratio',
            fill_value=np.nan  # Use NaN for empty cells
        )

        # Create the heatmap visualization
        plt.figure(figsize=(12, 10))

        # Create mask for invalid combinations (where short_SMA >= long_SMA)
        # and also for NaN values (which represent filtered-out points)
        mask = np.zeros_like(heatmap_data, dtype=bool)
        for i, long_sma in enumerate(heatmap_data.index):
            for j, short_sma in enumerate(heatmap_data.columns):
                if short_sma >= long_sma or np.isnan(heatmap_data.iloc[i, j]):
                    mask[i, j] = True

        # Plot the base heatmap with ALL filtered data
        ax = sns.heatmap(
            heatmap_data,
            mask=mask,
            cmap='coolwarm',  # Blue to red colormap
            annot=False,      # Don't annotate each cell with its value
            fmt='.4f',
            linewidths=0,
            cbar_kws={'label': 'Sharpe Ratio'}
        )

        # Invert the y-axis so smaller long_SMA values are at the top
        ax.invert_yaxis()

        # Plot max Sharpe point (Green Star)
        try:
            best_x_pos = np.where(heatmap_data.columns == max_sharpe_point[0])[0][0] + 0.5
            best_y_pos = np.where(heatmap_data.index == max_sharpe_point[1])[0][0] + 0.5
            plt.scatter(best_x_pos, best_y_pos, marker='*', color='lime', s=200,
                        edgecolor='black', zorder=5)
        except IndexError:
            print(f"Warning: Max Sharpe point at ({max_sharpe_point[0]}, {max_sharpe_point[1]}) not found in heatmap coordinates")

        # Only plot medoids (these are already from valid clusters)
        if medoids:
            for medoid in medoids:
                try:
                    x_pos = np.where(heatmap_data.columns == medoid[0])[0][0] + 0.5
                    y_pos = np.where(heatmap_data.index == medoid[1])[0][0] + 0.5
                    plt.scatter(x_pos, y_pos, marker='s', color='black', s=75, zorder=4)
                except IndexError:
                    print(f"Warning: Medoid at ({medoid[0]}, {medoid[1]}) not found in heatmap coordinates")

        # Plot top 5 medoids (Purple Diamonds)
        if top_medoids:
            for medoid in top_medoids:
                try:
                    x_pos = np.where(heatmap_data.columns == medoid[0])[0][0] + 0.5
                    y_pos = np.where(heatmap_data.index == medoid[1])[0][0] + 0.5
                    plt.scatter(x_pos, y_pos, marker='D', color='purple', s=100, zorder=5)
                except IndexError:
                    print(f"Warning: Top medoid at ({medoid[0]}, {medoid[1]}) not found in heatmap coordinates")

        # Create custom legend
        max_sharpe_handle = mlines.Line2D([], [], color='lime', marker='*', linestyle='None',
                                        markersize=15, markeredgecolor='black', label='Max Sharpe')
        medoid_handle = mlines.Line2D([], [], color='black', marker='s', linestyle='None',
                                    markersize=10, label='Medoids')
        top_medoid_handle = mlines.Line2D([], [], color='purple', marker='D', linestyle='None',
                                        markersize=10, label='Top 5 Medoids')

        # Add legend
        plt.legend(handles=[max_sharpe_handle, medoid_handle, top_medoid_handle],
                loc='best')

        # Set labels and title
        plt.title('Hierarchical Clustering Analysis (Sharpe Ratio)', fontsize=14)
        plt.xlabel('Short SMA (days)', fontsize=12)
        plt.ylabel('Long SMA (days)', fontsize=12)

        # Rotate tick labels
        plt.xticks(rotation=45, ha='right')
        plt.yticks(rotation=0)

        # Display plot
        plt.tight_layout()
        save_plot('Hierarchical_Cluster_Analysis.png')

    # Main execution block
    # Main execution block
    if __name__ == "__main__":
            # Set matplotlib backend explicitly
            # Main execution block
            # Set matplotlib backend explicitly
            matplotlib.use('Agg')  # Use non-interactive backend for headless environments

            print("Starting ATR-based SMA strategy analysis...")

            # Run the basic analysis first
            data, best_short, best_long, best_sharpe, best_trades = analyze_sma_results()

            if data is None:
                print("Error: Failed to load or analyze SMA results data.")
                exit(1)

            print("\nProceeding with cluster analysis...")

            # Run the cluster analysis to get medoids
            X_filtered, medoids, top_medoids, max_sharpe_point, labels = hierarchical_cluster_analysis()

            if X_filtered is None or medoids is None:
                print("Error: Cluster analysis failed.")
                exit(1)

            # Re-sort top_medoids to ensure they're in the right order
            if top_medoids is not None:
                print("Re-sorting top medoids by Sharpe ratio...")
                # CRITICAL FIX: Use float() to ensure proper numeric comparison
                top_medoids = sorted(top_medoids, key=lambda x: float(x[2]), reverse=True)
                for idx, medoid in enumerate(top_medoids, 1):
                    print(f"Verified Medoid {idx}: Short SMA={int(medoid[0])}, Long SMA={int(medoid[1])}, "
                        f"Sharpe={float(medoid[2]):.4f}, Trades={int(medoid[3])}")

            print("\nPlotting strategy performance...")

            # Plot strategy performance with the best parameters AND top medoids using ATR-based position sizing
            market_data = plot_strategy_performance(
                best_short, best_long, top_medoids, 
                big_point_value=big_point_value,
                slippage=slippage,
                capital=capital,
                atr_period=atr_period
            )
            
            # Run the bimonthly out-of-sample comparison between best Sharpe and top medoids
            if top_medoids and len(top_medoids) > 0:
                print("\nPerforming bimonthly out-of-sample comparison with hierarchical clustering...")
                bimonthly_sharpe_df = bimonthly_out_of_sample_comparison(
                    market_data, 
                    best_short, 
                    best_long, 
                    top_medoids,  # Pass the entire list of top medoids
                    big_point_value=big_point_value,
                    slippage=slippage,
                    capital=capital,
                    atr_period=atr_period
                )
            else:
                print("No top medoids found. Cannot run bimonthly comparison.")
                
            print("\nAnalysis complete! All plots and result files have been saved to the output directory.")
            print(f"Output directory: {output_dir}")

# Call the main function to execute the script
if __name__ == "__main__":
    main()