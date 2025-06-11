import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import matplotlib.lines as mlines
from matplotlib.patches import Circle
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from datetime import datetime
import calendar
import read_ts
import openpyxl

import os
import matplotlib.pyplot as plt
import matplotlib

from input_gen import MIN_TRADES, MAX_TRADES, MIN_ELEMENTS_PER_CLUSTER, DEFAULT_NUM_CLUSTERS, ELBOW_THRESHOLD
from input_gen import TICKER, START_DATE, END_DATE, TRAIN_TEST_SPLIT, ATR_PERIOD, TRADING_CAPITAL
from SMA_Strategy import SMAStrategy

import glob
import json

SYMBOL = TICKER.replace('=F', '')  # Define SYMBOL based on TICKER

def save_plot(plot_name):
    output_dir = os.path.join(".", 'output2', TICKER.replace('=F', ''))
    os.makedirs(output_dir, exist_ok=True)
    plt.savefig(os.path.join(output_dir, plot_name))
    plt.close()

def calculate_elbow_curve(X_scaled, max_clusters=20):
    """
    Calculate the elbow curve for KMeans clustering
    
    Parameters:
    X_scaled: numpy array - Scaled data points
    max_clusters: int - Maximum number of clusters to try
    
    Returns:
    distortions: list - Distortion values for each k
    k_values: list - K values used
    optimal_k: int - Optimal number of clusters based on elbow method
    """
    print("\nCalculating elbow curve...")
    distortions = []
    k_values = range(1, max_clusters + 1)
    
    for k in k_values:
        if k == 1:
            # For k=1, distortion is just the sum of squared distances to mean
            distortion = np.sum((X_scaled - np.mean(X_scaled, axis=0)) ** 2)
            distortions.append(distortion)
            continue
            
        kmeans = KMeans(n_clusters=k, random_state=42, n_init=10)
        kmeans.fit(X_scaled)
        distortions.append(kmeans.inertia_)
    
    # Calculate the percentage changes
    pct_changes = np.diff(distortions) / np.array(distortions)[:-1] * 100

    # Find optimal k using the threshold method
    optimal_k = 1
    for i, pct_change in enumerate(pct_changes):
        if abs(pct_change) < ELBOW_THRESHOLD:
            optimal_k = i + 1  # +1 because we start from k=1
            break
    
    # Plot the elbow curve
    plt.figure(figsize=(10, 6))
    plt.plot(k_values, distortions, 'bo-')
    plt.axvline(x=optimal_k, color='r', linestyle='--', 
                label=f'Optimal k={optimal_k}\n(threshold={ELBOW_THRESHOLD}%)')
    plt.xlabel('Number of Clusters (k)')
    plt.ylabel('Distortion')
    plt.title(f'{SYMBOL} KMeans Elbow Method for Optimal k')
    plt.grid(True)
    plt.legend()
    plt.tight_layout()
    save_plot(f'{SYMBOL}_KMeans_Elbow_Curve.png')
    
    print(f"Optimal number of clusters from elbow method: {optimal_k}")
    return distortions, k_values, optimal_k

def main():
    def load_parameters():
        try:
            with open("parameters.json", "r") as file:
                parameters = json.load(file)
                big_point_value = parameters["big_point_value"]
                slippage = parameters["slippage"]
                capital = parameters.get("capital", TRADING_CAPITAL)
                atr_period = parameters.get("atr_period", ATR_PERIOD)
                return big_point_value, slippage, capital, atr_period
        except FileNotFoundError:
            print("Parameters file not found. Ensuring it was saved correctly in data_gather_gen.py.")
            return None, None, TRADING_CAPITAL, ATR_PERIOD

    # Load the parameters
    big_point_value, slippage, capital, atr_period = load_parameters()

    print(f"Big Point Value: {big_point_value}")
    print(f"Dynamic Slippage: {slippage}")
    print(f"Capital for Position Sizing: {capital:,}")
    print(f"ATR Period: {atr_period}")

    # Setup paths
    WORKING_DIR = "."  # Current directory
    DATA_DIR = os.path.join(WORKING_DIR, "data")

    # Define the output folder where the plots will be saved
    output_dir = os.path.join(WORKING_DIR, 'output2', SYMBOL)
    os.makedirs(output_dir, exist_ok=True)  # Create the directory if it doesn't exist

    # Function to save plots in the created folder
    def save_plot(plot_name):
        plt.savefig(os.path.join(output_dir, plot_name))  # Save plot to the symbol-specific folder
        plt.close()  # Close the plot to free up memory

    def find_futures_file(symbol, data_dir):
        """Find a data file for the specified futures symbol"""
        # First try a pattern that specifically looks for @SYMBOL
        pattern = f"*@{symbol}_*.dat"
        files = glob.glob(os.path.join(data_dir, pattern))
        
        if not files:
            # Try a pattern that looks for the symbol with an underscore or at boundary
            pattern = f"*_@{symbol}_*.dat"
            files = glob.glob(os.path.join(data_dir, pattern))
        
        if not files:
            # Try a more specific boundary pattern for the symbol
            pattern = f"*_{symbol}_*.dat"
            files = glob.glob(os.path.join(data_dir, pattern))
        
        if not files:
            # Last resort: less specific but better than nothing
            pattern = f"*@{symbol}*.dat"
            files = glob.glob(os.path.join(data_dir, pattern))
        
        return files[0] if files else None

    # Added functions from cluster_visualization.py
    def compute_medoids(X, labels, valid_clusters):
        """Compute medoids for each cluster (point with minimum distance to all other points in cluster)"""
        medoids = []

        for cluster_id in valid_clusters:
            # Get points in this cluster
            cluster_points = X[labels == cluster_id]

            if len(cluster_points) == 0:
                continue

            # Calculate pairwise distances within cluster
            min_total_distance = float('inf')
            medoid = None

            for i, point1 in enumerate(cluster_points):
                total_distance = 0
                for j, point2 in enumerate(cluster_points):
                    # Calculate Euclidean distance between points
                    distance = np.sqrt(np.sum((point1 - point2) ** 2))
                    total_distance += distance

                if total_distance < min_total_distance:
                    min_total_distance = total_distance
                    medoid = point1

            if medoid is not None:
                medoids.append(medoid)

        return medoids

    def cluster_analysis(file_path='sma_all_results.txt', min_trades=MIN_TRADES, max_trades=MAX_TRADES,
                    min_elements_per_cluster=MIN_ELEMENTS_PER_CLUSTER):
        """
        Perform clustering analysis on SMA optimization results to find robust parameter regions
        Now clusters based on short_SMA, long_SMA, and sharpe_ratio only (not trades)
        """
        print(f"\n----- CLUSTER ANALYSIS -----")
        print(f"Loading data from {file_path}...")

        # Load the data
        df = pd.read_csv(file_path)

        # Convert data to numpy array for easier processing
        X_full = df[['short_SMA', 'long_SMA', 'sharpe_ratio', 'trades']].values

        # Filter data by number of trades and ensure short_SMA < long_SMA
        X_filtered_full = X_full[(X_full[:, 0] < X_full[:, 1]) &  # short_SMA < long_SMA
                    (X_full[:, 3] >= min_trades) &  # trades >= min_trades
                    (X_full[:, 3] <= max_trades)]  # trades <= max_trades

        if len(X_filtered_full) == 0:
            print(
                f"No data points meet the criteria after filtering! Adjust min_trades ({min_trades}) and max_trades ({max_trades}).")
            return None, None, None, None, None

        # Create a version with only the 3 dimensions for clustering
        X_filtered = X_filtered_full[:, 0:3]  # Only short_SMA, long_SMA, and sharpe_ratio

        print(f"Filtered data to {len(X_filtered)} points with {min_trades}-{max_trades} trades")

        # Extract the fields for better scaling visibility
        short_sma_values = X_filtered[:, 0]
        long_sma_values = X_filtered[:, 1]
        sharpe_values = X_filtered[:, 2]
        trades_values = X_filtered_full[:, 3]

        print(f"Short SMA range: {short_sma_values.min()} to {short_sma_values.max()}")
        print(f"Long SMA range: {long_sma_values.min()} to {long_sma_values.max()}")
        print(f"Sharpe ratio range: {sharpe_values.min():.4f} to {sharpe_values.max():.4f}")
        print(f"Trades range: {trades_values.min()} to {trades_values.max()}")

        # Scale the data for clustering - using StandardScaler for each dimension
        # This addresses the issue where SMA values have much larger ranges than Sharpe ratio
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X_filtered)  # Only scale the 3 dimensions we use for clustering

        # Print scaling info for verification
        print("\nScaled data information:")
        scaled_short = X_scaled[:, 0]
        scaled_long = X_scaled[:, 1]
        scaled_sharpe = X_scaled[:, 2]

        print(f"Scaled Short SMA range: {scaled_short.min():.4f} to {scaled_short.max():.4f}")
        print(f"Scaled Long SMA range: {scaled_long.min():.4f} to {scaled_long.max():.4f}")
        print(f"Scaled Sharpe ratio range: {scaled_sharpe.min():.4f} to {scaled_sharpe.max():.4f}")

        # Determine number of clusters using elbow method
        print("\nDetermining optimal number of clusters using elbow method...")
        _, _, k = calculate_elbow_curve(X_scaled, max_clusters=min(50, len(X_scaled) // 2))
        print(f"Using {k} clusters based on elbow method (threshold={ELBOW_THRESHOLD})")

        # Apply KMeans clustering
        print(f"Performing KMeans clustering with k={k}...")
        kmeans = KMeans(n_clusters=k, random_state=42, n_init=10)
        kmeans.fit(X_scaled)

        # Get cluster labels
        labels = kmeans.labels_

        # Count elements per cluster
        unique_labels, counts = np.unique(labels, return_counts=True)
        cluster_sizes = dict(zip(unique_labels, counts))

        print("\nCluster sizes:")
        for cluster_id, size in cluster_sizes.items():
            print(f"Cluster {cluster_id}: {size} elements")

        # Filter clusters with enough elements
        valid_clusters = {i for i, count in cluster_sizes.items() if count >= min_elements_per_cluster}
        filtered_indices = np.array([i in valid_clusters for i in labels])

        # Filter data to only include points in valid clusters
        X_valid = X_filtered_full[filtered_indices]  # Use full data to get trades too
        labels_valid = labels[filtered_indices]

        # Compute medoids using the existing method that expects 4D data
        print("Computing medoids...")
        medoids = compute_medoids(X_valid, labels_valid, valid_clusters)

        # Compute centroids (only for the 3 dimensions we clustered on)
        centroids_scaled = kmeans.cluster_centers_
        # Create a version that can be inverse-transformed (matches the original dimension count)
        centroids = np.zeros((centroids_scaled.shape[0], 3))
        centroids[:, 0:3] = scaler.inverse_transform(centroids_scaled)

        # Print raw centroids for debugging
        print("\nCluster Centroids (in original space):")
        for i, centroid in enumerate(centroids):
            if i in valid_clusters:  # Only show valid clusters
                print(f"Centroid {i}: Short SMA={centroid[0]:.2f}, Long SMA={centroid[1]:.2f}, "
                    f"Sharpe={centroid[2]:.4f}")

        # Sort medoids by Sharpe ratio
        medoids_sorted = sorted(medoids, key=lambda x: x[2], reverse=True)
        top_medoids = medoids_sorted[:5]  # Get top 5 medoids by Sharpe ratio

        # Find max Sharpe ratio point overall
        max_sharpe_idx = np.argmax(df['sharpe_ratio'].values)
        max_sharpe_point = df.iloc[max_sharpe_idx][['short_SMA', 'long_SMA', 'sharpe_ratio', 'trades']].values

        # Print results
        print("\n----- CLUSTERING RESULTS -----")
        print(f"Max Sharpe point: Short SMA={int(max_sharpe_point[0])}, Long SMA={int(max_sharpe_point[1])}, "
            f"Sharpe={max_sharpe_point[2]:.4f}, Trades={int(max_sharpe_point[3])}")

        print("\nTop 5 Medoids (by Sharpe ratio):")
        for idx, medoid in enumerate(top_medoids, 1):
            print(f"Top {idx}: Short SMA={int(medoid[0])}, Long SMA={int(medoid[1])}, "
                f"Sharpe={medoid[2]:.4f}, Trades={int(medoid[3])}")

        # Create visualization with clustering results - pass the original labels and valid_clusters
        create_cluster_visualization(X_filtered_full, medoids, top_medoids, centroids, max_sharpe_point, 
                                labels=labels, valid_clusters=valid_clusters)

        return X_filtered_full, medoids, top_medoids, centroids, max_sharpe_point

    def create_cluster_visualization(X_filtered_full, medoids, top_medoids, centroids, max_sharpe_point, labels=None, valid_clusters=None):
        """
        Create a continuous heatmap visualization with cluster centers overlaid.
        Only plots data points and clusters that meet the filtering criteria.
        
        Parameters:
        X_filtered_full: array - Filter-compliant data points with shape (n_samples, 4) containing short_SMA, long_SMA, sharpe_ratio, trades
        medoids: list - List of medoids from valid clusters
        top_medoids: list - List of top medoids by Sharpe ratio
        centroids: array - Centroids of clusters
        max_sharpe_point: array - Point with maximum Sharpe ratio
        labels: array - Cluster labels for each point in X_filtered_full (optional)
        valid_clusters: set - Set of valid cluster IDs that meet the min_elements_per_cluster requirement (optional)
        """
        print("Creating cluster visualization...")

        # Load the full dataset, but we'll only use it for creating the heatmap grid
        data = pd.read_csv('sma_all_results.txt')
        
        # Create filtered dataframe from the filtered points that meet trade requirements
        # This ensures we only visualize points that meet the trade count requirements
        filtered_df = pd.DataFrame(X_filtered_full, columns=['short_SMA', 'long_SMA', 'sharpe_ratio', 'trades'])
        
        # Create a pivot table for the heatmap using ONLY the filtered data points
        heatmap_data = filtered_df.pivot_table(
            index='long_SMA',
            columns='short_SMA',
            values='sharpe_ratio',
            fill_value=np.nan  # Use NaN for empty cells
        )

        # Create the heatmap visualization
        plt.figure(figsize=(12, 10))

        # Create mask for invalid combinations (where short_SMA >= long_SMA)
        # and also for NaN values (which represent filtered-out points)
        mask = np.zeros_like(heatmap_data, dtype=bool)
        for i, long_sma in enumerate(heatmap_data.index):
            for j, short_sma in enumerate(heatmap_data.columns):
                if short_sma >= long_sma or np.isnan(heatmap_data.iloc[i, j]):
                    mask[i, j] = True

        # Plot the base heatmap with ONLY filtered data
        ax = sns.heatmap(
            heatmap_data,
            mask=mask,
            cmap='coolwarm',  # Blue to red colormap
            annot=False,      # Don't annotate each cell with its value
            fmt='.4f',
            linewidths=0,
            cbar_kws={'label': 'Sharpe Ratio'}
        )

        # Invert the y-axis so smaller long_SMA values are at the top
        ax.invert_yaxis()

        # Plot max Sharpe point (Green Star)
        # Note: We need to convert from data values to plot coordinates
        try:
            best_x_pos = np.where(heatmap_data.columns == max_sharpe_point[0])[0][0] + 0.5
            best_y_pos = np.where(heatmap_data.index == max_sharpe_point[1])[0][0] + 0.5
            plt.scatter(best_x_pos, best_y_pos, marker='*', color='lime', s=200,
                        edgecolor='black', zorder=5)
        except IndexError:
            print(f"Warning: Max Sharpe point at ({max_sharpe_point[0]}, {max_sharpe_point[1]}) not found in heatmap coordinates")

        # Only plot medoids from valid clusters
        if medoids:
            for medoid in medoids:
                try:
                    x_pos = np.where(heatmap_data.columns == medoid[0])[0][0] + 0.5
                    y_pos = np.where(heatmap_data.index == medoid[1])[0][0] + 0.5
                    plt.scatter(x_pos, y_pos, marker='s', color='black', s=75, zorder=4)
                except IndexError:
                    print(f"Warning: Medoid at ({medoid[0]}, {medoid[1]}) not found in heatmap coordinates")

        # Plot top 5 medoids (Purple Diamonds)
        if top_medoids:
            for medoid in top_medoids:
                try:
                    x_pos = np.where(heatmap_data.columns == medoid[0])[0][0] + 0.5
                    y_pos = np.where(heatmap_data.index == medoid[1])[0][0] + 0.5
                    plt.scatter(x_pos, y_pos, marker='D', color='purple', s=100, zorder=5)
                except IndexError:
                    print(f"Warning: Top medoid at ({medoid[0]}, {medoid[1]}) not found in heatmap coordinates")

        # Only plot centroids from valid clusters
        print(f"Plotting centroids from valid clusters...")
        centroids_plotted = 0

        # If valid_clusters is not provided, assume all centroids are valid
        plot_centroids = centroids
        if valid_clusters is not None and labels is not None:
            # Only plot centroids from valid clusters
            plot_centroids = [centroids[i] for i in range(len(centroids)) if i in valid_clusters]
            print(f"Filtering centroids to only include valid clusters: {valid_clusters}")
        
        for i, centroid in enumerate(plot_centroids):
            # Get the actual raw centroid values
            short_sma = centroid[0]
            long_sma = centroid[1]

            print(f"Centroid {i}: raw values = ({short_sma}, {long_sma})")

            # First try exact values
            try:
                if (short_sma in heatmap_data.columns) and (long_sma in heatmap_data.index) and (short_sma < long_sma):
                    x_pos = np.where(heatmap_data.columns == short_sma)[0][0] + 0.5
                    y_pos = np.where(heatmap_data.index == long_sma)[0][0] + 0.5
                    plt.scatter(x_pos, y_pos, marker='o', color='blue', s=75, zorder=4)
                    centroids_plotted += 1
                    continue
            except (IndexError, TypeError):
                pass

            # Try rounded values
            try:
                short_sma_rounded = int(round(short_sma))
                long_sma_rounded = int(round(long_sma))

                print(f"  Rounded: ({short_sma_rounded}, {long_sma_rounded})")

                if (short_sma_rounded in heatmap_data.columns) and (long_sma_rounded in heatmap_data.index) and (
                        short_sma_rounded < long_sma_rounded):
                    x_pos = np.where(heatmap_data.columns == short_sma_rounded)[0][0] + 0.5
                    y_pos = np.where(heatmap_data.index == long_sma_rounded)[0][0] + 0.5
                    plt.scatter(x_pos, y_pos, marker='o', color='blue', s=75, zorder=4)
                    centroids_plotted += 1
                    continue
            except (IndexError, TypeError):
                pass

            # Try finding nearest valid point
            try:
                # Find nearest valid parameter values
                short_options = np.array(heatmap_data.columns)
                long_options = np.array(heatmap_data.index)

                # Find nearest short SMA
                short_idx = np.argmin(np.abs(short_options - short_sma))
                short_nearest = short_options[short_idx]

                # Find nearest long SMA
                long_idx = np.argmin(np.abs(long_options - long_sma))
                long_nearest = long_options[long_idx]

                print(f"  Nearest: ({short_nearest}, {long_nearest})")

                # Check if valid
                if short_nearest < long_nearest:
                    x_pos = np.where(heatmap_data.columns == short_nearest)[0][0] + 0.5
                    y_pos = np.where(heatmap_data.index == long_nearest)[0][0] + 0.5
                    plt.scatter(x_pos, y_pos, marker='o', color='blue', s=75, zorder=4, alpha=0.7)
                    centroids_plotted += 1
                else:
                    print(f"  Invalid nearest parameters (short >= long): {short_nearest} >= {long_nearest}")
            except (IndexError, TypeError) as e:
                print(f"  Error finding nearest point: {e}")

        print(f"Successfully plotted {centroids_plotted} out of {len(plot_centroids)} centroids")

        # Create custom legend
        max_sharpe_handle = mlines.Line2D([], [], color='lime', marker='*', linestyle='None',
                                        markersize=15, markeredgecolor='black', label='Max Sharpe')
        medoid_handle = mlines.Line2D([], [], color='black', marker='s', linestyle='None',
                                    markersize=10, label='Medoids')
        top_medoid_handle = mlines.Line2D([], [], color='purple', marker='D', linestyle='None',
                                        markersize=10, label='Top 5 Medoids')
        centroid_handle = mlines.Line2D([], [], color='blue', marker='o', linestyle='None',
                                        markersize=10, label='Centroids')

        # Add legend
        plt.legend(handles=[max_sharpe_handle, medoid_handle, top_medoid_handle, centroid_handle],
                loc='best')

        # Set labels and title
        plt.title(f'{SYMBOL} SMA Parameter Clustering Analysis (Sharpe Ratio)', fontsize=14)
        plt.xlabel('Short SMA (days)', fontsize=12)
        plt.ylabel('Long SMA (days)', fontsize=12)

        # Rotate tick labels
        plt.xticks(rotation=45, ha='right')
        plt.yticks(rotation=0)

        # Display plot
        plt.tight_layout()
        save_plot(f'{SYMBOL}_KMeans_Cluster_Analysis.png')

    # Load the SMA simulation results
    def analyze_sma_results(file_path='sma_all_results.txt'):
        print(f"Loading simulation results from {file_path}...")

        # Load the data from the CSV file
        data = pd.read_csv(file_path)

        # Print basic information about the data
        print(f"Loaded {len(data)} simulation results")
        print(f"Short SMA range: {data['short_SMA'].min()} to {data['short_SMA'].max()}")
        print(f"Long SMA range: {data['long_SMA'].min()} to {data['long_SMA'].max()}")
        print(f"Sharpe ratio range: {data['sharpe_ratio'].min():.4f} to {data['sharpe_ratio'].max():.4f}")

        # Find the best Sharpe ratio
        best_idx = data['sharpe_ratio'].idxmax()
        best_short_sma = data.loc[best_idx, 'short_SMA']
        best_long_sma = data.loc[best_idx, 'long_SMA']
        best_sharpe = data.loc[best_idx, 'sharpe_ratio']
        best_trades = data.loc[best_idx, 'trades']

        print(f"\nBest parameters:")
        print(f"Short SMA: {best_short_sma}")
        print(f"Long SMA: {best_long_sma}")
        print(f"Sharpe Ratio: {best_sharpe:.6f}")
        print(f"Number of Trades: {best_trades}")

        # Create a pivot table for the heatmap
        heatmap_data = data.pivot_table(
            index='long_SMA',
            columns='short_SMA',
            values='sharpe_ratio'
        )

        # Create the heatmap visualization
        plt.figure(figsize=(12, 10))

        # Create a mask for invalid combinations (where short_SMA >= long_SMA)
        mask = np.zeros_like(heatmap_data, dtype=bool)
        for i, long_sma in enumerate(heatmap_data.index):
            for j, short_sma in enumerate(heatmap_data.columns):
                if short_sma >= long_sma:
                    mask[i, j] = True

        # Plot the heatmap with the mask
        ax = sns.heatmap(
            heatmap_data,
            mask=mask,
            cmap='coolwarm',  # Blue to red colormap
            annot=False,  # Don't annotate each cell with its value
            fmt='.4f',
            linewidths=0,
            cbar_kws={'label': 'Sharpe Ratio'}
        )

        # Invert the y-axis so smaller long_SMA values are at the top
        ax.invert_yaxis()

        # Find the position of the best Sharpe ratio in the heatmap
        best_y = heatmap_data.index.get_loc(best_long_sma)
        best_x = heatmap_data.columns.get_loc(best_short_sma)

        # Add a star to mark the best Sharpe ratio
        # We need to add 0.5 to center the marker in the cell
        ax.add_patch(Circle((best_x + 0.5, best_y + 0.5), 0.4, facecolor='none',
                            edgecolor='white', lw=2))
        plt.plot(best_x + 0.5, best_y + 0.5, 'w*', markersize=10)

        # Set labels and title
        plt.title(f'{SYMBOL} SMA Optimization Heatmap (Best Sharpe: {best_sharpe:.4f} at {best_short_sma}/{best_long_sma})',
                fontsize=14)
        plt.xlabel('Short SMA (days)', fontsize=12)
        plt.ylabel('Long SMA (days)', fontsize=12)

        # Rotate tick labels for better readability
        plt.xticks(rotation=45, ha='right')
        plt.yticks(rotation=0)

        # Add a text annotation for the best parameters
        plt.annotate(
            f'Best: Short={best_short_sma}, Long={best_long_sma}\nSharpe={best_sharpe:.4f}, Trades={best_trades}',
            xy=(best_x + 0.5, best_y + 0.5),
            xytext=(best_x + 5, best_y + 5),
            arrowprops=dict(arrowstyle="->", color='white'),
            color='white',
            backgroundcolor='black',
            bbox=dict(boxstyle="round,pad=0.3", fc="black", alpha=0.7)
        )

        # Display the plot
        plt.tight_layout()
        save_plot(f'{SYMBOL}_KMeans_Heatmap.png')

        # Return the data and best parameters
        return data, best_short_sma, best_long_sma, best_sharpe, best_trades


    # Update the plot_strategy_performance function in data_analysis.py
# Around line 587 in the function, replace the existing plotting code with:

    def plot_strategy_performance(short_sma, long_sma, top_medoids=None, big_point_value=big_point_value,
                            slippage=slippage, capital=capital, atr_period=atr_period):
        """
        Plot the strategy performance using the best SMA parameters and include top medoids
        Uses the SMAStrategy class for consistent logic across the codebase

        Parameters:
        short_sma: int - The short SMA period
        long_sma: int - The long SMA period
        top_medoids: list - List of top medoids, each as (short_sma, long_sma, sharpe, trades)
        big_point_value: float - Big point value for the futures contract
        slippage: float - Slippage value in price units
        capital: float - Capital allocation for position sizing
        atr_period: int - Period for ATR calculation
        """
        print(f"\n----- PLOTTING STRATEGY PERFORMANCE -----")
        print(f"Using Short SMA: {short_sma}, Long SMA: {long_sma}")
        print(f"Trading with ATR-based position sizing (Capital: ${capital:,}, ATR Period: {atr_period})")
        if top_medoids:
            print(f"Including top {len(top_medoids)} medoids")

        # Load data from local file
        print(f"Loading {TICKER} data from local files...")
        data_file = find_futures_file(SYMBOL, DATA_DIR)
        if not data_file:
            print(f"Error: No data file found for {TICKER} in {DATA_DIR}")
            exit(1)

        print(f"Found data file: {os.path.basename(data_file)}")
        print(f"File size: {os.path.getsize(data_file)} bytes")

        # Load the data from local file
        all_data = read_ts.read_ts_ohlcv_dat(data_file)
        data_obj = all_data[0]
        ohlc_data = data_obj.data.copy()

        # Convert the OHLCV data to the format expected by the strategy
        data = ohlc_data.rename(columns={
            'datetime': 'Date',
            'open': 'Open',
            'high': 'High',
            'low': 'Low',
            'close': 'Close',
            'volume': 'Volume'
        })
        data.set_index('Date', inplace=True)

        # Add a warm-up period before the start date
        original_start_idx = None
        if START_DATE and END_DATE:
            # Calculate warm-up period for SMA calculation (longest SMA + buffer)
            warm_up_days = max(short_sma, long_sma) * 3  # Use 3x the longest SMA as warm-up
            
            # Convert dates to datetime
            start_date = pd.to_datetime(START_DATE)
            end_date = pd.to_datetime(END_DATE)
            
            # Adjust start date for warm-up
            adjusted_start = start_date - pd.Timedelta(days=warm_up_days)
            
            # Filter with the extended date range
            extended_data = data[(data.index >= adjusted_start) & (data.index <= end_date)]
            
            # Store the index where the actual analysis should start
            if not extended_data.empty:
                # Find the closest index to our original start date
                original_start_idx = extended_data.index.get_indexer([start_date], method='nearest')[0]
            
            print(f"Added {warm_up_days} days warm-up period before {START_DATE}")
            data = extended_data
        
        # Create a dictionary to store results for each strategy
        strategies = {
            'Best': {'short_sma': short_sma, 'long_sma': long_sma}
        }

        # Add medoids in their original order, including original Sharpe and Trades
        if top_medoids:
            print("\nUSING THESE MEDOIDS (IN ORIGINAL ORDER):")
            for i, medoid in enumerate(top_medoids, 1):
                strategies[f'Medoid {i}'] = {
                    'short_sma': int(medoid[0]),
                    'long_sma': int(medoid[1]),
                    'original_sharpe': float(medoid[2]),  # Store the original Sharpe ratio
                    'original_trades': int(medoid[3])     # Store the original number of trades
                }
                print(f"Medoid {i}: SMA({int(medoid[0])}/{int(medoid[1])}) - Original Sharpe: {float(medoid[2]):.4f}, Trades: {int(medoid[3])}")

        # Apply the proper strategy for each parameter set
        for name, params in strategies.items():
            # Calculate centered SMAs directly
            data[f'SMA_Short_{name}'] = data['Close'].rolling(window=params['short_sma'], center=True).mean()
            data[f'SMA_Long_{name}'] = data['Close'].rolling(window=params['long_sma'], center=True).mean()
            
            # Create a strategy instance for each parameter set
            sma_strategy = SMAStrategy(
                short_sma=params['short_sma'],
                long_sma=params['long_sma'],
                big_point_value=big_point_value,
                slippage=slippage,
                capital=capital,
                atr_period=atr_period
            )

            # Apply the strategy
            data = sma_strategy.apply_strategy(
                data.copy(),
                strategy_name=name
            )

            # Add returns calculation (for compatibility with old code)
            data[f'Returns_{name}'] = data[f'Daily_PnL_{name}'] / capital
            data[f'Cumulative_Returns_{name}'] = (1 + data[f'Returns_{name}']).cumprod()
            data[f'Capital_{name}'] = capital * data[f'Cumulative_Returns_{name}']

        # Trim data to the original date range if we added warm-up period
        if original_start_idx is not None:
            data_for_evaluation = data.iloc[original_start_idx:]
            print(f"Trimmed warm-up period. Original data length: {len(data)}, Evaluation data length: {len(data_for_evaluation)}")
        else:
            data_for_evaluation = data
        
        # Calculate split index for in-sample/out-of-sample
        split_index = int(len(data_for_evaluation) * TRAIN_TEST_SPLIT)
        split_date = data_for_evaluation.index[split_index]

        # Create color palette for strategies
        colors = ['blue', 'green', 'purple', 'orange', 'brown', 'pink']

        # Create the performance visualization with only two panels (removed the middle panel)
        plt.figure(figsize=(14, 12))

        # Plot price and SMA
        plt.subplot(2, 1, 1)
        plt.plot(data_for_evaluation.index, data_for_evaluation['Close'], label=f'{SYMBOL} Price', color='black', alpha=0.5)
        
        # Use the centered SMAs for plotting
        for name, params in strategies.items():
            if name == 'Best':
                plt.plot(data_for_evaluation.index, data_for_evaluation[f'SMA_Short_{name}'], 
                        label=f'Short SMA ({params["short_sma"]})', color='orange')
                plt.plot(data_for_evaluation.index, data_for_evaluation[f'SMA_Long_{name}'], 
                        label=f'Long SMA ({params["long_sma"]})', color='blue')
        
        # Mark the train/test split
        plt.axvline(x=split_date, color='black', linestyle='--', alpha=0.7)
        
        plt.legend(loc='upper left')
        plt.title(f'{SYMBOL} Price and SMA Indicators')
        plt.ylabel('Price')
        plt.grid(True, alpha=0.3)

        # Plot cumulative P&L (second subplot)
        plt.subplot(2, 1, 2)

        # Reset all strategies to start from zero at the beginning
        for i, (name, params) in enumerate(strategies.items()):
            color = colors[i % len(colors)]
            
            # Get the daily P&L and calculate cumulative
            daily_pnl = data_for_evaluation[f'Daily_PnL_{name}']
            cumulative_pnl = daily_pnl.cumsum()
            
            # Plot full period P&L
            plt.plot(data_for_evaluation.index, cumulative_pnl,
                    label=f'{name} ({params["short_sma"]}/{params["long_sma"]})', color=color)

            # Plot out-of-sample portion with thicker line
            plt.plot(data_for_evaluation.index[split_index:], cumulative_pnl.iloc[split_index:],
                    color=color, linewidth=2.5, alpha=0.7)

        plt.axvline(x=split_date, color='black', linestyle='--',
                    label=f'Train/Test Split ({int(TRAIN_TEST_SPLIT * 100)}%/{int((1 - TRAIN_TEST_SPLIT) * 100)}%)')
        plt.axhline(y=0.0, color='gray', linestyle='-', alpha=0.5, label='Break-even')
        plt.legend(loc='upper left')
        plt.title(f'{SYMBOL} Strategy Cumulative P&L')
        plt.ylabel('P&L ($)')
        plt.grid(True, alpha=0.3)

        plt.tight_layout()
        save_plot(f'{SYMBOL}_KMeans_Multiple_Strategy_Plots.png')

        # Print detailed performance metrics for all strategies with in-sample and out-of-sample breakdown
        print("\n----- PERFORMANCE SUMMARY -----")

        # Create DataFrame to store performance data
        performance_data = []

        # IN-SAMPLE PERFORMANCE
        print("\nIN-SAMPLE PERFORMANCE:")
        header = f"{'Strategy':<10} | {'Short/Long':<10} | {'P&L':>12} | {'Avg Pos':>8} | {'Sharpe':>7} | {'Trades':>6}"
        separator = "-" * len(header)
        print(separator)
        print(header)
        print(separator)

        for name, params in strategies.items():
            short = params['short_sma']
            long = params['long_sma']

            # Get in-sample data
            in_sample = data.iloc[:split_index]

            # Calculate in-sample metrics
            in_sample_daily_pnl = in_sample[f'Daily_PnL_{name}']
            in_sample_cumulative_pnl = in_sample[f'Daily_PnL_{name}'].sum()
            
            # Calculate average position size for in-sample
            avg_pos_size = in_sample[f'Position_Size_{name}'].mean()

            # Use original Sharpe and Trades for medoids, calculate for Best
            if name.startswith('Medoid'):
                in_sample_sharpe = params['original_sharpe']
                in_sample_trades = params['original_trades']
            else:
                # Calculate Sharpe ratio (annualized) for Best
                if in_sample_daily_pnl.std() > 0:
                    in_sample_sharpe = in_sample_daily_pnl.mean() / in_sample_daily_pnl.std() * np.sqrt(252)
                else:
                    in_sample_sharpe = 0
                # Count in-sample trades for Best
                in_sample_trades = in_sample[f'Position_Change_{name}'].sum()

            # Create row text
            row = f"{name:<10} | {short:>4}/{long:<5} | ${in_sample_cumulative_pnl:>10,.2f} | {avg_pos_size:>8.2f} | {in_sample_sharpe:>6.3f} | {in_sample_trades:>6}"

            # Print row
            print(row)

            # Store the data for later use
            performance_data.append({
                'Period': 'In-Sample',
                'Strategy': name,
                'Short_SMA': short,
                'Long_SMA': long,
                'PnL': in_sample_cumulative_pnl,
                'Avg_Position': avg_pos_size,
                'Sharpe': in_sample_sharpe,
                'Trades': in_sample_trades
            })

        print(separator)

        # OUT-OF-SAMPLE PERFORMANCE
        print("\nOUT-OF-SAMPLE PERFORMANCE:")
        print(separator)
        print(header)
        print(separator)

        for name, params in strategies.items():
            short = params['short_sma']
            long = params['long_sma']

            # Get out-of-sample data
            out_sample = data.iloc[split_index:]

            # Calculate out-of-sample metrics
            out_sample_daily_pnl = out_sample[f'Daily_PnL_{name}']

            # Cumulative P&L just for the out-of-sample period
            out_sample_cumulative_pnl = out_sample[f'Daily_PnL_{name}'].sum()
            
            # Calculate average position size for out-of-sample
            avg_pos_size = out_sample[f'Position_Size_{name}'].mean()

            # Calculate Sharpe ratio (annualized)
            if out_sample_daily_pnl.std() > 0:
                out_sample_sharpe = out_sample_daily_pnl.mean() / out_sample_daily_pnl.std() * np.sqrt(252)
            else:
                out_sample_sharpe = 0

            # Count out-of-sample trades
            out_sample_trades = out_sample[f'Position_Change_{name}'].sum()

            # Create row text
            row = f"{name:<10} | {short:>4}/{long:<5} | ${out_sample_cumulative_pnl:>10,.2f} | {avg_pos_size:>8.2f} | {out_sample_sharpe:>6.3f} | {out_sample_trades:>6}"

            # Print row
            print(row)

            # Store the data
            performance_data.append({
                'Period': 'Out-of-Sample',
                'Strategy': name,
                'Short_SMA': short,
                'Long_SMA': long,
                'PnL': out_sample_cumulative_pnl,
                'Avg_Position': avg_pos_size,
                'Sharpe': out_sample_sharpe,
                'Trades': out_sample_trades
            })

        print(separator)

        # FULL PERIOD PERFORMANCE
        print("\nFULL PERIOD PERFORMANCE:")
        print(separator)
        print(header)
        print(separator)

        for name, params in strategies.items():
            short = params['short_sma']
            long = params['long_sma']

            # Calculate full period metrics
            full_daily_pnl = data[f'Daily_PnL_{name}']
            full_cumulative_pnl = full_daily_pnl.sum()
            
            # Calculate average position size for full period
            avg_pos_size = data[f'Position_Size_{name}'].mean()
            max_pos_size = data[f'Position_Size_{name}'].max()

            # Calculate Sharpe ratio (annualized)
            if full_daily_pnl.std() > 0:
                full_sharpe = full_daily_pnl.mean() / full_daily_pnl.std() * np.sqrt(252)
            else:
                full_sharpe = 0

            # Count full period trades
            full_trades = data[f'Position_Change_{name}'].sum()

            # Create row text
            row = f"{name:<10} | {short:>4}/{long:<5} | ${full_cumulative_pnl:>10,.2f} | {avg_pos_size:>8.2f} | {full_sharpe:>6.3f} | {full_trades:>6}"

            # Print row
            print(row)

            # Store the data
            performance_data.append({
                'Period': 'Full',
                'Strategy': name,
                'Short_SMA': short,
                'Long_SMA': long,
                'PnL': full_cumulative_pnl,
                'Avg_Position': avg_pos_size,
                'Max_Position': max_pos_size,
                'Sharpe': full_sharpe,
                'Trades': full_trades
            })
            
            # Additional metrics for the best strategy
            if name == 'Best':
                print(f"\nAdditional metrics for Best strategy:")
                print(f"Maximum position size: {max_pos_size:.2f} contracts")
                print(f"Average ATR value: {data['ATR_Best'].mean():.4f}")
                
                # Calculate drawdown
                peak = data[f'Cumulative_PnL_{name}'].cummax()
                drawdown = data[f'Cumulative_PnL_{name}'] - peak
                max_drawdown = drawdown.min()
                
                print(f"Maximum drawdown: ${max_drawdown:.2f}")
                
                # Calculate win rate using trimmed data
                daily_win_rate = (data[f'Daily_PnL_{name}'] > 0).mean() * 100
                print(f"Daily win rate: {daily_win_rate:.2f}%")

        print(separator)

        return data_for_evaluation  # Return the trimmed data instead of full data

    def bimonthly_out_of_sample_comparison(data, best_short_sma, best_long_sma, top_medoids, min_sharpe=0.2, 
                                    big_point_value=big_point_value, slippage=slippage,
                                    capital=capital, atr_period=atr_period):
        """
        Compare bimonthly (2-month) performance between the best Sharpe strategy and a portfolio of top medoids
        using ATR-based position sizing.
        
        Parameters:
        data: DataFrame with market data
        best_short_sma: int - The short SMA period for the best Sharpe strategy
        best_long_sma: int - The long SMA period for the best Sharpe strategy
        top_medoids: list - List of top medoids, each as (short_sma, long_sma, sharpe, trades)
        min_sharpe: float - Minimum Sharpe ratio threshold for medoids to be included
        big_point_value: float - Big point value for the futures contract
        dynamic_slippage: float - Slippage value in price units
        capital: float - Capital allocation for position sizing
        atr_period: int - Period for ATR calculation
        """
        print(f"\n----- BIMONTHLY OUT-OF-SAMPLE COMPARISON -----")
        print(f"Best Sharpe: ({best_short_sma}/{best_long_sma})")
        print(f"Using ATR-based position sizing (Capital: ${capital:,}, ATR Period: {atr_period})")
        
        # Handle the case where top_medoids is None
        if top_medoids is None:
            print("No medoids provided. Comparison cannot be performed.")
            return None
        
        # Check if top_medoids is a single tuple (rather than a list of tuples)
        if isinstance(top_medoids, tuple) and len(top_medoids) >= 3:
            # Convert single tuple to a list with one tuple
            medoids_list = [top_medoids]
        else:
            # Assume it's already a list
            medoids_list = top_medoids
        
        # Take at most 3 medoids and filter by minimum Sharpe
        filtered_medoids = []
        for m in medoids_list[:3]:
            # Check if we can access the required elements (support numpy arrays, tuples, and lists)
            try:
                # Extract Sharpe ratio and check if it meets the threshold
                short_sma = m[0]
                long_sma = m[1]
                sharpe = float(m[2])  # Convert to float to handle numpy types
                trades = m[3]
                
                if sharpe >= min_sharpe:
                    filtered_medoids.append(m)
            except (IndexError, TypeError) as e:
                print(f"Error processing medoid: {e}")
        
        if not filtered_medoids:
            print(f"No medoids have a Sharpe ratio >= {min_sharpe}. Comparison cannot be performed.")
            return None
        
        print(f"Creating portfolio of {len(filtered_medoids)} medoids with Sharpe ratio >= {min_sharpe}:")
        for i, medoid in enumerate(filtered_medoids, 1):
            print(f"Medoid {i}: ({int(medoid[0])}/{int(medoid[1])}) - Sharpe: {float(medoid[2]):.4f}")
        
        # Download historical data if not already provided
        if data is None:
            # Load data from local file
            print(f"Loading {TICKER} data from local files...")
            data_file = find_futures_file(SYMBOL, DATA_DIR)
            if not data_file:
                print(f"Error: No data file found for {TICKER} in {DATA_DIR}")
                exit(1)
            
            print(f"Found data file: {os.path.basename(data_file)}")
            print(f"File size: {os.path.getsize(data_file)} bytes")

            # Load the data from local file
            all_data = read_ts.read_ts_ohlcv_dat(data_file)
            data_obj = all_data[0]
            ohlc_data = data_obj.data.copy()

            # Convert the OHLCV data to the format expected by the strategy
            data = ohlc_data.rename(columns={
                'datetime': 'Date',
                'open': 'Open',
                'high': 'High',
                'low': 'Low',
                'close': 'Close',
                'volume': 'Volume'
            })
            data.set_index('Date', inplace=True)
            
            # Filter data to match the date range if specified in input.py
            if START_DATE and END_DATE:
                data = data[(data.index >= pd.to_datetime(START_DATE)) & (data.index <= pd.to_datetime(END_DATE))]
                print(f"Filtered data to date range: {START_DATE} to {END_DATE}")
        
        # Create strategies
        strategies = {
            'Best': {'short_sma': best_short_sma, 'long_sma': best_long_sma}
        }
        
        # Add filtered medoids
        for i, medoid in enumerate(filtered_medoids, 1):
            strategies[f'Medoid_{i}'] = {'short_sma': int(medoid[0]), 'long_sma': int(medoid[1])}
        
        # Apply each strategy to the data
        for name, params in strategies.items():
            strategy = SMAStrategy(
                short_sma=params['short_sma'],
                long_sma=params['long_sma'],
                big_point_value=big_point_value,
                slippage=slippage,
                capital=capital,
                atr_period=atr_period
            )
            
            # Apply the strategy
            data = strategy.apply_strategy(
                data.copy(),
                strategy_name=name
            )
        
        # Get the out-of-sample split date
        split_index = int(len(data) * TRAIN_TEST_SPLIT)
        split_date = data.index[split_index]
        print(f"Out-of-sample period starts on: {split_date.strftime('%Y-%m-%d')}")
        
        # Get out-of-sample data
        oos_data = data.iloc[split_index:].copy()
        
        # Add a year and bimonthly period columns for grouping (each year has 6 bimonthly periods)
        oos_data['year'] = oos_data.index.year.astype(int)
        oos_data['bimonthly'] = ((oos_data.index.month - 1) // 2 + 1).astype(int)
        
        # Create simplified period labels with just the start month (YYYY-MM)
        oos_data['period_label'] = oos_data.apply(
            lambda row: f"{int(row['year'])}-{int((row['bimonthly'] - 1) * 2 + 1):02d}",
            axis=1
        )
        
        # Create a DataFrame to store bimonthly Sharpe ratios
        bimonthly_sharpe = []
        
        # Group by year and bimonthly period, calculate Sharpe ratio for each period
        for period_label, group in oos_data.groupby('period_label'):
            # Skip periods with too few trading days
            if len(group) < 10:
                continue
                
            # Create a bimonthly result entry
            year, start_month = period_label.split('-')
            year = int(year)
            start_month = int(start_month)
            
            bimonthly_result = {
                'period_label': period_label,
                'date': pd.Timestamp(year=year, month=start_month, day=15),  # Middle of first month in period
                'trading_days': len(group),
            }
            
            # Calculate Sharpe ratio for each strategy in this period
            for name in strategies.keys():
                # Get returns for this strategy in this period
                returns = group[f'Daily_PnL_{name}']
                
                # Calculate Sharpe ratio (annualized)
                if len(returns) > 1 and returns.std() > 0:
                    sharpe = returns.mean() / returns.std() * np.sqrt(252)
                else:
                    sharpe = 0
                    
                bimonthly_result[f'{name}_sharpe'] = sharpe
                bimonthly_result[f'{name}_return'] = returns.sum()  # Total P&L for the period
            
            # Calculate the normalized average of medoid Sharpe ratios
            medoid_sharpes = [bimonthly_result[f'Medoid_{i}_sharpe'] for i in range(1, len(filtered_medoids) + 1)]
            bimonthly_result['Avg_Medoid_sharpe'] = sum(medoid_sharpes) / len(filtered_medoids)
            
            # Calculate the normalized average of medoid returns
            medoid_returns = [bimonthly_result[f'Medoid_{i}_return'] for i in range(1, len(filtered_medoids) + 1)]
            bimonthly_result['Avg_Medoid_return'] = sum(medoid_returns) / len(filtered_medoids)
            
            bimonthly_sharpe.append(bimonthly_result)
        
        # Convert to DataFrame
        bimonthly_sharpe_df = pd.DataFrame(bimonthly_sharpe)
        
        # Sort the DataFrame by date for proper chronological display
        if not bimonthly_sharpe_df.empty:
            bimonthly_sharpe_df = bimonthly_sharpe_df.sort_values('date')
        
        # Add rounded values to dataframe for calculations
        bimonthly_sharpe_df['Best_sharpe_rounded'] = np.round(bimonthly_sharpe_df['Best_sharpe'], 2)
        bimonthly_sharpe_df['Avg_Medoid_sharpe_rounded'] = np.round(bimonthly_sharpe_df['Avg_Medoid_sharpe'], 2)
        
        # Print detailed comparison of Sharpe ratios
        print("\nDetailed Sharpe ratio comparison by period:")
        print(f"{'Period':<12} | {'Best Sharpe':>12} | {'Medoid Portfolio':>16} | {'Difference':>12} | {'Portfolio Wins':<14}")
        print("-" * 80)
        
        for idx, row in bimonthly_sharpe_df.iterrows():
            period = row['period_label']
            best_sharpe = row['Best_sharpe']
            avg_medoid_sharpe = row['Avg_Medoid_sharpe']
            best_rounded = row['Best_sharpe_rounded']
            avg_medoid_rounded = row['Avg_Medoid_sharpe_rounded']
            
            diff = avg_medoid_sharpe - best_sharpe
            portfolio_wins = avg_medoid_sharpe > best_sharpe
            
            print(f"{period:<12} | {best_sharpe:12.6f} | {avg_medoid_sharpe:16.6f} | {diff:12.6f} | {portfolio_wins!s:<14}")
        
        # Calculate win rate using raw values
        portfolio_wins = sum(bimonthly_sharpe_df['Avg_Medoid_sharpe'] > bimonthly_sharpe_df['Best_sharpe'])
        total_periods = len(bimonthly_sharpe_df)
        win_percentage = (portfolio_wins / total_periods) * 100 if total_periods > 0 else 0
        
        # Calculate win rate using rounded values (for alternative comparison)
        rounded_wins = sum(bimonthly_sharpe_df['Avg_Medoid_sharpe_rounded'] > bimonthly_sharpe_df['Best_sharpe_rounded'])
        rounded_win_percentage = (rounded_wins / total_periods) * 100 if total_periods > 0 else 0
        
        print(f"\nBimonthly periods analyzed: {total_periods}")
        print(f"Medoid Portfolio Wins: {portfolio_wins} of {total_periods} periods ({win_percentage:.2f}%)")
        print(f"Using rounded values (2 decimal places): {rounded_wins} of {total_periods} periods ({rounded_win_percentage:.2f}%)")
        
        # Create a bar plot to compare bimonthly Sharpe ratios
        plt.figure(figsize=(14, 8))
        
        # Set up x-axis dates
        x = np.arange(len(bimonthly_sharpe_df))
        width = 0.35  # Width of the bars
        
        # Create bars
        plt.bar(x - width/2, bimonthly_sharpe_df['Best_sharpe'], width, 
            label=f'Best Sharpe ({best_short_sma}/{best_long_sma})', color='blue')
        plt.bar(x + width/2, bimonthly_sharpe_df['Avg_Medoid_sharpe'], width, 
            label=f'Medoid Portfolio ({len(filtered_medoids)} strategies)', color='green')
        
        # Add a horizontal line at Sharpe = 0
        plt.axhline(y=0, color='black', linestyle='-', alpha=0.3)
        
        # Create medoid description for the title
        medoid_desc = ", ".join([f"({int(m[0])}/{int(m[1])})" for m in filtered_medoids])
        
        # Customize the plot - using rounded win percentage instead of raw
        plt.title(f'{SYMBOL} Kmeans Bimonthly Sharpe Ratio Comparison (Out-of-Sample Period)\n' + 
                f'Medoid Portfolio [{medoid_desc}] outperformed {rounded_win_percentage:.2f}% of the time', 
                fontsize=14)
        plt.xlabel('Bimonthly Period (Start Month)', fontsize=12)
        plt.ylabel('Sharpe Ratio (Annualized)', fontsize=12)
        
        # Simplified x-tick labels with just the period start month
        plt.xticks(x, bimonthly_sharpe_df['period_label'], rotation=45)
        
        plt.grid(axis='y', linestyle='--', alpha=0.7)
        
        # Create legend with both strategies - moved to bottom to avoid overlap with title
        plt.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=2,
                frameon=True, fancybox=True, framealpha=0.9, fontsize=10)
        
        # Add a text box with rounded win percentage instead of raw - moved to right side
        plt.annotate(f'Medoid Portfolio Win Rate: {rounded_win_percentage:.2f}%\n'
                    f'({rounded_wins} out of {total_periods} periods)\n'
                    f'Portfolio: {len(filtered_medoids)} medoids with Sharpe ≥ {min_sharpe}\n'
                    f'ATR-Based Position Sizing (${capital:,}, {atr_period} days)',
                    xy=(0.7, 0.95), xycoords='axes fraction',
                    bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="gray", alpha=0.8),
                    fontsize=12)
        
        plt.tight_layout(rect=[0, 0, 1, 0.95])  # Add extra space at the bottom for the legend
        save_plot(f'{SYMBOL}_KMeans_Bimonthly_Comparison.png')
        
        
        # Save win percentage and medoid parameters to Excel file
        try:
            import openpyxl
            
            # Path to the Excel file
            excel_file = r"C:\Users\Admin\Documents\darbas\Results.xlsx"
            
            # Check if file exists
            if not os.path.exists(excel_file):
                print(f"Excel file not found at: {excel_file}")
                return bimonthly_sharpe_df
                
            print(f"Updating Excel file with K-means results for {SYMBOL}...")
            
            # Load the workbook
            wb = openpyxl.load_workbook(excel_file)
            
            # Get the active sheet
            sheet = wb.active
            
            # Find the row with the ticker symbol or the first empty row
            row = 3  # Start from row 3 (assuming rows 1-2 have headers)
            ticker_row = None
            
            while True:
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value == SYMBOL:
                    # Found the ticker symbol
                    ticker_row = row
                    break
                elif cell_value is None:
                    # Found an empty row
                    ticker_row = row
                    # Write the ticker symbol in column A
                    sheet.cell(row=ticker_row, column=1).value = SYMBOL
                    break
                row += 1
            
            # Round the win percentage to one decimal place
            rounded_win_percentage_1dp = round(rounded_win_percentage, 1)
            
            # Write the win percentage in column B (K-means)
            sheet.cell(row=ticker_row, column=2).value = rounded_win_percentage_1dp
            
            # Write the medoid parameters in the respective cluster columns
            # K-means clusters start at column E (5)
            for i, medoid in enumerate(filtered_medoids):
                if i >= 3:  # Only use up to 3 clusters
                    break
                    
                # Calculate column index: E=5, F=6, G=7 for Cluster1, Cluster2, Cluster3
                column_idx = 5 + i
                
                # Format as "short/long" (e.g., "5/20")
                param_value = f"{int(medoid[0])}/{int(medoid[1])}"
                
                # Write to Excel
                sheet.cell(row=ticker_row, column=column_idx).value = param_value
            
            # Write the best Sharpe parameters in column M (13)
            best_sharpe_params = f"{best_short_sma}/{best_long_sma}"
            sheet.cell(row=ticker_row, column=13).value = best_sharpe_params
            
            # Save the workbook
            wb.save(excel_file)
            
            print(f"Excel file updated successfully. Added {SYMBOL} with K-means win rate {rounded_win_percentage_1dp}% in row {ticker_row}")
            print(f"Added best Sharpe parameters {best_sharpe_params} in column M")
            
        except Exception as e:
            print(f"Error updating Excel file: {e}")
        
        # Return the bimonthly Sharpe ratio data
        return bimonthly_sharpe_df

    def analyze_full_oos_performance(data, best_short_sma, best_long_sma, top_medoids,
                               big_point_value=1, slippage=0, capital=1000000, atr_period=14):
        """
        Plot a comparison between the best strategy and a portfolio of top 3 medoids.
        Only shows the P&L comparison plot without the price/SMA indicators.
        Also writes TRUE/FALSE to Excel column T based on whether the Portfolio
        outperforms the Best strategy in terms of out-of-sample Sharpe ratio.
        
        Parameters:
        data: DataFrame with market data that already has strategies applied
        best_short_sma: int - The short SMA period for the best strategy
        best_long_sma: int - The long SMA period for the best strategy
        top_medoids: list - List of top medoids by Sharpe ratio
        big_point_value: float - Big point value for the futures contract
        slippage: float - Slippage value in price units
        capital: float - Capital allocation for position sizing
        atr_period: int - Period for ATR calculation
        """
        print(f"\n----- FULL OUT-OF-SAMPLE PERFORMANCE ANALYSIS -----")
        print(f"Using Best Strategy: Short SMA: {best_short_sma}, Long SMA: {best_long_sma}")
        print(f"Creating portfolio from top 3 medoids")
        
        # Calculate split index for in-sample/out-of-sample
        split_index = int(len(data) * TRAIN_TEST_SPLIT)
        split_date = data.index[split_index]
        
        # Create dictionary of all strategies to calculate portfolio
        all_strategies = {
            'Best': {'short_sma': best_short_sma, 'long_sma': best_long_sma}
        }
        
        # Add top 3 medoids to strategies dictionary for calculation
        medoid_strategies = []
        for i, medoid in enumerate(top_medoids[:3], 1):
            strategy_name = f'Medoid {i}'  # Use space instead of underscore to match existing columns
            medoid_strategies.append(strategy_name)
            all_strategies[strategy_name] = {
                'short_sma': int(medoid[0]),
                'long_sma': int(medoid[1]),
                'original_sharpe': float(medoid[2]),
                'original_trades': int(medoid[3])
            }
            print(f"Added {strategy_name} to portfolio: SMA({int(medoid[0])}/{int(medoid[1])}) - Original Sharpe: {float(medoid[2]):.4f}")
        
        # Create the portfolio PnL by averaging the medoid strategies
        print("Calculating portfolio PnL as average of medoid strategies...")
        
        # Check if we have enough data to create a portfolio
        if len(medoid_strategies) == 0:
            print("Error: No medoid strategies found to create portfolio")
            return data
        
        # Create portfolio daily PnL series
        portfolio_daily_pnl = pd.Series(0, index=data.index)
        valid_strategies = 0
        
        # Print available columns for debugging
        print("\nAvailable columns that might contain Daily PnL data:")
        pnl_cols = [col for col in data.columns if 'Daily_PnL' in col]
        for col in pnl_cols:
            print(f"  - {col}")
        
        for strategy in medoid_strategies:
            # Try both naming conventions
            daily_pnl_col = f'Daily_PnL_{strategy}'
            
            if daily_pnl_col in data.columns:
                portfolio_daily_pnl += data[daily_pnl_col]
                valid_strategies += 1
                print(f"Added {daily_pnl_col} to portfolio")
            else:
                print(f"Warning: Column {daily_pnl_col} not found")
        
        # If no valid strategies found, we can't create a portfolio
        if valid_strategies == 0:
            print("Error: No valid strategy data found for portfolio. Plotting only Best strategy.")
            # Set portfolio to same as Best for plotting purposes
            portfolio_daily_pnl = data['Daily_PnL_Best'].copy()
            valid_strategies = 1
        else:
            # Calculate average (divide by number of valid strategies)
            portfolio_daily_pnl = portfolio_daily_pnl / valid_strategies
        
        # Calculate portfolio cumulative PnL
        portfolio_cumulative_pnl = portfolio_daily_pnl.cumsum()
        
        # Create the performance visualization with only the P&L plot
        plt.figure(figsize=(14, 8))
        
        # Store out-of-sample Sharpe ratios
        oos_sharpe_ratios = {}
        
        # Plot Best strategy and Portfolio
        colors = ['blue', 'green']
        strategies_to_plot = {
            'Best': {'data': data[f'Cumulative_PnL_Best'], 'daily': data[f'Daily_PnL_Best'], 'color': colors[0]},
            'Portfolio': {'data': portfolio_cumulative_pnl, 'daily': portfolio_daily_pnl, 'color': colors[1]}
        }
        
        # Get the starting values for each line to normalize them to zero
        for name, strategy_data in strategies_to_plot.items():
            # Get the first value to normalize
            start_value = strategy_data['data'].iloc[0] if not strategy_data['data'].empty else 0
            
            # Calculate in-sample and out-of-sample Sharpe ratios
            in_sample_returns = strategy_data['daily'].iloc[:split_index]
            out_sample_returns = strategy_data['daily'].iloc[split_index:]
            
            # Calculate annualized Sharpe ratios
            in_sample_sharpe = in_sample_returns.mean() / in_sample_returns.std() * np.sqrt(252) if in_sample_returns.std() > 0 else 0
            out_sample_sharpe = out_sample_returns.mean() / out_sample_returns.std() * np.sqrt(252) if out_sample_returns.std() > 0 else 0
            
            # Store out-of-sample Sharpe ratio
            oos_sharpe_ratios[name] = out_sample_sharpe
            
            # Create label with strategy info and Sharpe ratios
            if name == 'Best':
                params_str = f"({best_short_sma}/{best_long_sma})"
            else:
                # Create a string with all the medoid parameters
                medoid_params = []
                for i, medoid in enumerate(top_medoids[:3], 1):
                    medoid_params.append(f"{int(medoid[0])}/{int(medoid[1])}")
                params_str = f"({', '.join(medoid_params)})"
            
            label = f'{name} {params_str} [IS Sharpe: {in_sample_sharpe:.2f}, OOS Sharpe: {out_sample_sharpe:.2f}]'
            
            # Normalize the data to start at zero
            normalized_pnl = strategy_data['data'] - start_value
            
            # Plot full period P&L (normalized)
            plt.plot(data.index, normalized_pnl, label=label, color=strategy_data['color'])
            
            # Plot out-of-sample portion with thicker line
            plt.plot(data.index[split_index:], normalized_pnl.iloc[split_index:],
                    color=strategy_data['color'], linewidth=2.5, alpha=0.7)
        
        plt.axvline(x=split_date, color='black', linestyle='--',
                    label=f'Train/Test Split ({int(TRAIN_TEST_SPLIT * 100)}%/{int((1 - TRAIN_TEST_SPLIT) * 100)}%)')
        plt.axhline(y=0.0, color='gray', linestyle='-', alpha=0.5, label='Break-even')
        plt.legend(loc='upper left')
        plt.title(f'{SYMBOL} Strategy Comparison: Best vs Portfolio of Medoids', fontsize=14)
        plt.ylabel('P&L ($)', fontsize=12)
        plt.grid(True, alpha=0.3)
        
        plt.tight_layout()
        save_plot(f'{SYMBOL}_KMeans_Full_OOS_Performance_Analysis.png')
        
        print(f"Full out-of-sample performance analysis plot saved to '{SYMBOL}_KMeans_Full_OOS_Performance_Analysis.png'")
        
        # Determine if Portfolio outperforms Best in terms of OOS Sharpe
        portfolio_wins = oos_sharpe_ratios['Portfolio'] > oos_sharpe_ratios['Best']
        print(f"Portfolio out-of-sample Sharpe: {oos_sharpe_ratios['Portfolio']:.4f}")
        print(f"Best out-of-sample Sharpe: {oos_sharpe_ratios['Best']:.4f}")
        print(f"Portfolio outperforms Best in OOS: {portfolio_wins}")
        
        # Save to Excel
        try:
            import openpyxl

            # Path to the Excel file
            excel_file = os.path.join(WORKING_DIR, "Results.xlsx")

            # Check if file exists
            if not os.path.exists(excel_file):
                print(f"Excel file not found at: {excel_file}")
                return data

            print(f"Updating Excel file with OOS comparison results for {SYMBOL}...")

            # Load the workbook
            wb = openpyxl.load_workbook(excel_file)

            # Get the active sheet
            sheet = wb.active

            # Find the row with the ticker symbol
            row = 3  # Start from row 3 (assuming rows 1-2 have headers)
            ticker_row = None

            while True:
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value == SYMBOL:
                    # Found the ticker symbol
                    ticker_row = row
                    break
                elif cell_value is None:
                    # Found an empty row
                    ticker_row = row
                    # Write the ticker symbol in column A
                    sheet.cell(row=ticker_row, column=1).value = SYMBOL
                    break
                row += 1

            # Write TRUE/FALSE in column T (20) based on whether Portfolio outperforms Best
            sheet.cell(row=ticker_row, column=20).value = "TRUE" if portfolio_wins else "FALSE"

            # Save the workbook
            wb.save(excel_file)

            print(f"Excel file updated successfully. Added {SYMBOL} with K-means portfolio comparison {portfolio_wins} in column T")

        except Exception as e:
            print(f"Error updating Excel file: {e}")
        
        return data

    # Main execution block
    # Set matplotlib backend explicitly
    matplotlib.use('Agg')  # Use non-interactive backend for headless environments
    
    print("Starting ATR-based SMA strategy analysis...")
    
    # Run the basic analysis first
    data, best_short, best_long, best_sharpe, best_trades = analyze_sma_results()
    
    if data is None:
        print("Error: Failed to load or analyze SMA results data.")
        exit(1)
    
    print("\nProceeding with cluster analysis...")
    
    # Run the cluster analysis to get medoids
    X_filtered, medoids, top_medoids, centroids, max_sharpe_point = cluster_analysis()
    
    if X_filtered is None or medoids is None:
        print("Error: Cluster analysis failed.")
        exit(1)
    
    # Re-sort top_medoids to ensure they're in the right order
    if top_medoids is not None:
        print("Re-sorting top medoids by Sharpe ratio...")
        top_medoids = sorted(top_medoids, key=lambda x: float(x[2]), reverse=True)
        for idx, medoid in enumerate(top_medoids, 1):
            print(f"Verified Medoid {idx}: Short SMA={int(medoid[0])}, Long SMA={int(medoid[1])}, "
                f"Sharpe={float(medoid[2]):.4f}, Trades={int(medoid[3])}")
    
    print("\nPlotting strategy performance...")
    
    # Plot strategy performance with the best parameters AND top medoids using ATR-based position sizing
    market_data = plot_strategy_performance(
        best_short, best_long, top_medoids,
        big_point_value=big_point_value,
        slippage=slippage,
        capital=capital,
        atr_period=atr_period
    )
    
    # Run the bimonthly out-of-sample comparison between best Sharpe and top medoids
    if top_medoids and len(top_medoids) > 0:
        print("\nPerforming bimonthly out-of-sample comparison...")
        bimonthly_sharpe_df = bimonthly_out_of_sample_comparison(
            market_data,
            best_short,
            best_long,
            top_medoids,  # Pass the entire top_medoids list
            big_point_value=big_point_value,
            slippage=slippage,
            capital=capital,
            atr_period=atr_period
        )
    else:
        print("No top medoids found. Cannot run bimonthly comparison.")
    
    # After your bimonthly comparison code
    if top_medoids and len(top_medoids) > 0:
        print("\nPerforming full out-of-sample performance analysis...")
        full_oos_results = analyze_full_oos_performance(
            market_data,
            best_short,
            best_long,
            top_medoids,
            big_point_value=big_point_value,
            slippage=slippage,
            capital=capital,
            atr_period=atr_period
        )
    else:
        print("No top medoids found. Cannot run full OOS analysis.")
    
    print("\nAnalysis complete! All plots and result files have been saved to the output directory.")
    print(f"Output directory: {output_dir}")

# Call the main function to execute the script
if __name__ == "__main__":
    main()