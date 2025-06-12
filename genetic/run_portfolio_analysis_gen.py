import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import glob
import json
import openpyxl
import warnings
import read_ts
from input_gen import *
from SMA_Strategy import SMAStrategy

# warnings.filterwarnings('ignore')

# Fallback defaults (in case Excel/data lookups fail)
#ATR_PERIOD      = 70
#TRADING_CAPITAL = 6000

def get_slippage_for_symbol(symbol, data_dir):
    """
    Read `sessions_slippages.xlsx` inside data_dir and return the numeric slippage
    corresponding to `symbol` (e.g. "@BO", "@AD", etc.).
    """
    excel_path = os.path.join(data_dir, "sessions_slippages.xlsx")
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Slippage Excel file not found at {excel_path!r}")

    df = pd.read_excel(excel_path)
    if df.shape[1] < 4:
        raise ValueError(f"Expected at least 4 columns in {excel_path!r}, but saw: {df.columns.tolist()}")

    df["SymbolUpper"] = df.iloc[:, 1].astype(str).str.upper()
    lookup = symbol.replace("=F", "").upper()
    matches = df[df["SymbolUpper"] == lookup]
    if matches.empty:
        raise ValueError(f"Symbol '{symbol}' not found in column B of {excel_path!r}")

    slippage_value = matches.iloc[0, 3]  # column D (index 3)
    if pd.isna(slippage_value) or not isinstance(slippage_value, (int, float)):
        raise ValueError(f"Invalid slippage for '{symbol}': {slippage_value!r}")

    return float(slippage_value)


def bimonthly_kmeans_vs_best_comparison(portfolio_df, save_plot):
    """
    Compare bimonthly (2-month) performance between K-means Portfolio and Best Strategy Portfolio.

    Parameters:
    portfolio_df: DataFrame with daily portfolio returns and cumulative P&L

    Returns:
    DataFrame with bimonthly comparison results
    """
    print(f"\n----- BIMONTHLY K-MEANS VS BEST STRATEGY COMPARISON -----")

    # Define the display window (same as your visualization function)
    start_plot = pd.to_datetime("2014-01-01")
    end_plot = pd.to_datetime("2024-12-31")

    # Slice to the plotting window
    df = portfolio_df.loc[start_plot:end_plot].copy()
    if df.empty:
        print(f"No data between {start_plot.date()} and {end_plot.date()}. Cannot perform comparison.")
        return None

    # Compute the split index for out-of-sample analysis
    split_idx = int(len(df) * TRAIN_TEST_SPLIT)
    split_date = df.index[split_idx]
    print(f"Out-of-sample period starts on: {split_date.strftime('%Y-%m-%d')}")

    # Get out-of-sample data only
    oos_data = df.iloc[split_idx:].copy()

    if oos_data.empty:
        print("No out-of-sample data available. Cannot perform comparison.")
        return None

    # Add time grouping columns
    oos_data['year'] = oos_data.index.year.astype(int)
    oos_data['bimonthly'] = ((oos_data.index.month - 1) // 2 + 1).astype(int)
    oos_data['period_label'] = oos_data.apply(
        lambda row: f"{int(row['year'])}-{int((row['bimonthly'] - 1) * 2 + 1):02d}", axis=1
    )

    # Compute bimonthly Sharpe ratios and returns
    bimonthly_results = []

    for period_label, group in oos_data.groupby('period_label'):
        if len(group) < 10:  # Skip periods with too few trading days
            continue

        year, start_month = period_label.split('-')
        year = int(year)
        start_month = int(start_month)

        entry = {
            'period_label': period_label,
            'date': pd.Timestamp(year=year, month=start_month, day=15),
            'trading_days': len(group)
        }

        # Calculate Sharpe ratio and returns for Best Strategy Portfolio
        best_returns = group["Portfolio_1_Best"]
        if len(best_returns) > 1 and best_returns.std() > 0:
            entry['Best_sharpe'] = best_returns.mean() / best_returns.std() * np.sqrt(252)
        else:
            entry['Best_sharpe'] = 0.0
        entry['Best_return'] = best_returns.sum()

        # Calculate Sharpe ratio and returns for K-means Portfolio
        kmeans_returns = group["Portfolio_2_Kmeans"]
        if len(kmeans_returns) > 1 and kmeans_returns.std() > 0:
            entry['Kmeans_sharpe'] = kmeans_returns.mean() / kmeans_returns.std() * np.sqrt(252)
        else:
            entry['Kmeans_sharpe'] = 0.0
        entry['Kmeans_return'] = kmeans_returns.sum()

        bimonthly_results.append(entry)

    bimonthly_df = pd.DataFrame(bimonthly_results)
    if bimonthly_df.empty:
        print("No valid bimonthly periods found. Cannot create comparison.")
        return None

    bimonthly_df = bimonthly_df.sort_values('date')

    # Round Sharpe ratios for comparison
    bimonthly_df['Best_sharpe_rounded'] = np.round(bimonthly_df['Best_sharpe'], 2)
    bimonthly_df['Kmeans_sharpe_rounded'] = np.round(bimonthly_df['Kmeans_sharpe'], 2)

    # Print detailed comparison
    print("\nDetailed Sharpe ratio comparison by period:")
    print(f"{'Period':<12} | {'Best Strategy':>15} | {'K-means Portfolio':>18} | {'Difference':>12} | {'K-means Wins':<13}")
    print("-" * 85)

    for idx, row in bimonthly_df.iterrows():
        period = row['period_label']
        best_sharpe = row['Best_sharpe']
        kmeans_sharpe = row['Kmeans_sharpe']
        diff = kmeans_sharpe - best_sharpe
        kmeans_wins = kmeans_sharpe > best_sharpe

        print(f"{period:<12} | {best_sharpe:15.6f} | {kmeans_sharpe:18.6f} | {diff:12.6f} | {kmeans_wins!s:<13}")

    # Calculate win rates
    total_periods = len(bimonthly_df)
    kmeans_wins = (bimonthly_df['Kmeans_sharpe_rounded'] > bimonthly_df['Best_sharpe_rounded']).sum()
    kmeans_win_rate = (kmeans_wins / total_periods) * 100 if total_periods > 0 else 0

    print(f"\n{'='*60}")
    print("K-MEANS VS BEST STRATEGY - BIMONTHLY WIN RATE:")
    print(f"{'='*60}")
    print(f"Total bimonthly periods analyzed: {total_periods}")
    print(f"K-means Portfolio wins: {kmeans_wins} of {total_periods} periods ({kmeans_win_rate:.2f}%)")
    print(f"Best Strategy wins: {total_periods - kmeans_wins} of {total_periods} periods ({100 - kmeans_win_rate:.2f}%)")
    print(f"{'='*60}")

    # Create the comparison plot
    plt.figure(figsize=(16, 8))
    x = np.arange(len(bimonthly_df))
    width = 0.35

    plt.bar(x - width/2, bimonthly_df['Best_sharpe'], width,
           label='Best Strategy Portfolio', color='blue', alpha=0.8)
    plt.bar(x + width/2, bimonthly_df['Kmeans_sharpe'], width,
           label='K-means Cluster Portfolio', color='green', alpha=0.8)

    plt.axhline(y=0, color='black', linestyle='-', alpha=0.3)

    plt.title(
        f"K-means Portfolio vs Best Strategy - Bimonthly Sharpe Ratio Comparison (Out-of-Sample)\n"
        f"K-means Portfolio outperformed {kmeans_win_rate:.2f}% of the time",
        fontsize=14
    )
    plt.xlabel('Bimonthly Period (Start Month)', fontsize=12)
    plt.ylabel('Sharpe Ratio (Annualized)', fontsize=12)

    plt.xticks(x, bimonthly_df['period_label'], rotation=45, ha='right')
    plt.grid(axis='y', linestyle='--', alpha=0.7)

    plt.legend(loc='upper center',
              bbox_to_anchor=(0.5, -0.15),
              ncol=2,
              frameon=True,
              fancybox=True,
              framealpha=0.9,
              fontsize=11)

    plt.annotate(
        f'K-means Portfolio Win Rate: {kmeans_win_rate:.2f}%\n'
        f'({kmeans_wins} of {total_periods} periods)\n'
        f'Out-of-Sample Analysis',
        xy=(0.7, 0.95),
        xycoords='axes fraction',
        bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="gray", alpha=0.8),
        fontsize=12
    )

    plt.tight_layout()
    save_plot("kmeans_vs_best_bimonthly_comparison.png")

    return bimonthly_df


def bimonthly_hierarchical_vs_best_comparison(portfolio_df, save_plot):
    """
    Compare bimonthly (2-month) performance between Hierarchical Portfolio and Best Strategy Portfolio.

    Parameters:
    portfolio_df: DataFrame with daily portfolio returns and cumulative P&L

    Returns:
    DataFrame with bimonthly comparison results
    """
    print(f"\n----- BIMONTHLY HIERARCHICAL VS BEST STRATEGY COMPARISON -----")

    # Define the display window (same as your visualization function)
    start_plot = pd.to_datetime("2014-01-01")
    end_plot = pd.to_datetime("2024-12-31")

    # Slice to the plotting window
    df = portfolio_df.loc[start_plot:end_plot].copy()
    if df.empty:
        print(f"No data between {start_plot.date()} and {end_plot.date()}. Cannot perform comparison.")
        return None

    # Compute the split index for out-of-sample analysis
    split_idx = int(len(df) * TRAIN_TEST_SPLIT)
    split_date = df.index[split_idx]
    print(f"Out-of-sample period starts on: {split_date.strftime('%Y-%m-%d')}")

    # Get out-of-sample data only
    oos_data = df.iloc[split_idx:].copy()

    if oos_data.empty:
        print("No out-of-sample data available. Cannot perform comparison.")
        return None

    # Add time grouping columns
    oos_data['year'] = oos_data.index.year.astype(int)
    oos_data['bimonthly'] = ((oos_data.index.month - 1) // 2 + 1).astype(int)
    oos_data['period_label'] = oos_data.apply(
        lambda row: f"{int(row['year'])}-{int((row['bimonthly'] - 1) * 2 + 1):02d}", axis=1
    )

    # Compute bimonthly Sharpe ratios and returns
    bimonthly_results = []

    for period_label, group in oos_data.groupby('period_label'):
        if len(group) < 10:  # Skip periods with too few trading days
            continue

        year, start_month = period_label.split('-')
        year = int(year)
        start_month = int(start_month)

        entry = {
            'period_label': period_label,
            'date': pd.Timestamp(year=year, month=start_month, day=15),
            'trading_days': len(group)
        }

        # Calculate Sharpe ratio and returns for Best Strategy Portfolio
        best_returns = group["Portfolio_1_Best"]
        if len(best_returns) > 1 and best_returns.std() > 0:
            entry['Best_sharpe'] = best_returns.mean() / best_returns.std() * np.sqrt(252)
        else:
            entry['Best_sharpe'] = 0.0
        entry['Best_return'] = best_returns.sum()

        # Calculate Sharpe ratio and returns for Hierarchical Portfolio
        hier_returns = group["Portfolio_3_Hierarchical"]
        if len(hier_returns) > 1 and hier_returns.std() > 0:
            entry['Hierarchical_sharpe'] = hier_returns.mean() / hier_returns.std() * np.sqrt(252)
        else:
            entry['Hierarchical_sharpe'] = 0.0
        entry['Hierarchical_return'] = hier_returns.sum()

        bimonthly_results.append(entry)

    bimonthly_df = pd.DataFrame(bimonthly_results)
    if bimonthly_df.empty:
        print("No valid bimonthly periods found. Cannot create comparison.")
        return None

    bimonthly_df = bimonthly_df.sort_values('date')

    # Round Sharpe ratios for comparison
    bimonthly_df['Best_sharpe_rounded'] = np.round(bimonthly_df['Best_sharpe'], 2)
    bimonthly_df['Hierarchical_sharpe_rounded'] = np.round(bimonthly_df['Hierarchical_sharpe'], 2)

    # Print detailed comparison
    print("\nDetailed Sharpe ratio comparison by period:")
    print(f"{'Period':<12} | {'Best Strategy':>15} | {'Hierarchical Portfolio':>23} | {'Difference':>12} | {'Hierarchical Wins':<17}")
    print("-" * 95)

    for idx, row in bimonthly_df.iterrows():
        period = row['period_label']
        best_sharpe = row['Best_sharpe']
        hier_sharpe = row['Hierarchical_sharpe']
        diff = hier_sharpe - best_sharpe
        hier_wins = hier_sharpe > best_sharpe

        print(f"{period:<12} | {best_sharpe:15.6f} | {hier_sharpe:23.6f} | {diff:12.6f} | {hier_wins!s:<17}")

    # Calculate win rates
    total_periods = len(bimonthly_df)
    hier_wins = (bimonthly_df['Hierarchical_sharpe_rounded'] > bimonthly_df['Best_sharpe_rounded']).sum()
    hier_win_rate = (hier_wins / total_periods) * 100 if total_periods > 0 else 0

    print(f"\n{'='*60}")
    print("HIERARCHICAL VS BEST STRATEGY - BIMONTHLY WIN RATE:")
    print(f"{'='*60}")
    print(f"Total bimonthly periods analyzed: {total_periods}")
    print(f"Hierarchical Portfolio wins: {hier_wins} of {total_periods} periods ({hier_win_rate:.2f}%)")
    print(f"Best Strategy wins: {total_periods - hier_wins} of {total_periods} periods ({100 - hier_win_rate:.2f}%)")
    print(f"{'='*60}")

    # Create the comparison plot
    plt.figure(figsize=(16, 8))
    x = np.arange(len(bimonthly_df))
    width = 0.35

    plt.bar(x - width/2, bimonthly_df['Best_sharpe'], width,
           label='Best Strategy Portfolio', color='blue', alpha=0.8)
    plt.bar(x + width/2, bimonthly_df['Hierarchical_sharpe'], width,
           label='Hierarchical Cluster Portfolio', color='purple', alpha=0.8)

    plt.axhline(y=0, color='black', linestyle='-', alpha=0.3)

    plt.title(
        f"Hierarchical Portfolio vs Best Strategy - Bimonthly Sharpe Ratio Comparison (Out-of-Sample)\n"
        f"Hierarchical Portfolio outperformed {hier_win_rate:.2f}% of the time",
        fontsize=14
    )
    plt.xlabel('Bimonthly Period (Start Month)', fontsize=12)
    plt.ylabel('Sharpe Ratio (Annualized)', fontsize=12)

    plt.xticks(x, bimonthly_df['period_label'], rotation=45, ha='right')
    plt.grid(axis='y', linestyle='--', alpha=0.7)

    plt.legend(loc='upper center',
              bbox_to_anchor=(0.5, -0.15),
              ncol=2,
              frameon=True,
              fancybox=True,
              framealpha=0.9,
              fontsize=11)

    plt.annotate(
        f'Hierarchical Portfolio Win Rate: {hier_win_rate:.2f}%\n'
        f'({hier_wins} of {total_periods} periods)\n'
        f'Out-of-Sample Analysis',
        xy=(0.7, 0.95),
        xycoords='axes fraction',
        bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="gray", alpha=0.8),
        fontsize=12
    )

    plt.tight_layout()
    save_plot("hierarchical_vs_best_bimonthly_comparison.png")

    return bimonthly_df


def portfolio_analysis_main():
    """
    Main function to analyze and compare three portfolios.
    For each symbol, we:
      1) Read slippage from Excel via get_slippage_for_symbol(...)
      2) Load .dat data to discover big_point_value from data_obj
      3) Use those per-symbol values when instantiating SMAStrategy
    """

    # 1) Paths and setup
    WORKING_DIR = '.'  # Assuming the script is run from the project root.
    DATA_DIR     = os.path.join(WORKING_DIR, "data")
    excel_file   = os.path.join(WORKING_DIR, "Results.xlsx")

    # Output folder for plots & CSVs
    output_dir = os.path.join(WORKING_DIR, 'output1')
    os.makedirs(output_dir, exist_ok=True)

    def save_plot(plot_name):
        """Save the current Matplotlib figure into output_dir."""
        plt.savefig(os.path.join(output_dir, plot_name), dpi=300, bbox_inches='tight')
        plt.show()
        plt.close()

    def find_futures_file(symbol, data_dir):
        """Find the first .dat file matching @SYMBOL patterns."""
        patterns = [
            f"*@{symbol}_*.dat",
            f"*_@{symbol}_*.dat",
            f"*_{symbol}_*.dat",
            f"*@{symbol}*.dat",
        ]
        for patt in patterns:
            matches = glob.glob(os.path.join(data_dir, patt))
            if matches:
                return matches[0]
        return None

    def parse_strategy_params(param_str):
        """Parse "short/long" (e.g. '5/20') into integers."""
        if isinstance(param_str, str) and '/' in param_str:
            try:
                s, l = param_str.split('/')
                return int(s), int(l)
            except:
                return None, None
        return None, None

    def load_excel_strategies():
        """
        Read your Results.xlsx (in WORKING_DIR) to find, for each symbol:
          - best_sharpe   (column M)
          - kmeans_clusters  (columns E/F/G)
          - hierarchical_clusters (columns I/J/K)
        """
        print(f"Loading strategy parameters from {excel_file}...")
        if not os.path.exists(excel_file):
            print(f"Excel file not found: {excel_file!r}")
            return {}

        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active

        strategies = {}
        row = 3  # assume data begins on row 3
        while True:
            symbol = sheet.cell(row=row, column=1).value
            if symbol is None:
                break

            strategies[symbol] = {
                'best_sharpe':         None,
                'kmeans_clusters':     [],
                'hierarchical_clusters': []
            }

            # Best Sharpe (column 13 = M)
            best_params = sheet.cell(row=row, column=13).value
            if best_params:
                s, l = parse_strategy_params(best_params)
                if s and l:
                    strategies[symbol]['best_sharpe'] = (s, l)

            # K-means cluster (columns E=5, F=6, G=7)
            for col in (5, 6, 7):
                param = sheet.cell(row=row, column=col).value
                if param:
                    s, l = parse_strategy_params(param)
                    if s and l:
                        strategies[symbol]['kmeans_clusters'].append((s, l))

            # Hierarchical (columns I=9, J=10, K=11)
            for col in (9, 10, 11):
                param = sheet.cell(row=row, column=col).value
                if param:
                    s, l = parse_strategy_params(param)
                    if s and l:
                        strategies[symbol]['hierarchical_clusters'].append((s, l))

            row += 1

        wb.close()

        total = len(strategies)
        best_ct = sum(1 for v in strategies.values() if v['best_sharpe'])
        k_ct    = sum(1 for v in strategies.values() if v['kmeans_clusters'])
        h_ct    = sum(1 for v in strategies.values() if v['hierarchical_clusters'])
        print(f"Loaded strategies for {total} symbols: "
              f"{best_ct} best-Sharpe, {k_ct} with K-means, {h_ct} with hierarchical.")
        return strategies

    def load_and_prepare_data(symbol, strategies_dict):
        """
        Load this symbol's raw .dat via read_ts_ohlcv_dat, rename columns,
        apply any warm-up if START_DATE/END_DATE are set, and return:
          - data (a DataFrame with columns [Open,High,Low,Close,Volume,...])
          - original_start_idx (index where we begin evaluating after warm-up)
        """
        print(f"\nLoading data for {symbol}...")
        data_file = find_futures_file(symbol, DATA_DIR)
        if not data_file:
            print(f"  No data file found for {symbol}")
            return None, None
        print(f"  Found data file: {os.path.basename(data_file)}")

        try:
            all_data = read_ts.read_ts_ohlcv_dat(data_file)  # your existing data loader
            data_obj = all_data[0]
            df = data_obj.data.copy().rename(columns={
                'datetime': 'Date',
                'open':     'Open',
                'high':     'High',
                'low':      'Low',
                'close':    'Close',
                'volume':   'Volume'
            })
            df.set_index('Date', inplace=True)
        except Exception as e:
            print(f"  Error loading data: {e}")
            return None, None

        # Determine longest SMA needed:
        all_strats = []
        if strategies_dict[symbol]['best_sharpe']:
            all_strats.append(strategies_dict[symbol]['best_sharpe'])
        all_strats.extend(strategies_dict[symbol]['kmeans_clusters'])
        all_strats.extend(strategies_dict[symbol]['hierarchical_clusters'])
        if not all_strats:
            print(f"  No valid strategies for {symbol}")
            return None, None
        max_sma = max(max(s, l) for s, l in all_strats)

        original_start_idx = None
        if START_DATE and END_DATE:
            # Suppose you want 3×longest_SMA days of warm-up:
            warm_up_days = max_sma * 3
            sd = pd.to_datetime(START_DATE)
            ed = pd.to_datetime(END_DATE)
            adj_start = sd - pd.Timedelta(days=warm_up_days)
            sub = df[(df.index >= adj_start) & (df.index <= ed)]
            if not sub.empty:
                idx0 = sub.index.get_indexer([sd], method='nearest')[0]
                original_start_idx = idx0
                print(f"  Added {warm_up_days} days warm-up before {START_DATE}")
                df = sub
            else:
                print(f"  No data in extended range for {symbol}")
                return None, None

        print(f"  Data shape: {df.shape}, Range: {df.index.min()} to {df.index.max()}")
        return df, original_start_idx

    def apply_strategies_to_data(df, symbol, strategies_dict, original_start_idx):
        """
        Apply best / kmeans / hierarchical SMAs for a single symbol,
        using that symbol's big_point_value and slippage (fetched here).
        Return the processed DataFrame (with all Daily_PnL_ and Cumulative_PnL_ columns)
        plus a dictionary of which strategy‐names were applied.
        """
        print(f"  Applying strategies for {symbol}...")

        # 1) Immediately fetch this symbol's slippage from Excel:
        try:
            slippage = get_slippage_for_symbol(symbol, DATA_DIR)
            print(f"    Using slippage from Excel: {slippage}")
        except Exception as e:
            print(f"    ERROR fetching slippage for {symbol}: {e}")
            print("    Falling back to slippage = 0.0")
            slippage = 0.0

        # 2) Also read big_point_value directly from the data object (if we kept it):
        #    We can re‐open read_ts_ohlcv_dat or rely on 'df' having a hidden attribute.
        #    Easiest: re-load the data object briefly to fetch big_point_value:
        data_file = find_futures_file(symbol, DATA_DIR)
        all_data  = read_ts.read_ts_ohlcv_dat(data_file)
        data_obj  = all_data[0]
        big_point_value = data_obj.big_point_value
        print(f"    Using big_point_value from data: {big_point_value}")
        print(f"    Using ATR_PERIOD: {ATR_PERIOD}")
        print(f"    Using TRADING_CAPITAL: {TRADING_CAPITAL}")

        applied = {'best': [], 'kmeans': [], 'hierarchical': []}
        count = 0

        # Best Sharpe:
        if strategies_dict[symbol]['best_sharpe']:
            s, l = strategies_dict[symbol]['best_sharpe']
            name = f"Best_{symbol}"
            try:
                strategy = SMAStrategy(
                    short_sma=s,
                    long_sma=l,
                    big_point_value=big_point_value,
                    slippage=       slippage,
                    capital=        TRADING_CAPITAL,
                    atr_period=     ATR_PERIOD
                )
                df = strategy.apply_strategy(df.copy(), strategy_name=name)
                applied['best'].append(name)
                count += 1
                print(f"    Applied Best: SMA({s}/{l})")
            except Exception as e:
                print(f"    Error applying Best SMA({s}/{l}): {e}")

        # K-means cluster SMAs:
        for i, (s, l) in enumerate(strategies_dict[symbol]['kmeans_clusters'], start=1):
            name = f"Kmeans_{i}_{symbol}"
            try:
                strategy = SMAStrategy(
                    short_sma=s,
                    long_sma=l,
                    big_point_value=big_point_value,
                    slippage=       slippage,
                    capital=        TRADING_CAPITAL,
                    atr_period=     ATR_PERIOD
                )
                df = strategy.apply_strategy(df.copy(), strategy_name=name)
                applied['kmeans'].append(name)
                count += 1
                print(f"    Applied K-means {i}: SMA({s}/{l})")
            except Exception as e:
                print(f"    Error applying K-means {i}: {e}")

        # Hierarchical cluster SMAs:
        for i, (s, l) in enumerate(strategies_dict[symbol]['hierarchical_clusters'], start=1):
            name = f"Hierarchical_{i}_{symbol}"
            try:
                strategy = SMAStrategy(
                    short_sma=s,
                    long_sma=l,
                    big_point_value=big_point_value,
                    slippage=       slippage,
                    capital=        TRADING_CAPITAL,
                    atr_period=     ATR_PERIOD
                )
                df = strategy.apply_strategy(df.copy(), strategy_name=name)
                applied['hierarchical'].append(name)
                count += 1
                print(f"    Applied Hierarchical {i}: SMA({s}/{l})")
            except Exception as e:
                print(f"    Error applying Hierarchical {i}: {e}")

        # 3) Trim warm-up (reset cumulative P&L) if needed
        if original_start_idx is not None and original_start_idx > 0:
            print(f"    Trimming warm-up (original length = {len(df)})")
            all_strat_names = (
                applied['best'] +
                applied['kmeans'] +
                applied['hierarchical']
            )
            for strat_name in all_strat_names:
                daily_col = f"Daily_PnL_{strat_name}"
                cum_col   = f"Cumulative_PnL_{strat_name}"
                if cum_col in df.columns and daily_col in df.columns:
                    df[f"{cum_col}_Original"] = df[cum_col].copy()
                    df[cum_col] = 0.0
                    df.loc[df.index[original_start_idx:], cum_col] = (
                        df.loc[df.index[original_start_idx:], daily_col].cumsum()
                    )

            df = df.iloc[original_start_idx:].copy()
            print(f"    Trimmed data shape: {df.shape}")

        print(f"    Successfully applied {count} strategies")
        return df, applied

    def build_portfolios(all_data_dict, all_applied_dict):
        """
        Build three portfolio streams (best, kmeans, hierarchical).
        Return a DataFrame with daily & cumulative P&L columns.
        """
        print("\n----- BUILDING PORTFOLIOS -----")
        all_dates = []
        coverage = {}

        for sym, df in all_data_dict.items():
            if df is not None:
                for d in df.index:
                    all_dates.append(d)
                    coverage.setdefault(d, []).append(sym)

        if not all_dates:
            print("No dates found. Aborting.")
            return None

        min_req = max(1, len(all_data_dict) // 3)
        common = [d for d, syms in coverage.items() if len(syms) >= min_req]
        if common:
            common = sorted(common)
            print(f"Using {len(common)} dates (≥{min_req} symbols).")
        else:
            common = sorted(set(all_dates))
            print(f"No date meets threshold; using all {len(common)} dates.")

        idx = pd.DatetimeIndex(common)
        port_df = pd.DataFrame(index=idx)

        # Portfolio 1: Best Strategy
        best_cols = []
        for sym, df in all_data_dict.items():
            if df is not None and all_applied_dict[sym]['best']:
                strat_name = all_applied_dict[sym]['best'][0]
                daily_col  = f"Daily_PnL_{strat_name}"
                if daily_col in df.columns:
                    aligned = df.reindex(idx)[daily_col].fillna(0)
                    port_df[f"Best_{sym}"] = aligned
                    best_cols.append(f"Best_{sym}")
                    print(f"  Added Best strategy PnL for {sym}")

        if best_cols:
            bd = port_df[best_cols]
            act = (bd != 0).sum(axis=1)
            port_df["Portfolio_1_Best"] = (bd.sum(axis=1) / act.replace(0, np.nan)).fillna(0)
            print(f"Portfolio 1: {len(best_cols)} symbols")
        else:
            port_df["Portfolio_1_Best"] = 0
            print("Portfolio 1: No best strategies")

        # Portfolio 2: K-means
        km_cols = []
        for sym, df in all_data_dict.items():
            if df is not None and all_applied_dict[sym]['kmeans']:
                series_list = []
                for nm in all_applied_dict[sym]['kmeans']:
                    col = f"Daily_PnL_{nm}"
                    if col in df.columns:
                        series_list.append(df.reindex(idx)[col].fillna(0))
                if series_list:
                    tmp = pd.concat(series_list, axis=1)
                    act = (tmp != 0).sum(axis=1)
                    avg = (tmp.sum(axis=1) / act.replace(0, np.nan)).fillna(0)
                    port_df[f"Kmeans_Avg_{sym}"] = avg
                    km_cols.append(f"Kmeans_Avg_{sym}")
                    print(f"  Added K-means avg PnL for {sym}")

        if km_cols:
            kd = port_df[km_cols]
            act2 = (kd != 0).sum(axis=1)
            port_df["Portfolio_2_Kmeans"] = (kd.sum(axis=1) / act2.replace(0, np.nan)).fillna(0)
            print(f"Portfolio 2: {len(km_cols)} symbols")
        else:
            port_df["Portfolio_2_Kmeans"] = 0
            print("Portfolio 2: No K-means portfolios")

        # Portfolio 3: Hierarchical
        hi_cols = []
        for sym, df in all_data_dict.items():
            if df is not None and all_applied_dict[sym]['hierarchical']:
                series_list = []
                for nm in all_applied_dict[sym]['hierarchical']:
                    col = f"Daily_PnL_{nm}"
                    if col in df.columns:
                        series_list.append(df.reindex(idx)[col].fillna(0))
                if series_list:
                    tmp = pd.concat(series_list, axis=1)
                    act = (tmp != 0).sum(axis=1)
                    avg = (tmp.sum(axis=1) / act.replace(0, np.nan)).fillna(0)
                    port_df[f"Hierarchical_Avg_{sym}"] = avg
                    hi_cols.append(f"Hierarchical_Avg_{sym}")
                    print(f"  Added Hierarchical avg PnL for {sym}")

        if hi_cols:
            hd = port_df[hi_cols]
            act3 = (hd != 0).sum(axis=1)
            port_df["Portfolio_3_Hierarchical"] = (hd.sum(axis=1) / act3.replace(0, np.nan)).fillna(0)
            print(f"Portfolio 3: {len(hi_cols)} symbols")
        else:
            port_df["Portfolio_3_Hierarchical"] = 0
            print("Portfolio 3: No hierarchical portfolios")

        # Compute cumulative P&L columns
        port_df["Portfolio_1_Best_Cumulative"]         = port_df["Portfolio_1_Best"].cumsum()
        port_df["Portfolio_2_Kmeans_Cumulative"]       = port_df["Portfolio_2_Kmeans"].cumsum()
        port_df["Portfolio_3_Hierarchical_Cumulative"] = port_df["Portfolio_3_Hierarchical"].cumsum()

        return port_df

    def create_portfolio_visualizations(port_df):
        """
        Plot the cumulative P&L comparison (2014–2024), zero‐based at 2014-01-01,
        and add in-sample, out-of-sample, and full-period Sharpe ratios into
        each portfolio's legend entry (top-left).
        """
        # 1) Define the display window:
        start_plot = pd.to_datetime("2014-01-01")
        end_plot   = pd.to_datetime("2024-12-31")

        # 2) Slice the DataFrame to that window:
        df = port_df.loc[start_plot:end_plot].copy()
        if df.empty:
            print(f"No data between {start_plot.date()} and {end_plot.date()}. Nothing to plot.")
            return

        # 3) Compute 70% train / 30% test split index & date:
        split_idx = int(len(df) * TRAIN_TEST_SPLIT)
        split_dt  = df.index[split_idx]

        # 4) Zero‐base each cumulative P&L so that all curves start at zero:
        adj_best = df["Portfolio_1_Best_Cumulative"] - df["Portfolio_1_Best_Cumulative"].iloc[0]
        adj_km   = df["Portfolio_2_Kmeans_Cumulative"] - df["Portfolio_2_Kmeans_Cumulative"].iloc[0]
        adj_hl   = df["Portfolio_3_Hierarchical_Cumulative"] - df["Portfolio_3_Hierarchical_Cumulative"].iloc[0]

        # 5) Extract daily‐return (non‐cumulative) series for Sharpe calculation:
        ret_best_full = df["Portfolio_1_Best"]
        ret_km_full   = df["Portfolio_2_Kmeans"]
        ret_hl_full   = df["Portfolio_3_Hierarchical"]

        # 6) Split into in‐sample and out‐of‐sample:
        ret_best_in  = ret_best_full.iloc[:split_idx]
        ret_best_out = ret_best_full.iloc[split_idx:]

        ret_km_in    = ret_km_full.iloc[:split_idx]
        ret_km_out   = ret_km_full.iloc[split_idx:]

        ret_hl_in    = ret_hl_full.iloc[:split_idx]
        ret_hl_out   = ret_hl_full.iloc[split_idx:]

        # 7) Helper to compute annualized Sharpe ratio (252 trading days)
        def annualized_sharpe(returns):
            mu = returns.mean()
            sd = returns.std()
            return (mu / sd * np.sqrt(252)) if (sd > 0) else 0.0

        # 8) Compute Sharpe ratios:
        sharpe_best_in   = annualized_sharpe(ret_best_in)
        sharpe_best_out  = annualized_sharpe(ret_best_out)
        sharpe_best_full = annualized_sharpe(ret_best_full)

        sharpe_km_in     = annualized_sharpe(ret_km_in)
        sharpe_km_out    = annualized_sharpe(ret_km_out)
        sharpe_km_full   = annualized_sharpe(ret_km_full)

        sharpe_hl_in     = annualized_sharpe(ret_hl_in)
        sharpe_hl_out    = annualized_sharpe(ret_hl_out)
        sharpe_hl_full   = annualized_sharpe(ret_hl_full)

        # 9) Build legend labels that include Sharpe ratios:
        label_best = (
            f"Best Strategy Portfolio\n"
            f"  In:  {sharpe_best_in:.3f}, "
            f"Out: {sharpe_best_out:.3f}, "
            f"Full: {sharpe_best_full:.3f}"
        )
        label_km = (
            f"K-means Cluster Portfolio\n"
            f"  In:  {sharpe_km_in:.3f}, "
            f"Out: {sharpe_km_out:.3f}, "
            f"Full: {sharpe_km_full:.3f}"
        )
        label_hl = (
            f"Hierarchical Cluster Portfolio\n"
            f"  In:  {sharpe_hl_in:.3f}, "
            f"Out: {sharpe_hl_out:.3f}, "
            f"Full: {sharpe_hl_full:.3f}"
        )

        # 10) Plot the cumulative P&L curves:
        plt.figure(figsize=(15, 6))
        colors = ["blue", "green", "purple"]

        # Full‐window lines (2014–2024):
        line_best, = plt.plot(df.index, adj_best, color=colors[0], linewidth=2)
        line_km,   = plt.plot(df.index, adj_km,   color=colors[1], linewidth=2)
        line_hl,   = plt.plot(df.index, adj_hl,   color=colors[2], linewidth=2)

        # Highlight out‐of‐sample segment with thicker, semi-transparent lines:
        plt.plot(df.index[split_idx:], adj_best.iloc[split_idx:], color=colors[0], linewidth=3, alpha=0.8)
        plt.plot(df.index[split_idx:], adj_km.iloc[split_idx:],   color=colors[1], linewidth=3, alpha=0.8)
        plt.plot(df.index[split_idx:], adj_hl.iloc[split_idx:],   color=colors[2], linewidth=3, alpha=0.8)

        # Vertical train/test split:
        plt.axvline(
            x=split_dt,
            color="black",
            linestyle="--",
            alpha=0.7,
            label=f"Train/Test Split ({int(TRAIN_TEST_SPLIT*100)}%/{int((1-TRAIN_TEST_SPLIT)*100)}%)"
        )

        # Zero (break-even) line:
        plt.axhline(
            y=0,
            color="gray",
            linestyle="-",
            alpha=0.5,
            label="Break-even"
        )

        plt.title("Portfolio Cumulative P&L Comparison (2014–2024, All Start at 0)", fontsize=16)
        plt.ylabel("Cumulative P&L ($)", fontsize=12)
        plt.grid(True, alpha=0.3)

        # 11) Build a combined legend entry in the top-left:
        # We manually specify the handles and labels so that our three portfolios (with Sharpe info)
        # appear first, followed by the split and break-even lines.
        handles = [line_best, line_km, line_hl]
        labels  = [label_best, label_km, label_hl]

        # Add the train/test split and break-even handles to the legend as well:
        split_handle = plt.Line2D([], [], color="black", linestyle="--", alpha=0.7)
        zero_handle  = plt.Line2D([], [], color="gray", linestyle="-", alpha=0.5)
        handles.extend([split_handle, zero_handle])
        labels.extend([
            f"Train/Test Split ({int(TRAIN_TEST_SPLIT*100)}%/{int((1-TRAIN_TEST_SPLIT)*100)}%)",
            "Break-even"
        ])

        # Place the legend in the top-left corner, using 'monospace' for better alignment:
        plt.legend(handles, labels, loc="upper left", fontsize=9, framealpha=0.8, prop={"family":"monospace"})

        plt.tight_layout()
        save_plot("portfolio_cumulative_pnl_with_sharpe_in_legend.png")

    # --------------------------------------------
    # MAIN EXECUTION
    # --------------------------------------------
    print("=" * 60)
    print("MULTI-STRATEGY PORTFOLIO ANALYSIS (DYNAMIC SLIPPAGE)")
    print("=" * 60)

    # 1) Load strategy grid from Excel
    strategies_dict = load_excel_strategies()
    if not strategies_dict:
        print("No strategies found. Exiting.")
        return None

    # 2) Loop over each symbol, load & prepare data, then apply strategies:
    all_data_dict       = {}
    all_strategies_dict = {}

    for symbol in strategies_dict:
        # Skip if no strategies at all:
        if (not strategies_dict[symbol]['best_sharpe'] and
            not strategies_dict[symbol]['kmeans_clusters'] and
            not strategies_dict[symbol]['hierarchical_clusters']):
            continue

        # a) Load and prepare raw OHLC data (with warm-up)
        df, original_start_idx = load_and_prepare_data(symbol, strategies_dict)
        if df is None:
            continue

        # b) Apply all SMA strategies for this symbol, using the slippage & BPV we fetched:
        processed_df, applied_names = apply_strategies_to_data(
            df, symbol, strategies_dict, original_start_idx
        )

        if processed_df is not None:
            all_data_dict[symbol]       = processed_df
            all_strategies_dict[symbol] = applied_names

    print(f"\nSuccessfully processed {len(all_data_dict)} symbols.")

    if not all_data_dict:
        print("No valid data was processed. Exiting.")
        return None

    # 3) Build portfolios (daily & cumulative P&L)
    portfolio_df = build_portfolios(all_data_dict, all_strategies_dict)
    if portfolio_df is None:
        print("Failed to build portfolios. Exiting.")
        return None

    # 4) Plot only the single cumulative P&L comparison chart
    create_portfolio_visualizations(portfolio_df)

    # 5) Perform bimonthly comparisons
    print("\nPerforming bimonthly portfolio comparisons...")

    # First comparison: K-means vs Best
    kmeans_bimonthly_results = bimonthly_kmeans_vs_best_comparison(portfolio_df, save_plot)

    # Second comparison: Hierarchical vs Best
    hierarchical_bimonthly_results = bimonthly_hierarchical_vs_best_comparison(portfolio_df, save_plot)

    print("\nPortfolio analysis complete. (No parameters.json needed.)")
    return portfolio_df


if __name__ == "__main__":
    portfolio_analysis_main()