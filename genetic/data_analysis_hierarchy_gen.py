import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import matplotlib.lines as mlines
from matplotlib.patches import Circle
from sklearn.cluster import AgglomerativeClustering
from sklearn.preprocessing import StandardScaler
from datetime import datetime
import scipy.cluster.hierarchy as shc
from scipy.spatial.distance import pdist, squareform
import calendar
import os
import glob
import read_ts
import matplotlib.pyplot as plt
import numpy as np
from scipy.spatial.distance import cdist
import matplotlib.pyplot as plt
import scipy.cluster.hierarchy as shc
from scipy.spatial.distance import pdist
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import numpy as np
from matplotlib.lines import Line2D
import matplotlib
import openpyxl

from input_gen import MIN_TRADES, MAX_TRADES, MIN_ELEMENTS_PER_CLUSTER, DEFAULT_NUM_CLUSTERS, ELBOW_THRESHOLD
from input_gen import TICKER, START_DATE, END_DATE, TRAIN_TEST_SPLIT, ATR_PERIOD, TRADING_CAPITAL
from SMA_Strategy import SMAStrategy

import json

SYMBOL = TICKER.replace('=F', '')  # Define SYMBOL based on TICKER

# Function to save plots in the created folder
def save_plot(plot_name):
    output_dir = os.path.join(".", 'output1', TICKER.replace('=F', ''))
    os.makedirs(output_dir, exist_ok=True)
    plt.savefig(os.path.join(output_dir, plot_name))
    plt.close()

def save_results_to_excel(ticker_row, sheet, data):
    """
    Save hierarchical strategy results to Excel with specific column mappings.
    
    Column mappings:
    C (3) - Hierarchical win rate
    H-J (8-10) - Top 3 hierarchical medoids
    AC (29) - Portfolio OOS Sharpe (Hierarchical)
    """
    try:
        # Symbol is already written in column A by the calling function
        
        # Column C - Hierarchical win rate
        if 'hierarchical_win_rate' in data:
            sheet.cell(row=ticker_row, column=3).value = round(data['hierarchical_win_rate'], 1)
        
        # Columns H-J - Top 3 hierarchical medoids
        if 'hierarchical_medoids' in data:
            for i, medoid in enumerate(data['hierarchical_medoids']):
                if i >= 3:  # Only use up to 3 medoids
                    break
                column_idx = 8 + i  # H=8, I=9, J=10
                param_value = f"{int(medoid[0])}/{int(medoid[1])}"
                sheet.cell(row=ticker_row, column=column_idx).value = param_value
        
        # Column AC - Portfolio OOS Sharpe (Hierarchical)
        if 'portfolio_oos_sharpe_hierarchical' in data:
            sheet.cell(row=ticker_row, column=29).value = round(data['portfolio_oos_sharpe_hierarchical'], 4)
            
    except Exception as e:
        print(f"Error writing data to Excel: {e}")
        raise

def update_excel_results(data_dict, results_file='Results.xlsx'):
    """
    Update the Excel results file with strategy performance data.
    Creates the file if it doesn't exist.
    """
    try:
        if not os.path.exists(results_file):
            print(f"Creating new Excel results file: {results_file}")
            wb = openpyxl.Workbook()
        else:
            print(f"Updating existing Excel file: {results_file}")
            wb = openpyxl.load_workbook(results_file)
        
        sheet = wb.active
        
        # Find or create row for this symbol
        row = 3  # Start from row 3 (assuming rows 1-2 have headers)
        ticker_row = None
        
        while True:
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value == SYMBOL:
                ticker_row = row
                break
            elif cell_value is None:
                ticker_row = row
                sheet.cell(row=ticker_row, column=1).value = SYMBOL
                break
            row += 1
            
        # Save the data to appropriate columns
        save_results_to_excel(ticker_row, sheet, data_dict)
        
        # Save the workbook
        wb.save(results_file)
        print(f"Excel file updated successfully for {SYMBOL} in row {ticker_row}")
        
    except Exception as e:
        print(f"Error updating Excel file: {e}")
        raise

def calculate_elbow_curve(X_scaled, max_clusters=20):
    """
    Calculate the elbow curve for KMeans clustering
    
    Parameters:
    X_scaled: numpy array - Scaled data points
    max_clusters: int - Maximum number of clusters to try
    
    Returns:
    distortions: list - Distortion values for each k
    k_values: list - K values used
    optimal_k: int - Optimal number of clusters based on elbow method
    """
    print("\nCalculating elbow curve...")
    distortions = []
    k_values = range(1, max_clusters + 1)
    
    for k in k_values:
        if k == 1:
            # For k=1, distortion is just the sum of squared distances to mean
            distortion = np.sum((X_scaled - np.mean(X_scaled, axis=0)) ** 2)
            distortions.append(distortion)
            continue
            
        clustering = AgglomerativeClustering(n_clusters=k)
        clustering.fit(X_scaled)
        
        # Calculate distortion for current clustering
        centroids = np.array([X_scaled[clustering.labels_ == i].mean(axis=0) for i in range(k)])
        distortion = np.sum([np.sum((X_scaled[clustering.labels_ == i] - centroids[i]) ** 2) 
                           for i in range(k)])
        distortions.append(distortion)
    
    # Calculate the percentage changes
    pct_changes = np.diff(distortions) / np.array(distortions)[:-1] * 100
    
    # Find optimal k using the threshold method
    optimal_k = 1
    for i, pct_change in enumerate(pct_changes):
        if abs(pct_change) < ELBOW_THRESHOLD:
            optimal_k = i + 1  # +1 because we start from k=1
            break
    
    # Plot the elbow curve
    plt.figure(figsize=(10, 6))
    plt.plot(k_values, distortions, 'bo-')
    plt.axvline(x=optimal_k, color='r', linestyle='--', 
                label=f'Optimal k={optimal_k}\n(threshold={ELBOW_THRESHOLD}%)')
    plt.xlabel('Number of Clusters (k)')
    plt.ylabel('Distortion')
    plt.title(f'{SYMBOL} Hierarchical Elbow Method for Optimal k')
    plt.grid(True)
    plt.legend()
    plt.tight_layout()
    save_plot(f'{SYMBOL}_Hierarchical_Elbow_Curve.png')
    
    print(f"Optimal number of clusters from elbow method: {optimal_k}")
    return distortions, k_values, optimal_k

def main():
    def load_parameters():
        try:
            with open("parameters.json", "r") as file:
                parameters = json.load(file)
                big_point_value = parameters["big_point_value"]
                slippage = parameters["slippage"]
                capital = parameters.get("capital", TRADING_CAPITAL)
                atr_period = parameters.get("atr_period", ATR_PERIOD)
                return big_point_value, slippage, capital, atr_period
        except FileNotFoundError:
            print("Parameters file not found. Ensure it was saved correctly in data_gather_gen.py.")
            return None, None, TRADING_CAPITAL, ATR_PERIOD
        except KeyError as e:
            print(f"Missing key in parameters.json: {e}")
            print(f"Available keys: {', '.join(parameters.keys())}")
            return None, None, TRADING_CAPITAL, ATR_PERIOD

    # Load the parameters
    big_point_value, slippage, capital, atr_period = load_parameters()

    print(f"Big Point Value: {big_point_value}")
    print(f"Dynamic Slippage: {slippage}")
    print(f"Capital for Position Sizing: {capital:,}")
    print(f"ATR Period: {atr_period}")

    # Setup paths
    WORKING_DIR = "."  # Current directory
    DATA_DIR = os.path.join(WORKING_DIR, "data")
    SYMBOL = TICKER.replace('=F', '')

    # Define the output folder where the plots will be saved
    output_dir = os.path.join(WORKING_DIR, 'output1', SYMBOL)
    os.makedirs(output_dir, exist_ok=True)  # Create the directory if it doesn't exist



    def find_futures_file(symbol, data_dir):
        """Find a data file for the specified futures symbol"""
        # First try a pattern that specifically looks for @SYMBOL
        pattern = f"*@{symbol}_*.dat"
        files = glob.glob(os.path.join(data_dir, pattern))
        
        if not files:
            # Try a pattern that looks for the symbol with an underscore or at boundary
            pattern = f"*_@{symbol}_*.dat"
            files = glob.glob(os.path.join(data_dir, pattern))
        
        if not files:
            # Try a more specific boundary pattern for the symbol
            pattern = f"*_{symbol}_*.dat"
            files = glob.glob(os.path.join(data_dir, pattern))
        
        if not files:
            # Last resort: less specific but better than nothing
            pattern = f"*@{symbol}*.dat"
            files = glob.glob(os.path.join(data_dir, pattern))
        
        return files[0] if files else None

    # Load the SMA simulation results
    def analyze_sma_results(file_path='sma_all_results.txt'):
        print(f"Loading simulation results from {file_path}...")

        # Load the data from the CSV file
        data = pd.read_csv(file_path)

        # Print basic information about the data
        print(f"Loaded {len(data)} simulation results")
        print(f"Short SMA range: {data['short_SMA'].min()} to {data['short_SMA'].max()}")
        print(f"Long SMA range: {data['long_SMA'].min()} to {data['long_SMA'].max()}")
        print(f"Sharpe ratio range: {data['sharpe_ratio'].min():.4f} to {data['sharpe_ratio'].max():.4f}")

        # Find the best Sharpe ratio
        best_idx = data['sharpe_ratio'].idxmax()
        best_short_sma = data.loc[best_idx, 'short_SMA']
        best_long_sma = data.loc[best_idx, 'long_SMA']
        best_sharpe = data.loc[best_idx, 'sharpe_ratio']
        best_trades = data.loc[best_idx, 'trades']

        print(f"\nBest parameters:")
        print(f"Short SMA: {best_short_sma}")
        print(f"Long SMA: {best_long_sma}")
        print(f"Sharpe Ratio: {best_sharpe:.6f}")
        print(f"Number of Trades: {best_trades}")

        # Create a pivot table for the heatmap
        heatmap_data = data.pivot_table(
            index='long_SMA',
            columns='short_SMA',
            values='sharpe_ratio'
        )

        # Create the heatmap visualization
        plt.figure(figsize=(12, 10))

        # Create a mask for invalid combinations (where short_SMA >= long_SMA)
        mask = np.zeros_like(heatmap_data, dtype=bool)
        for i, long_sma in enumerate(heatmap_data.index):
            for j, short_sma in enumerate(heatmap_data.columns):
                if short_sma >= long_sma:
                    mask[i, j] = True

        # Plot the heatmap with the mask
        ax = sns.heatmap(
            heatmap_data,
            mask=mask,
            cmap='coolwarm',  # Blue to red colormap
            annot=False,  # Don't annotate each cell with its value
            fmt='.4f',
            linewidths=0,
            cbar_kws={'label': 'Sharpe Ratio'}
        )

        # Invert the y-axis so smaller long_SMA values are at the top
        ax.invert_yaxis()

        # Find the position of the best Sharpe ratio in the heatmap
        best_y = heatmap_data.index.get_loc(best_long_sma)
        best_x = heatmap_data.columns.get_loc(best_short_sma)

        # Add a star to mark the best Sharpe ratio
        # We need to add 0.5 to center the marker in the cell
        ax.add_patch(Circle((best_x + 0.5, best_y + 0.5), 0.4, facecolor='none',
                            edgecolor='white', lw=2))
        plt.plot(best_x + 0.5, best_y + 0.5, 'w*', markersize=10)

        # Set labels and title
        plt.title(f'{SYMBOL} SMA Optimization Heatmap (Best Sharpe: {best_sharpe:.4f} at {best_short_sma}/{best_long_sma})',
                fontsize=14)
        plt.xlabel('Short SMA (days)', fontsize=12)
        plt.ylabel('Long SMA (days)', fontsize=12)

        # Rotate tick labels for better readability
        plt.xticks(rotation=45, ha='right')
        plt.yticks(rotation=0)

        # Add a text annotation for the best parameters
        plt.annotate(
            f'Best: Short={best_short_sma}, Long={best_long_sma}\nSharpe={best_sharpe:.4f}, Trades={best_trades}',
            xy=(best_x + 0.5, best_y + 0.5),
            xytext=(best_x + 5, best_y + 5),
            arrowprops=dict(arrowstyle="->", color='white'),
            color='white',
            backgroundcolor='black',
            bbox=dict(boxstyle="round,pad=0.3", fc="black", alpha=0.7)
        )

        # Display and save the plot
        plt.tight_layout()


        # Return the data and best parameters
        return data, best_short_sma, best_long_sma, best_sharpe, best_trades

    def compute_hierarchical_medoids(X, labels, valid_clusters):
        """
        Compute medoids for each valid cluster from hierarchical clustering
        A medoid is the data point in a cluster that has the minimum average distance to all other points in the cluster

        Parameters:
        X: numpy array of shape (n_samples, n_features) - Original data points (not scaled)
        labels: numpy array of shape (n_samples,) - Cluster labels for each data point
        valid_clusters: set - Set of valid cluster IDs

        Returns:
        list of tuples - Each tuple contains (short_SMA, long_SMA, sharpe_ratio, trades) for each medoid
        """

        # Initialize list to store medoids for each cluster
        medoids = []

        # Process each valid cluster
        for cluster_id in valid_clusters:
            # Extract points belonging to this cluster
            cluster_indices = np.where(labels == cluster_id)[0]
            cluster_points = X[cluster_indices]

            if len(cluster_points) == 0:
                continue

            if len(cluster_points) == 1:
                # If only one point in cluster, it's the medoid
                medoids.append(tuple(cluster_points[0]))
                continue

            # Compute pairwise distances between all points in this cluster
            distances = cdist(cluster_points, cluster_points, metric='euclidean')

            # For each point, compute sum of distances to all other points
            total_distances = np.sum(distances, axis=1)

            # Find index of point with minimum total distance
            medoid_idx = np.argmin(total_distances)

            # Get the actual medoid point
            medoid = cluster_points[medoid_idx]

            # Store the medoid as a tuple
            medoids.append(tuple(medoid))

            # Print medoid info
            print(f"Cluster {cluster_id} medoid: Short SMA={int(medoid[0])}, Long SMA={int(medoid[1])}, "
                f"Sharpe={medoid[2]:.4f}, Trades={int(medoid[3])}")

        return medoids

    def create_dendrogram(X_scaled, method='ward', figsize=(12, 8), color_threshold=None, truncate_mode=None,
                        p=10, save_path=None):
        """
        Create and display a dendrogram for hierarchical clustering

        Parameters:
        X_scaled: numpy array - Scaled data points
        method: str - Linkage method ('ward', 'complete', 'average', 'single')
        figsize: tuple - Figure size
        color_threshold: float - The threshold to apply for coloring the dendrogram
        truncate_mode: str - 'lastp' for last p branches, 'level' for no more than p levels
        p: int - Used with truncate_mode
        save_path: str - Path to save the dendrogram image

        Returns:
        Z: numpy array - The hierarchical clustering linkage matrix
        """
        # Use a sample if there are too many points
        if len(X_scaled) > 1000:
            np.random.seed(42)  # For reproducibility
            sample_indices = np.random.choice(len(X_scaled), size=1000, replace=False)
            X_sample = X_scaled[sample_indices]
            print(f"Using a random sample of 1000 points for dendrogram creation (out of {len(X_scaled)} total points)")
        else:
            X_sample = X_scaled

        # Compute the distance matrix
        dist_matrix = pdist(X_sample, metric='euclidean')

        # Compute the linkage matrix
        Z = shc.linkage(dist_matrix, method=method)

        # Print some statistics about the linkage
        print(f"\nDendrogram using {method} linkage method:")
        print(f"Number of data points: {len(X_sample)}")
        print(f"Cophenetic correlation: {shc.cophenet(Z, dist_matrix)[0]:.4f}")

        # Create a new figure
        plt.figure(figsize=figsize)

        # Set the background color
        plt.gca().set_facecolor('white')

        # Set a title
        plt.title(f'{SYMBOL} Hierarchical Clustering Dendrogram ({method} linkage)', fontsize=14)

        # Generate the dendrogram
        dendrogram = shc.dendrogram(
            Z,
            truncate_mode='lastp',  # Only show the last p merges
            p=30,  # Show only 30 merges
            leaf_rotation=90.,
            leaf_font_size=10.,
            show_contracted=True,
            color_threshold=color_threshold
        )

        # Add labels and axes
        plt.xlabel('Sample Index or Cluster Size', fontsize=12)
        plt.ylabel('Distance', fontsize=12)

        # Add a horizontal line to indicate a distance threshold if specified
        if color_threshold is not None:
            plt.axhline(y=color_threshold, color='crimson', linestyle='--',
                        label=f'Threshold: {color_threshold:.2f}')
            plt.legend()

        # Adjust layout
        plt.tight_layout()

        # Save the plot
        save_plot(f'{SYMBOL}_Hierarchical_Dendrogram.png')

        return Z

    def hierarchical_cluster_analysis(file_path='sma_all_results.txt', min_trades=MIN_TRADES, max_trades=MAX_TRADES,
                                min_elements_per_cluster=MIN_ELEMENTS_PER_CLUSTER):
        """
        Perform hierarchical clustering analysis on SMA optimization results to find robust parameter regions

        Parameters:
        file_path: str - Path to the file with SMA results
        min_trades: int - Minimum number of trades to consider
        max_trades: int - Maximum number of trades to consider
        min_elements_per_cluster: int - Minimum number of elements per cluster

        Returns:
        tuple - (X_filtered, medoids, top_medoids, max_sharpe_point, labels)
        """
        print(f"\n----- HIERARCHICAL CLUSTER ANALYSIS -----")
        print(f"Loading data from {file_path}...")

        # Load the data
        df = pd.read_csv(file_path)

        # Convert data to numpy array for easier processing
        X = df[['short_SMA', 'long_SMA', 'sharpe_ratio', 'trades']].values

        # Filter data by number of trades and ensure short_SMA < long_SMA
        X_filtered = X[(X[:, 0] < X[:, 1]) &  # short_SMA < long_SMA
                    (X[:, 3] >= min_trades) &  # trades >= min_trades
                    (X[:, 3] <= max_trades)]  # trades <= max_trades

        if len(X_filtered) == 0:
            print(
                f"No data points meet the criteria after filtering! Adjust min_trades ({min_trades}) and max_trades ({max_trades}).")
            return None, None, None, None, None

        print(f"Filtered data to {len(X_filtered)} points with {min_trades}-{max_trades} trades")

        # Extract the fields for better scaling visibility
        short_sma_values = X_filtered[:, 0]
        long_sma_values = X_filtered[:, 1]
        sharpe_values = X_filtered[:, 2]
        trades_values = X_filtered[:, 3]

        print(f"Short SMA range: {short_sma_values.min()} to {short_sma_values.max()}")
        print(f"Long SMA range: {long_sma_values.min()} to {long_sma_values.max()}")
        print(f"Sharpe ratio range: {sharpe_values.min():.4f} to {sharpe_values.max():.4f}")
        print(f"Trades range: {trades_values.min()} to {trades_values.max()}")

        # Scale the data for clustering - using StandardScaler for each dimension
        # This addresses the issue where SMA values have much larger ranges than Sharpe ratio
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X_filtered)

        # Print scaling info for verification
        print("\nScaled data information:")
        scaled_short = X_scaled[:, 0]
        scaled_long = X_scaled[:, 1]
        scaled_sharpe = X_scaled[:, 2]
        scaled_trades = X_scaled[:, 3]

        print(f"Scaled Short SMA range: {scaled_short.min():.4f} to {scaled_short.max():.4f}")
        print(f"Scaled Long SMA range: {scaled_long.min():.4f} to {scaled_long.max():.4f}")
        print(f"Scaled Sharpe ratio range: {scaled_sharpe.min():.4f} to {scaled_sharpe.max():.4f}")
        print(f"Scaled Trades range: {scaled_trades.min():.4f} to {scaled_trades.max():.4f}")

        # Create a dendrogram to help choose the number of clusters
        print("\nCreating dendrogram to visualize hierarchical structure...")
        linkage_method = 'ward'  # Ward minimizes the variance within clusters
        Z = create_dendrogram(X_scaled, method=linkage_method,
                            figsize=(12, 8))

        # Calculate optimal number of clusters using elbow method
        print("\nDetermining optimal number of clusters using elbow method...")
        _, _, k = calculate_elbow_curve(X_scaled, max_clusters=min(50, len(X_scaled) // 2))
        print(f"Using {k} clusters based on elbow method (threshold={ELBOW_THRESHOLD}%)")

        # Apply hierarchical clustering
        print(f"Performing hierarchical clustering with {k} clusters using {linkage_method} linkage...")
        if linkage_method == 'ward':
            # Ward linkage requires euclidean distance
            hierarchical = AgglomerativeClustering(
                n_clusters=k,
                linkage=linkage_method
            )
        else:
            # For other linkage methods, we can specify affinity
            hierarchical = AgglomerativeClustering(
                n_clusters=k,
                affinity='euclidean',
                linkage=linkage_method
            )

        # Fit the model and get cluster labels
        labels = hierarchical.fit_predict(X_scaled)

        # Count elements per cluster
        unique_labels, counts = np.unique(labels, return_counts=True)
        cluster_sizes = dict(zip(unique_labels, counts))

        print("\nCluster sizes:")
        for cluster_id, size in cluster_sizes.items():
            print(f"Cluster {cluster_id}: {size} elements")

        # Filter clusters with enough elements
        valid_clusters = {i for i, count in cluster_sizes.items() if count >= min_elements_per_cluster}
        if not valid_clusters:
            print(f"No clusters have at least {min_elements_per_cluster} elements! Reducing threshold to 1...")
            valid_clusters = set(unique_labels)
        else:
            print(f"Using {len(valid_clusters)} clusters with at least {min_elements_per_cluster} elements")

        filtered_indices = np.array([i in valid_clusters for i in labels])

        # Filter data to only include points in valid clusters
        X_valid = X_filtered[filtered_indices]
        labels_valid = labels[filtered_indices]

        # Compute medoids of hierarchical clusters
        print("Computing medoids for each cluster...")
        medoids = compute_hierarchical_medoids(X_valid, labels_valid, valid_clusters)

        # Find max Sharpe ratio point overall
        max_sharpe_idx = np.argmax(df['sharpe_ratio'].values)
        max_sharpe_point = df.iloc[max_sharpe_idx][['short_SMA', 'long_SMA', 'sharpe_ratio', 'trades']].values

        # Sort medoids by Sharpe ratio
        medoids_sorted = sorted(medoids, key=lambda x: x[2], reverse=True)
        top_medoids = medoids_sorted[:5]  # Get top 5 medoids by Sharpe ratio

        # Print results
        print("\n----- HIERARCHICAL CLUSTERING RESULTS -----")
        print(f"Max Sharpe point: Short SMA={int(max_sharpe_point[0])}, Long SMA={int(max_sharpe_point[1])}, "
            f"Sharpe={max_sharpe_point[2]:.4f}, Trades={int(max_sharpe_point[3])}")

        print("\nTop 5 Medoids (by Sharpe ratio):")
        for idx, medoid in enumerate(top_medoids, 1):
            print(f"Top {idx}: Short SMA={int(medoid[0])}, Long SMA={int(medoid[1])}, "
                f"Sharpe={medoid[2]:.4f}, Trades={int(medoid[3])}")

        # Create visualization with clustering results
        create_hierarchical_cluster_visualization(X_filtered, medoids, top_medoids, max_sharpe_point, labels)

        # Save clustering results to CSV
        clustering_results = []
        for idx, medoid in enumerate(medoids_sorted):
            clustering_results.append({
                'Rank': idx + 1,
                'Short_SMA': int(medoid[0]),
                'Long_SMA': int(medoid[1]),
                'Sharpe': medoid[2],
                'Trades': int(medoid[3])
            })
        
        clustering_df = pd.DataFrame(clustering_results)

        # Return exactly 5 values
        return X_filtered, medoids, top_medoids, max_sharpe_point, labels

    # Update the plot_strategy_performance function in data_analysis_hierarchy.py
# Around line 581 in the function, replace the existing plotting code with:

    def plot_strategy_performance(short_sma, long_sma, top_medoids=None, big_point_value=big_point_value,
                            slippage=slippage, capital=capital, atr_period=atr_period):
        """
        Plot the strategy performance using the best SMA parameters and include top medoids
        Uses the SMAStrategy class for consistent logic across the codebase

        Parameters:
        short_sma: int - The short SMA period
        long_sma: int - The long SMA period
        top_medoids: list - List of top medoids, each as (short_sma, long_sma, sharpe, trades)
        big_point_value: float - Big point value for the futures contract
        slippage: float - Slippage value in price units
        capital: float - Capital allocation for position sizing
        atr_period: int - Period for ATR calculation
        """
        print(f"\n----- PLOTTING STRATEGY PERFORMANCE -----")
        print(f"Using Short SMA: {short_sma}, Long SMA: {long_sma}")
        print(f"Trading with ATR-based position sizing (Capital: ${capital:,}, ATR Period: {atr_period})")
        if top_medoids:
            print(f"Including top {len(top_medoids)} medoids")

        # Load data from local files
        print(f"Loading {TICKER} data from local files...")
        data_file = find_futures_file(SYMBOL, DATA_DIR)
        if not data_file:
            print(f"Error: No data file found for {TICKER} in {DATA_DIR}")
            exit(1)
        
        print(f"Found data file: {os.path.basename(data_file)}")
        print(f"File size: {os.path.getsize(data_file)} bytes")

        # Load the data from local file
        all_data = read_ts.read_ts_ohlcv_dat(data_file)
        data_obj = all_data[0]
        ohlc_data = data_obj.data.copy()

        # Convert the OHLCV data to the format expected by the strategy
        data = ohlc_data.rename(columns={
            'datetime': 'Date',
            'open': 'Open',
            'high': 'High',
            'low': 'Low',
            'close': 'Close',
            'volume': 'Volume'
        })
        data.set_index('Date', inplace=True)

        # Add a warm-up period before the start date
        original_start_idx = None
        if START_DATE and END_DATE:
            # Calculate warm-up period for SMA calculation (longest SMA + buffer)
            warm_up_days = max(short_sma, long_sma) * 3  # Use 3x the longest SMA as warm-up
            
            # Convert dates to datetime
            start_date = pd.to_datetime(START_DATE)
            end_date = pd.to_datetime(END_DATE)
            
            # Adjust start date for warm-up
            adjusted_start = start_date - pd.Timedelta(days=warm_up_days)
            
            # Filter with the extended date range
            extended_data = data[(data.index >= adjusted_start) & (data.index <= end_date)]
            
            # Store the index where the actual analysis should start
            if not extended_data.empty:
                # Find the closest index to our original start date
                original_start_idx = extended_data.index.get_indexer([start_date], method='nearest')[0]
            
            print(f"Added {warm_up_days} days warm-up period before {START_DATE}")
            data = extended_data
        
        # Simplify the data structure if needed
        if hasattr(data.columns, 'get_level_values'):
            data.columns = data.columns.get_level_values(0)

        # Create a dictionary to store results for each strategy
        strategies = {
            'Best': {'short_sma': short_sma, 'long_sma': long_sma}
        }

        # Add medoids in their original order, including original Sharpe and Trades
        if top_medoids:
            print("\nUSING THESE MEDOIDS (IN ORIGINAL ORDER):")
            for i, medoid in enumerate(top_medoids, 1):
                strategies[f'Medoid {i}'] = {
                    'short_sma': int(medoid[0]),
                    'long_sma': int(medoid[1]),
                    'original_sharpe': float(medoid[2]),  # Store the original Sharpe ratio
                    'original_trades': int(medoid[3])     # Store the original number of trades
                }
                print(f"Medoid {i}: SMA({int(medoid[0])}/{int(medoid[1])}) - Original Sharpe: {float(medoid[2]):.4f}, Trades: {int(medoid[3])}")

        # Apply the proper strategy for each parameter set
        for name, params in strategies.items():
            # Calculate centered SMAs directly
            data[f'SMA_Short_{name}'] = data['Close'].rolling(window=params['short_sma'], center=True).mean()
            data[f'SMA_Long_{name}'] = data['Close'].rolling(window=params['long_sma'], center=True).mean()
            
            # Create a strategy instance for each parameter set
            sma_strategy = SMAStrategy(
                short_sma=params['short_sma'],
                long_sma=params['long_sma'],
                big_point_value=big_point_value,
                slippage=slippage,
                capital=capital,
                atr_period=atr_period
            )

            # Apply the strategy
            data = sma_strategy.apply_strategy(
                data.copy(),
                strategy_name=name
            )

        # Trim data to the original date range if we added warm-up period
        if original_start_idx is not None:
            data_for_evaluation = data.iloc[original_start_idx:]
            print(f"Trimmed warm-up period. Original data length: {len(data)}, Evaluation data length: {len(data_for_evaluation)}")
        else:
            data_for_evaluation = data
        
        # Calculate split index for in-sample/out-of-sample
        split_index = int(len(data_for_evaluation) * TRAIN_TEST_SPLIT)
        split_date = data_for_evaluation.index[split_index]

        # Create color palette for strategies
        colors = ['blue', 'green', 'purple', 'orange', 'brown', 'pink']

        # Create the performance visualization with just two panels
        plt.figure(figsize=(14, 12))

        # Plot price and SMA (first subplot)
        plt.subplot(2, 1, 1)
        plt.plot(data_for_evaluation.index, data_for_evaluation['Close'], label=f'{SYMBOL} Price', color='black', alpha=0.5)
        
        # Use the centered SMAs for plotting
        for name, params in strategies.items():
            if name == 'Best':
                plt.plot(data_for_evaluation.index, data_for_evaluation[f'SMA_Short_{name}'], 
                        label=f'Short SMA ({params["short_sma"]})', color='orange')
                plt.plot(data_for_evaluation.index, data_for_evaluation[f'SMA_Long_{name}'], 
                        label=f'Long SMA ({params["long_sma"]})', color='blue')
        
        # Mark the train/test split
        plt.axvline(x=split_date, color='black', linestyle='--', alpha=0.7)
        
        plt.legend(loc='upper left')
        plt.title(f'{SYMBOL} Price and SMA Indicators')
        plt.ylabel('Price')
        plt.grid(True, alpha=0.3)

        # Plot cumulative P&L (second subplot)
        plt.subplot(2, 1, 2)

        for i, (name, params) in enumerate(strategies.items()):
            color = colors[i % len(colors)]

            # Plot full period P&L
            plt.plot(data_for_evaluation.index, data_for_evaluation[f'Cumulative_PnL_{name}'],
                    label=f'{name} ({params["short_sma"]}/{params["long_sma"]})', color=color)

            # Plot out-of-sample portion with thicker line
            plt.plot(data_for_evaluation.index[split_index:], data_for_evaluation[f'Cumulative_PnL_{name}'].iloc[split_index:],
                    color=color, linewidth=2.5, alpha=0.7)

        plt.axvline(x=split_date, color='black', linestyle='--',
                    label=f'Train/Test Split ({int(TRAIN_TEST_SPLIT * 100)}%/{int((1 - TRAIN_TEST_SPLIT) * 100)}%)')
        plt.axhline(y=0.0, color='gray', linestyle='-', alpha=0.5, label='Break-even')
        plt.legend(loc='upper left')
        plt.title(f'{SYMBOL} Strategy Cumulative P&L')
        plt.ylabel('P&L ($)')
        plt.grid(True, alpha=0.3)

        plt.tight_layout()
        save_plot(f'{SYMBOL}_Hierarchical_Strategy_Performance.png')

        # Create a list to store performance data for saving to file
        performance_data = []

        # Print detailed performance metrics for all strategies with in-sample and out-of-sample breakdown
        print("\n----- PERFORMANCE SUMMARY -----")

        # IN-SAMPLE PERFORMANCE
        print("\nIN-SAMPLE PERFORMANCE:")
        header = f"{'Strategy':<10} | {'Short/Long':<10} | {'P&L':>12} | {'Avg Pos':>8} | {'Sharpe':>7} | {'Trades':>6}"
        separator = "-" * len(header)
        print(separator)
        print(header)
        print(separator)

        for name, params in strategies.items():
            short = params['short_sma']
            long = params['long_sma']

            # Get in-sample data
            in_sample = data_for_evaluation.iloc[:split_index]

            # Calculate in-sample metrics
            in_sample_daily_pnl = in_sample[f'Daily_PnL_{name}']
            in_sample_cumulative_pnl = in_sample[f'Daily_PnL_{name}'].sum()
            
            # Calculate average position size for in-sample
            avg_pos_size = in_sample[f'Position_Size_{name}'].mean()

            # Use original Sharpe and Trades for medoids, calculate for Best
            if name.startswith('Medoid'):
                in_sample_sharpe = params['original_sharpe']
                in_sample_trades = params['original_trades']
            else:
                # Calculate Sharpe ratio (annualized) for Best
                if in_sample_daily_pnl.std() > 0:
                    in_sample_sharpe = in_sample_daily_pnl.mean() / in_sample_daily_pnl.std() * np.sqrt(252)
                else:
                    in_sample_sharpe = 0
                # Count in-sample trades for Best
                in_sample_trades = in_sample[f'Position_Change_{name}'].sum()

            # Create row text
            row = f"{name:<10} | {short:>4}/{long:<5} | ${in_sample_cumulative_pnl:>10,.2f} | {avg_pos_size:>8.2f} | {in_sample_sharpe:>6.3f} | {in_sample_trades:>6}"

            # Print row
            print(row)

            # Store the data for later use
            performance_data.append({
                'Period': 'In-Sample',
                'Strategy': name,
                'Short_SMA': short,
                'Long_SMA': long,
                'PnL': in_sample_cumulative_pnl,
                'Avg_Position': avg_pos_size,
                'Sharpe': in_sample_sharpe,
                'Trades': in_sample_trades
            })

        print(separator)

        # OUT-OF-SAMPLE PERFORMANCE
        print("\nOUT-OF-SAMPLE PERFORMANCE:")
        print(separator)
        print(header)
        print(separator)

        for name, params in strategies.items():
            short = params['short_sma']
            long = params['long_sma']

            # Get out-of-sample data
            out_sample = data_for_evaluation.iloc[split_index:]

            # Calculate out-of-sample metrics
            out_sample_daily_pnl = out_sample[f'Daily_PnL_{name}']

            # Cumulative P&L just for the out-of-sample period
            out_sample_cumulative_pnl = out_sample[f'Daily_PnL_{name}'].sum()
            
            # Calculate average position size for out-of-sample
            avg_pos_size = out_sample[f'Position_Size_{name}'].mean()

            # Calculate Sharpe ratio (annualized)
            if out_sample_daily_pnl.std() > 0:
                out_sample_sharpe = out_sample_daily_pnl.mean() / out_sample_daily_pnl.std() * np.sqrt(252)
            else:
                out_sample_sharpe = 0

            # Count out-of-sample trades
            out_sample_trades = out_sample[f'Position_Change_{name}'].sum()

            # Create row text
            row = f"{name:<10} | {short:>4}/{long:<5} | ${out_sample_cumulative_pnl:>10,.2f} | {avg_pos_size:>8.2f} | {out_sample_sharpe:>6.3f} | {out_sample_trades:>6}"

            # Print row
            print(row)

            # Store the data
            performance_data.append({
                'Period': 'Out-of-Sample',
                'Strategy': name,
                'Short_SMA': short,
                'Long_SMA': long,
                'PnL': out_sample_cumulative_pnl,
                'Avg_Position': avg_pos_size,
                'Sharpe': out_sample_sharpe,
                'Trades': out_sample_trades
            })

        print(separator)

        # FULL PERIOD PERFORMANCE
        print("\nFULL PERIOD PERFORMANCE:")
        print(separator)
        print(header)
        print(separator)

        for name, params in strategies.items():
            short = params['short_sma']
            long = params['long_sma']

            # Calculate full period metrics
            full_daily_pnl = data_for_evaluation[f'Daily_PnL_{name}']
            full_cumulative_pnl = full_daily_pnl.sum()
            
            # Calculate average position size for full period
            avg_pos_size = data_for_evaluation[f'Position_Size_{name}'].mean()
            max_pos_size = data_for_evaluation[f'Position_Size_{name}'].max()

            # Calculate Sharpe ratio (annualized)
            if full_daily_pnl.std() > 0:
                full_sharpe = full_daily_pnl.mean() / full_daily_pnl.std() * np.sqrt(252)
            else:
                full_sharpe = 0

            # Count full period trades
            full_trades = data_for_evaluation[f'Position_Change_{name}'].sum()

            # Create row text
            row = f"{name:<10} | {short:>4}/{long:<5} | ${full_cumulative_pnl:>10,.2f} | {avg_pos_size:>8.2f} | {full_sharpe:>6.3f} | {full_trades:>6}"

            # Print row
            print(row)

            # Store the data
            performance_data.append({
                'Period': 'Full',
                'Strategy': name,
                'Short_SMA': short,
                'Long_SMA': long,
                'PnL': full_cumulative_pnl,
                'Avg_Position': avg_pos_size,
                'Max_Position': max_pos_size,
                'Sharpe': full_sharpe,
                'Trades': full_trades
            })
            
            # Additional metrics for the best strategy
            if name == 'Best':
                print(f"\nAdditional metrics for Best strategy:")
                print(f"Maximum position size: {max_pos_size:.2f} contracts")
                print(f"Average ATR value: {data_for_evaluation['ATR_Best'].mean():.4f}")
                
                # Calculate drawdown
                peak = data_for_evaluation[f'Cumulative_PnL_{name}'].cummax()
                drawdown = data_for_evaluation[f'Cumulative_PnL_{name}'] - peak
                max_drawdown = drawdown.min()
                
                print(f"Maximum drawdown: ${max_drawdown:.2f}")
                
                # Calculate win rate
                daily_win_rate = (data_for_evaluation[f'Daily_PnL_{name}'] > 0).mean() * 100
                print(f"Daily win rate: {daily_win_rate:.2f}%")

        print(separator)



        return data_for_evaluation

    def bimonthly_out_of_sample_comparison(data, best_short_sma, best_long_sma, top_medoids, min_sharpe=0.2, 
                                    big_point_value=big_point_value, slippage=slippage,
                                    capital=capital, atr_period=atr_period):
        """
        Compare bimonthly (2-month) performance between the best Sharpe strategy and a portfolio of top medoids
        using ATR-based position sizing.
        
        Parameters:
        data: DataFrame with market data
        best_short_sma: int - The short SMA period for the best Sharpe strategy
        best_long_sma: int - The long SMA period for the best Sharpe strategy
        top_medoids: list - List of top medoids, each as (short_sma, long_sma, sharpe, trades)
        min_sharpe: float - Minimum Sharpe ratio threshold for medoids to be included
        big_point_value: float - Big point value for the futures contract
        dynamic_slippage: float - Slippage value in price units
        capital: float - Capital allocation for position sizing
        atr_period: int - Period for ATR calculation
        """
        print(f"\n----- BIMONTHLY OUT-OF-SAMPLE COMPARISON -----")
        print(f"Best Sharpe: ({best_short_sma}/{best_long_sma})")
        print(f"Using ATR-based position sizing (Capital: ${capital:,}, ATR Period: {atr_period})")
        
        # Handle the case where top_medoids is None
        if top_medoids is None:
            print("No medoids provided. Comparison cannot be performed.")
            return None
        
        # Check if top_medoids is a single tuple (rather than a list of tuples)
        if isinstance(top_medoids, tuple) and len(top_medoids) >= 3:
            # Convert single tuple to a list with one tuple
            medoids_list = [top_medoids]
        else:
            # Assume it's already a list
            medoids_list = top_medoids
        
        # Take at most 3 medoids and filter by minimum Sharpe
        filtered_medoids = []
        for m in medoids_list[:3]:
            # Check if we can access the required elements (support numpy arrays, tuples, and lists)
            try:
                # Extract Sharpe ratio and check if it meets the threshold
                short_sma = m[0]
                long_sma = m[1]
                sharpe = float(m[2])  # Convert to float to handle numpy types
                trades = m[3]
                
                if sharpe >= min_sharpe:
                    filtered_medoids.append(m)
            except (IndexError, TypeError) as e:
                print(f"Error processing medoid: {e}")
        
        if not filtered_medoids:
            print(f"No medoids have a Sharpe ratio >= {min_sharpe}. Comparison cannot be performed.")
            return None
        
        print(f"Creating portfolio of {len(filtered_medoids)} medoids with Sharpe ratio >= {min_sharpe}:")
        for i, medoid in enumerate(filtered_medoids, 1):
            print(f"Medoid {i}: ({int(medoid[0])}/{int(medoid[1])}) - Sharpe: {float(medoid[2]):.4f}")
        
        # Download historical data if not already provided
        if data is None:
            # Load data from local file
            print(f"Loading {TICKER} data from local files...")
            data_file = find_futures_file(SYMBOL, DATA_DIR)
            if not data_file:
                print(f"Error: No data file found for {TICKER} in {DATA_DIR}")
                exit(1)
            
            print(f"Found data file: {os.path.basename(data_file)}")
            print(f"File size: {os.path.getsize(data_file)} bytes")

            # Load the data from local file
            all_data = read_ts.read_ts_ohlcv_dat(data_file)
            data_obj = all_data[0]
            ohlc_data = data_obj.data.copy()

            # Convert the OHLCV data to the format expected by the strategy
            data = ohlc_data.rename(columns={
                'datetime': 'Date',
                'open': 'Open',
                'high': 'High',
                'low': 'Low',
                'close': 'Close',
                'volume': 'Volume'
            })
            data.set_index('Date', inplace=True)
            
            # Filter data to match the date range if specified in input.py
            if START_DATE and END_DATE:
                data = data[(data.index >= pd.to_datetime(START_DATE)) & (data.index <= pd.to_datetime(END_DATE))]
                print(f"Filtered data to date range: {START_DATE} to {END_DATE}")
                
            # Simplify the data structure if needed
            if hasattr(data.columns, 'get_level_values'):
                data.columns = data.columns.get_level_values(0)
        
        # Create strategies
        strategies = {
            'Best': {'short_sma': best_short_sma, 'long_sma': best_long_sma}
        }
        
        # Add filtered medoids
        for i, medoid in enumerate(filtered_medoids, 1):
            strategies[f'Medoid_{i}'] = {'short_sma': int(medoid[0]), 'long_sma': int(medoid[1])}
        
        # Apply each strategy to the data
        for name, params in strategies.items():
            strategy = SMAStrategy(
                short_sma=params['short_sma'],
                long_sma=params['long_sma'],
                big_point_value=big_point_value,
                slippage=slippage,
                capital=capital,
                atr_period=atr_period
            )
            
            # Apply the strategy
            data = strategy.apply_strategy(
                data.copy(),
                strategy_name=name
            )
        
        # Get the out-of-sample split date
        split_index = int(len(data) * TRAIN_TEST_SPLIT)
        split_date = data.index[split_index]
        print(f"Out-of-sample period starts on: {split_date.strftime('%Y-%m-%d')}")
        
        # Get out-of-sample data
        oos_data = data.iloc[split_index:].copy()
        
        # Add a year and bimonthly period columns for grouping (each year has 6 bimonthly periods)
        oos_data['year'] = oos_data.index.year.astype(int)
        oos_data['bimonthly'] = ((oos_data.index.month - 1) // 2 + 1).astype(int)
        
        # Create simplified period labels with just the start month (YYYY-MM)
        oos_data['period_label'] = oos_data.apply(
            lambda row: f"{int(row['year'])}-{int((row['bimonthly'] - 1) * 2 + 1):02d}",
            axis=1
        )
        
        # Create a DataFrame to store bimonthly Sharpe ratios
        bimonthly_sharpe = []
        
        # Group by year and bimonthly period, calculate Sharpe ratio for each period
        for period_label, group in oos_data.groupby('period_label'):
            # Skip periods with too few trading days
            if len(group) < 10:
                continue
                
            # Create a bimonthly result entry
            year, start_month = period_label.split('-')
            year = int(year)
            start_month = int(start_month)
            
            bimonthly_result = {
                'period_label': period_label,
                'date': pd.Timestamp(year=year, month=start_month, day=15),  # Middle of first month in period
                'trading_days': len(group),
            }
            
            # Calculate Sharpe ratio for each strategy in this period
            for name in strategies.keys():
                # Get returns for this strategy in this period
                returns = group[f'Daily_PnL_{name}']
                
                # Calculate Sharpe ratio (annualized)
                if len(returns) > 1 and returns.std() > 0:
                    sharpe = returns.mean() / returns.std() * np.sqrt(252)
                else:
                    sharpe = 0
                    
                bimonthly_result[f'{name}_sharpe'] = sharpe
                bimonthly_result[f'{name}_return'] = returns.sum()  # Total P&L for the period
            
            # Calculate the normalized average of medoid Sharpe ratios
            medoid_sharpes = [bimonthly_result[f'Medoid_{i}_sharpe'] for i in range(1, len(filtered_medoids) + 1)]
            bimonthly_result['Avg_Medoid_sharpe'] = sum(medoid_sharpes) / len(filtered_medoids)
            
            # Calculate the normalized average of medoid returns
            medoid_returns = [bimonthly_result[f'Medoid_{i}_return'] for i in range(1, len(filtered_medoids) + 1)]
            bimonthly_result['Avg_Medoid_return'] = sum(medoid_returns) / len(filtered_medoids)
            
            bimonthly_sharpe.append(bimonthly_result)
        
        # Convert to DataFrame
        bimonthly_sharpe_df = pd.DataFrame(bimonthly_sharpe)
        
        # Sort the DataFrame by date for proper chronological display
        if not bimonthly_sharpe_df.empty:
            bimonthly_sharpe_df = bimonthly_sharpe_df.sort_values('date')
        else:
            print(f"WARNING: No bimonthly periods found for {SYMBOL}. Cannot create chart.")
            return None
        
        # Add rounded values to dataframe for calculations
        bimonthly_sharpe_df['Best_sharpe_rounded'] = np.round(bimonthly_sharpe_df['Best_sharpe'], 2)
        bimonthly_sharpe_df['Avg_Medoid_sharpe_rounded'] = np.round(bimonthly_sharpe_df['Avg_Medoid_sharpe'], 2)
        
        # Print detailed comparison of Sharpe ratios
        print("\nDetailed Sharpe ratio comparison by period:")
        print(f"{'Period':<12} | {'Best Sharpe':>12} | {'Cluster Portfolio':>16} | {'Difference':>12} | {'Portfolio Wins':<14}")
        print("-" * 80)
        
        for idx, row in bimonthly_sharpe_df.iterrows():
            period = row['period_label']
            best_sharpe = row['Best_sharpe']
            avg_medoid_sharpe = row['Avg_Medoid_sharpe']
            best_rounded = row['Best_sharpe_rounded']
            avg_medoid_rounded = row['Avg_Medoid_sharpe_rounded']
            
            diff = avg_medoid_sharpe - best_sharpe
            portfolio_wins = avg_medoid_sharpe > best_sharpe
            
            print(f"{period:<12} | {best_sharpe:12.6f} | {avg_medoid_sharpe:16.6f} | {diff:12.6f} | {portfolio_wins!s:<14}")
        
        # Calculate win rate using rounded values
        total_periods = len(bimonthly_sharpe_df)
        rounded_wins = sum(bimonthly_sharpe_df['Avg_Medoid_sharpe_rounded'] > bimonthly_sharpe_df['Best_sharpe_rounded'])
        rounded_win_percentage = (rounded_wins / total_periods) * 100 if total_periods > 0 else 0
        
        print(f"\nBimonthly Win Rate (Portfolio vs Best): {rounded_win_percentage:.2f}% ({rounded_wins}/{total_periods} periods)")
        
        # Create a bar plot to compare bimonthly Sharpe ratios
        plt.figure(figsize=(14, 8))
        
        # Set up x-axis dates
        x = np.arange(len(bimonthly_sharpe_df))
        width = 0.35  # Width of the bars
        
        # Create bars
        plt.bar(x - width/2, bimonthly_sharpe_df['Best_sharpe'], width, 
            label=f'Best Sharpe ({best_short_sma}/{best_long_sma})', color='blue')
        plt.bar(x + width/2, bimonthly_sharpe_df['Avg_Medoid_sharpe'], width, 
            label=f'Cluster Portfolio ({len(filtered_medoids)} strategies)', color='green')
        
        # Add a horizontal line at Sharpe = 0
        plt.axhline(y=0, color='black', linestyle='-', alpha=0.3)
        
        # Create medoid description for the title
        medoid_desc = ", ".join([f"({int(m[0])}/{int(m[1])})" for m in filtered_medoids])
        
        # Customize the plot - using rounded win percentage instead of raw
        plt.title(f'{SYMBOL} Hierarchical Bimonthly Sharpe Ratio Comparison (Out-of-Sample Period)\n' + 
                f'Medoid Portfolio [{medoid_desc}] outperformed {rounded_win_percentage:.2f}% of the time', 
                fontsize=14)
        plt.xlabel('Bimonthly Period (Start Month)', fontsize=12)
        plt.ylabel('Sharpe Ratio (Annualized)', fontsize=12)
        
        # Simplified x-tick labels with just the period start month
        plt.xticks(x, bimonthly_sharpe_df['period_label'], rotation=45)
        
        plt.grid(axis='y', linestyle='--', alpha=0.7)
        
        # Create legend with both strategies - moved to bottom to avoid overlap with title
        plt.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=2,
                frameon=True, fancybox=True, framealpha=0.9, fontsize=10)
        
        # Add a text box with rounded win percentage instead of raw - moved to right side
        plt.annotate(f'Medoid Portfolio Win Rate: {rounded_win_percentage:.2f}%\n'
                    f'({rounded_wins} out of {total_periods} periods)\n'
                    f'Portfolio: {len(filtered_medoids)} medoids with Sharpe ≥ {min_sharpe}\n'
                    f'ATR-Based Position Sizing (${capital:,}, {atr_period} days)',
                    xy=(0.7, 0.95), xycoords='axes fraction',
                    bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="gray", alpha=0.8),
                    fontsize=12)
        
        plt.tight_layout(rect=[0, 0, 1, 0.95])  # Add extra space at the bottom for the legend
        save_plot(f'{SYMBOL}_Hierarchical_Bimonthly_Comparison.png')
        
        
        # Save results to Excel
        try:
            excel_data = {
                'hierarchical_win_rate': rounded_win_percentage,
                'hierarchical_medoids': filtered_medoids
            }
            update_excel_results(excel_data)
        except Exception as e:
            print(f"Warning: Could not update Excel file: {e}")

        # Return the bimonthly Sharpe ratio data
        return bimonthly_sharpe_df

    def create_hierarchical_cluster_visualization(X, medoids, top_medoids, max_sharpe_point, labels):
        """
        Create visualization of hierarchical clustering results with scatter plots
        Only includes points from valid clusters that meet min_trades and min_elements_per_cluster requirements

        Parameters:
        X: numpy array - Data points that already meet min_trades requirement
        medoids: list of tuples - Cluster medoids (these are already from valid clusters only)
        top_medoids: list of tuples - Top medoids by Sharpe ratio
        max_sharpe_point: tuple - Point with maximum Sharpe ratio
        labels: numpy array - Cluster labels for the points in X
        """
        print("Creating hierarchical cluster visualization...")
        
        # Determine which clusters are represented in the medoids
        # This is our definition of valid clusters - ones that have medoids
        valid_cluster_ids = set([labels[np.where((X[:, 0] == medoid[0]) & (X[:, 1] == medoid[1]))[0][0]] for medoid in medoids])
        print(f"Using {len(valid_cluster_ids)} valid clusters with medoids")
        
        # Filter X to only include points from valid clusters
        valid_indices = np.array([label in valid_cluster_ids for label in labels])
        X_valid = X[valid_indices]
        labels_valid = labels[valid_indices]
        
        print(f"Filtering to {len(X_valid)} points from valid clusters")
        
        # Create a DataFrame for the valid points to use in visualization
        filtered_df = pd.DataFrame(X_valid, columns=['short_SMA', 'long_SMA', 'sharpe_ratio', 'trades'])

        # Create a pivot table for the heatmap from the filtered data
        heatmap_data = filtered_df.pivot_table(
            index='long_SMA',
            columns='short_SMA',
            values='sharpe_ratio',
            fill_value=np.nan  # Use NaN for empty cells
        )
        
        # Create a new figure with 2 subplots (1x2)
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 8))

        # Get the number of unique valid clusters
        n_clusters = len(valid_cluster_ids)

        # Create a colormap with distinct colors for each cluster
        colors = plt.cm.tab10(np.linspace(0, 1, n_clusters))
        
        # Create a mapping from cluster ID to color index
        cluster_to_color = {cluster_id: i for i, cluster_id in enumerate(valid_cluster_ids)}

        # Plot 1: Short SMA vs Long SMA (colored by cluster)
        for cluster_id in valid_cluster_ids:
            # Get indices of points in this valid cluster
            cluster_points = X_valid[labels_valid == cluster_id]
            if len(cluster_points) == 0:
                continue

            # Get the color for this cluster
            color_idx = cluster_to_color[cluster_id]
            color = colors[color_idx % len(colors)]  # Use modulo to handle if we have more clusters than colors
            
            # Plot the points
            ax1.scatter(
                cluster_points[:, 0],  # Short SMA
                cluster_points[:, 1],  # Long SMA
                s=50, c=[color], alpha=0.7
            )

        # Plot the medoids with black edges
        for i, medoid in enumerate(medoids):
            ax1.scatter(
                medoid[0],  # Short SMA
                medoid[1],  # Long SMA
                s=150, c='white', alpha=1,
                edgecolors='black', linewidths=2,
                marker='o'
            )

        # Plot the top medoids with star markers
        for i, medoid in enumerate(top_medoids):
            ax1.scatter(
                medoid[0],  # Short SMA
                medoid[1],  # Long SMA
                s=200, c='gold', alpha=1,
                edgecolors='black', linewidths=1.5,
                marker='*'
            )
            
        # Plot the max Sharpe point with a diamond marker
        ax1.scatter(
            max_sharpe_point[0],  # Short SMA
            max_sharpe_point[1],  # Long SMA
            s=250, c='red', alpha=1,
            edgecolors='black', linewidths=2,
            marker='D'
        )

        # Add labels and title
        ax1.set_xlabel('Short SMA (days)', fontsize=12)
        ax1.set_ylabel('Long SMA (days)', fontsize=12)
        ax1.set_title('Hierarchical Clusters: Short SMA vs Long SMA', fontsize=14)
        ax1.grid(True, linestyle='--', alpha=0.7)

        # Add a diagonal line where short_SMA = long_SMA
        max_val = max(X_valid[:, 0].max(), X_valid[:, 1].max()) * 1.1
        ax1.plot([0, max_val], [0, max_val], 'k--', alpha=0.3)

        # Create a simplified legend with just the key elements
        custom_handles = [
            Line2D([0], [0], marker='o', color='w', markerfacecolor='w',
                markeredgecolor='black', markeredgewidth=2, markersize=10, label='Cluster Medoid'),
            Line2D([0], [0], marker='*', color='w', markerfacecolor='gold',
                markeredgecolor='black', markersize=10, label='Top Medoid'),
            Line2D([0], [0], marker='D', color='w', markerfacecolor='red',
                markeredgecolor='black', markersize=10, label='Best Sharpe Point')
        ]
        
        # Add custom handles to legend
        ax1.legend(handles=custom_handles, loc='upper right')
        
        # Plot 2: Create a mask for the heatmap where short_SMA >= long_SMA or values are NaN
        mask = np.zeros_like(heatmap_data, dtype=bool)
        for i, long_sma in enumerate(heatmap_data.index):
            for j, short_sma in enumerate(heatmap_data.columns):
                if short_sma >= long_sma or np.isnan(heatmap_data.iloc[i, j]):
                    mask[i, j] = True

        # Plot the heatmap on ax2
        sns.heatmap(
            heatmap_data,
            ax=ax2,
            mask=mask,
            cmap='coolwarm',  # Blue to red colormap
            annot=False,      # Don't annotate each cell with its value
            fmt='.4f',
            linewidths=0,
            cbar_kws={'label': 'Sharpe Ratio'}
        )
        
        # Invert the y-axis so smaller long_SMA values are at the top
        ax2.invert_yaxis()
        
        # Plot the medoids (Black Squares)
        for medoid in medoids:
            try:
                x_pos = np.where(heatmap_data.columns == medoid[0])[0][0] + 0.5
                y_pos = np.where(heatmap_data.index == medoid[1])[0][0] + 0.5
                ax2.scatter(x_pos, y_pos, marker='s', color='black', s=75, zorder=4)
            except IndexError:
                print(f"Warning: Medoid at ({medoid[0]}, {medoid[1]}) not found in heatmap coordinates")

        # Plot top medoids (Purple Diamonds)
        for medoid in top_medoids:
            try:
                x_pos = np.where(heatmap_data.columns == medoid[0])[0][0] + 0.5
                y_pos = np.where(heatmap_data.index == medoid[1])[0][0] + 0.5
                ax2.scatter(x_pos, y_pos, marker='D', color='purple', s=100, zorder=5)
            except IndexError:
                print(f"Warning: Top medoid at ({medoid[0]}, {medoid[1]}) not found in heatmap coordinates")

        # Plot max Sharpe point (Green Star)
        try:
            best_x_pos = np.where(heatmap_data.columns == max_sharpe_point[0])[0][0] + 0.5
            best_y_pos = np.where(heatmap_data.index == max_sharpe_point[1])[0][0] + 0.5
            ax2.scatter(best_x_pos, best_y_pos, marker='*', color='lime', s=200,
                    edgecolor='black', zorder=5)
        except IndexError:
            print(f"Warning: Max Sharpe point at ({max_sharpe_point[0]}, {max_sharpe_point[1]}) not found in heatmap coordinates")

        # Create custom legend for the heatmap
        max_sharpe_handle = mlines.Line2D([], [], color='lime', marker='*', linestyle='None',
                                    markersize=15, markeredgecolor='black', label='Max Sharpe')
        medoid_handle = mlines.Line2D([], [], color='black', marker='s', linestyle='None',
                                markersize=10, label='Clusters')
        top_medoid_handle = mlines.Line2D([], [], color='purple', marker='D', linestyle='None',
                                    markersize=10, label='Top 5 Clusters')

        # Add a legend to the second plot
        ax2.legend(handles=[max_sharpe_handle, medoid_handle, top_medoid_handle],
                loc='best')

        # Add labels and title
        ax2.set_title(f'{SYMBOL} Hierarchical Clustering Analysis (Sharpe Ratio)', fontsize=14)
        ax2.set_xlabel('Short SMA (days)', fontsize=12)
        ax2.set_ylabel('Long SMA (days)', fontsize=12)

        # Rotate tick labels
        ax2.set_xticklabels(ax2.get_xticklabels(), rotation=45, ha='right')
        ax2.set_yticklabels(ax2.get_yticklabels(), rotation=0)

        # Adjust layout
        plt.tight_layout()

        # Save the plot
        save_plot(f'{SYMBOL}_Hierarchical_Clusters_Visualization.png')

    def analyze_full_oos_performance(best_short_sma, best_long_sma, top_medoids,
                               big_point_value=1, slippage=0, capital=1000000, atr_period=14):
        """
        Plot a comparison between the best strategy and a portfolio of top 3 clusters for hierarchical clustering.
        The portfolio is the average of the top 3 cluster strategies' P&L, matching the KMeans approach.
        """
        print(f"\n----- FULL OUT-OF-SAMPLE PERFORMANCE ANALYSIS (Hierarchical) -----")
        print(f"Using Best Strategy: Short SMA: {best_short_sma}, Long SMA: {best_long_sma}")
        print(f"Creating portfolio from top 3 clusters")

        # Load data from local file
        print(f"Loading {TICKER} data from local files...")
        data_file = find_futures_file(SYMBOL, DATA_DIR)
        if not data_file:
            print(f"Error: No data file found for {TICKER} in {DATA_DIR}")
            return None
        print(f"Found data file: {os.path.basename(data_file)}")
        print(f"File size: {os.path.getsize(data_file)} bytes")
        all_data = read_ts.read_ts_ohlcv_dat(data_file)
        data_obj = all_data[0]
        ohlc_data = data_obj.data.copy()
        data = ohlc_data.rename(columns={
            'datetime': 'Date',
            'open': 'Open',
            'high': 'High',
            'low': 'Low',
            'close': 'Close',
            'volume': 'Volume'
        })
        data.set_index('Date', inplace=True)

        # Add a warm-up period before the start date
        original_start_idx = None
        if START_DATE and END_DATE:
            warm_up_days = max(best_short_sma, best_long_sma) * 3
            start_date = pd.to_datetime(START_DATE)
            end_date = pd.to_datetime(END_DATE)
            adjusted_start = start_date - pd.Timedelta(days=warm_up_days)
            extended_data = data[(data.index >= adjusted_start) & (data.index <= end_date)]
            if not extended_data.empty:
                original_start_idx = extended_data.index.get_indexer([start_date], method='nearest')[0]
            print(f"Added {warm_up_days} days warm-up period before {START_DATE}")
            data = extended_data

        # Create strategies dict
        strategies = {
            'Best': {'short_sma': best_short_sma, 'long_sma': best_long_sma}
        }
        cluster_names = []
        if top_medoids:
            for i, medoid in enumerate(top_medoids[:3], 1):
                name = f'Cluster {i}'
                strategies[name] = {
                    'short_sma': int(medoid[0]),
                    'long_sma': int(medoid[1])
                }
                cluster_names.append(name)

        # Apply strategies
        for name, params in strategies.items():
            data[f'SMA_Short_{name}'] = data['Close'].rolling(window=params['short_sma'], center=True).mean()
            data[f'SMA_Long_{name}'] = data['Close'].rolling(window=params['long_sma'], center=True).mean()
            sma_strategy = SMAStrategy(
                short_sma=params['short_sma'],
                long_sma=params['long_sma'],
                big_point_value=big_point_value,
                slippage=slippage,
                capital=capital,
                atr_period=atr_period
            )
            data = sma_strategy.apply_strategy(
                data.copy(),
                strategy_name=name
            )

        # Trim data to the original date range if we added warm-up period
        if original_start_idx is not None:
            data_for_evaluation = data.iloc[original_start_idx:]
            print(f"Trimmed warm-up period. Original data length: {len(data)}, Evaluation data length: {len(data_for_evaluation)}")
        else:
            data_for_evaluation = data

        # Calculate split index for in-sample/out-of-sample
        split_index = int(len(data_for_evaluation) * TRAIN_TEST_SPLIT)
        split_date = data_for_evaluation.index[split_index]

        # Build portfolio daily PnL as average of top 3 clusters
        if cluster_names:
            portfolio_daily_pnl = pd.Series(0, index=data_for_evaluation.index)
            valid_strategies = 0
            for name in cluster_names:
                col = f'Daily_PnL_{name}'
                if col in data_for_evaluation.columns:
                    portfolio_daily_pnl += data_for_evaluation[col]
                    valid_strategies += 1
            if valid_strategies > 0:
                portfolio_daily_pnl = portfolio_daily_pnl / valid_strategies
            else:
                portfolio_daily_pnl = data_for_evaluation['Daily_PnL_Best'].copy()
        else:
            portfolio_daily_pnl = data_for_evaluation['Daily_PnL_Best'].copy()
        portfolio_cumulative_pnl = portfolio_daily_pnl.cumsum()

        # Compute Sharpe ratios and build labels
        split_index = int(len(data_for_evaluation) * TRAIN_TEST_SPLIT)
        best_daily = data_for_evaluation['Daily_PnL_Best']
        best_cum = best_daily.cumsum()
        best_in = best_daily.iloc[:split_index]
        best_out = best_daily.iloc[split_index:]
        best_sharpe_in = best_in.mean() / best_in.std() * np.sqrt(252) if best_in.std() > 0 else 0
        best_sharpe_out = best_out.mean() / best_out.std() * np.sqrt(252) if best_out.std() > 0 else 0
        best_label = f'Best ({best_short_sma}/{best_long_sma}) [IS Sharpe: {best_sharpe_in:.2f}, OOS Sharpe: {best_sharpe_out:.2f}]'

        port_in = portfolio_daily_pnl.iloc[:split_index]
        port_out = portfolio_daily_pnl.iloc[split_index:]
        port_sharpe_in = port_in.mean() / port_in.std() * np.sqrt(252) if port_in.std() > 0 else 0
        port_sharpe_out = port_out.mean() / port_out.std() * np.sqrt(252) if port_out.std() > 0 else 0
        medoid_params = []
        for i, medoid in enumerate(top_medoids[:3], 1):
            medoid_params.append(f"{int(medoid[0])}/{int(medoid[1])}")
        port_label = f'Portfolio ({", ".join(medoid_params)}) [IS Sharpe: {port_sharpe_in:.2f}, OOS Sharpe: {port_sharpe_out:.2f}]'

        # Plot
        plt.figure(figsize=(14, 8))
        plt.plot(data_for_evaluation.index, best_cum, label=best_label, color='blue')
        plt.plot(data_for_evaluation.index, portfolio_cumulative_pnl, label=port_label, color='green')
        plt.plot(data_for_evaluation.index[split_index:], best_cum.iloc[split_index:], color='blue', linewidth=2.5, alpha=0.7)
        plt.plot(data_for_evaluation.index[split_index:], portfolio_cumulative_pnl.iloc[split_index:], color='green', linewidth=2.5, alpha=0.7)
        plt.axvline(x=split_date, color='black', linestyle='--',
                    label=f'Train/Test Split ({int(TRAIN_TEST_SPLIT * 100)}%/{int((1 - TRAIN_TEST_SPLIT) * 100)}%)')
        plt.axhline(y=0.0, color='gray', linestyle='-', alpha=0.5, label='Break-even')
        plt.legend(loc='upper left')
        plt.title(f'{SYMBOL} Hierarchical Strategy Comparison: Best vs Portfolio of Clusters', fontsize=14)
        plt.ylabel('P&L ($)', fontsize=12)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        save_plot(f'{SYMBOL}_Hierarchical_Full_OOS_Performance_Analysis.png')
        print(f"Full out-of-sample performance analysis plot saved to '{SYMBOL}_Hierarchical_Full_OOS_Performance_Analysis.png'")
        
        # Save OOS Sharpe ratios to Excel
        try:
            excel_data = {
                'portfolio_oos_sharpe_hierarchical': port_sharpe_out
            }
            update_excel_results(excel_data)
        except Exception as e:
            print(f"Warning: Could not update Excel file for OOS performance: {e}")

        return data_for_evaluation

    # Main execution block
    if __name__ == "__main__":
        # Set matplotlib backend explicitly
        matplotlib.use('Agg')  # Use non-interactive backend for headless environments
        
        print("Starting SMA strategy analysis with hierarchical clustering...")
        
        # Run the basic analysis first
        data, best_short, best_long, best_sharpe, best_trades = analyze_sma_results()

        if data is None:
            print("Error: Failed to load or analyze SMA results data.")
            exit(1)

        print("\nProceeding with hierarchical cluster analysis...")
        
        # Run the hierarchical cluster analysis to get medoids
        X_filtered, medoids, top_medoids, max_sharpe_point, labels = hierarchical_cluster_analysis()
        
        if X_filtered is None or medoids is None:
            print("Error: Hierarchical cluster analysis failed.")
            exit(1)

        print("\nPlotting strategy performance...")
        
        # Plot strategy performance with the best parameters AND top medoids using ATR-based position sizing
        market_data = plot_strategy_performance(
            best_short, best_long, top_medoids, 
            big_point_value=big_point_value, 
            slippage=slippage,
            capital=capital,
            atr_period=atr_period
        )
        
        # Run the bimonthly out-of-sample comparison between best Sharpe and top medoids
        if top_medoids and len(top_medoids) > 0:
            print("\nPerforming bimonthly out-of-sample comparison with hierarchical clustering...")
            bimonthly_sharpe_df = bimonthly_out_of_sample_comparison(
                market_data, 
                best_short, 
                best_long, 
                top_medoids,  # Pass the entire list of top medoids
                big_point_value=big_point_value,
                slippage=slippage,
                capital=capital,
                atr_period=atr_period
            )
        else:
            print("No top medoids found. Cannot run bimonthly comparison.")
            
        # After bimonthly_out_of_sample_comparison in main, call the new function if top_medoids are available
        if top_medoids and len(top_medoids) > 0:
            print("\nPerforming full out-of-sample performance analysis (Hierarchical)...")
            analyze_full_oos_performance(
                best_short,
                best_long,
                top_medoids,
                big_point_value=big_point_value,
                slippage=slippage,
                capital=capital,
                atr_period=atr_period
            )
        else:
            print("No top medoids found. Cannot run full OOS analysis (Hierarchical).")
            
        print("\nAnalysis complete! All plots and result files have been saved to the output directory.")
        print(f"Output directory: {output_dir}")

# Call the main function to execute the script
if __name__ == "__main__":
    main()